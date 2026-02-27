import glob
import os
import openpyxl

base_dir = r"c:\Users\arauj\OneDrive\Área de Trabalho\testes\brunozzi\automacao_etl\dados"
target_file = os.path.join(base_dir, "RLH021_jhonatan.brunozzi_344-0.xlsx")

if not os.path.exists(target_file):
    print(f"File not found: {target_file}")
    exit()

print(f"Checking file: {target_file}")
try:
    wb = openpyxl.load_workbook(target_file, read_only=True)
    
    for sheet_name in wb.sheetnames:
        print(f"Checking sheet: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        if rows:
            cols = [str(c) for c in rows[0]]
            print(f"Columns in '{sheet_name}': {cols}")
            if "Vel_Desl_Vazio_media" in cols:
                print(f"FOUND 'Vel_Desl_Vazio_media' in {sheet_name}!")
            else:
                print(f"NOT FOUND 'Vel_Desl_Vazio_media' in {sheet_name}!")
            
            if "Vel_Desl_Carregado_media" in cols:
                print(f"FOUND 'Vel_Desl_Carregado_media' in {sheet_name}!")
            else:
                print(f"NOT FOUND 'Vel_Desl_Carregado_media' in {sheet_name}!")
            
            # Check for Basculamento
            basc_cols = [c for c in cols if "basculamento" in c.lower()]
            if basc_cols:
                print(f"FOUND Basculamento cols in {sheet_name}: {basc_cols}")
            else:
                print(f"NOT FOUND Basculamento cols in {sheet_name}!")
        else:
            print(f"Sheet {sheet_name} is empty.")

except Exception as e:
    print(f"Error: {e}")
