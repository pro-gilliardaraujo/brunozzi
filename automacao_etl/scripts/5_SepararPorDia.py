import os
import sys
import pandas as pd
import glob
import warnings
import re
import json
from datetime import datetime
from openpyxl.utils import get_column_letter

# Suprimir avisos específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- CONFIGURAÇÕES ---
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIRETORIO_DADOS = os.path.join(BASE_DIR, "dados")
DIRETORIO_SAIDA = os.path.join(DIRETORIO_DADOS, "separados")
DIRETORIO_XLSX = os.path.join(DIRETORIO_SAIDA, "xlsx")
DIRETORIO_JSON = os.path.join(DIRETORIO_SAIDA, "json")

def extrair_periodo_nome_arquivo(nome_arquivo):
    """
    Extrai datas de início e fim do nome do arquivo.
    Padrão esperado: ...DD-MM-YYYY_DD-MM-YYYY...
    """
    match = re.search(r"(\d{2}-\d{2}-\d{4})_(\d{2}-\d{2}-\d{4})", nome_arquivo)
    if match:
        d1_str, d2_str = match.groups()
        try:
            d1 = pd.to_datetime(d1_str, format="%d-%m-%Y").date()
            d2 = pd.to_datetime(d2_str, format="%d-%m-%Y").date()
            return d1, d2
        except:
            pass
    return None, None

def ajustar_largura_colunas(worksheet):
    """Ajusta a largura das colunas automaticamente."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def formatar_coluna_data(worksheet, df):
    """Aplica formatação de data DD/MM/YYYY nas colunas de data."""
    col_idx_map = {name: i+1 for i, name in enumerate(df.columns)}
    
    # Procura colunas que parecem ser data
    for col_name in df.columns:
        if "Data" in col_name or "Início" in col_name or "Fim" in col_name:
            col_idx = col_idx_map[col_name]
            col_letter = get_column_letter(col_idx)
            
            # Aplica formato na coluna inteira (exceto cabeçalho)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.number_format = 'DD/MM/YYYY'

def encontrar_ultimo_tratado():
    """Encontra o arquivo tratado mais recente na pasta dados."""
    if not os.path.exists(DIRETORIO_DADOS):
        print(f"ERRO: Diretório de dados não encontrado: {DIRETORIO_DADOS}")
        return None

    # Tenta padrões diferentes para garantir que encontre o arquivo
    padroes = ["*tratado*.xlsx", "*_tratado.xlsx"]
    arquivos = []
    
    for p in padroes:
        caminho_padrao = os.path.join(DIRETORIO_DADOS, p)
        encontrados = glob.glob(caminho_padrao)
        arquivos.extend(encontrados)
    
    # Remove duplicatas
    arquivos = list(set(arquivos))
    
    if not arquivos:
        return None
    
    # Ordena por data de modificação (mais recente primeiro)
    arquivos.sort(key=os.path.getmtime, reverse=True)
    return arquivos[0]

def normalizar_nome_pasta(txt):
    if not txt: return "outros"
    txt = str(txt).lower().strip()
    if txt == "colhedora de cana": return "colhedora"
    if txt == "trator transbordo": return "transbordo"
    return txt.replace(" ", "_").replace("/", "-")

def main():
    print("=== INICIANDO SEPARAÇÃO COMPLETA POR DIA ===")
    
    # 1. Encontrar o arquivo
    arquivo_input = encontrar_ultimo_tratado()
    
    if not arquivo_input:
        print(f"ERRO: Nenhum arquivo tratado encontrado em: {DIRETORIO_DADOS}")
        print("Execute o script 2_Tratamento.py primeiro.")
        sys.exit(1)
        
    print(f"Lendo arquivo base: {os.path.basename(arquivo_input)}")
    
    try:
        # 2. Ler todas as abas do arquivo Excel
        xl = pd.ExcelFile(arquivo_input, engine='openpyxl')
        sheet_names = xl.sheet_names
        print(f"Abas encontradas: {sheet_names}")
        
        # Carregar todas as abas em um dicionário de DataFrames
        dfs = {}
        for sheet in sheet_names:
            print(f"  Carregando aba: {sheet}")
            dfs[sheet] = pd.read_excel(arquivo_input, sheet_name=sheet, engine='openpyxl')
        
        # 3. Identificar dias únicos (usando a aba 'Tratado' ou 'Original' como referência)
        # Preferência por 'Tratado' pois já passou pelo filtro de datas do 2_Tratamento.py
        df_ref = None
        if 'Tratado' in dfs:
            df_ref = dfs['Tratado']
        elif 'Original' in dfs:
            df_ref = dfs['Original']
        else:
            # Se não tiver Tratado nem Original, pega a primeira
            df_ref = dfs[sheet_names[0]]
            
        # Encontrar coluna de data na referência
        col_data_ref = None
        for col in df_ref.columns:
            if str(col).lower().strip() == 'data':
                col_data_ref = col
                break
        
        # Se não achou 'Data', tenta criar a partir de 'Data Hora Local'
        if not col_data_ref:
            print("  Coluna 'Data' explícita não encontrada na referência. Tentando derivar de 'Data Hora Local'...")
            for col in df_ref.columns:
                if str(col).lower().strip() == 'data hora local':
                    # Criar coluna Data no df_ref
                    df_ref['Data'] = pd.to_datetime(df_ref[col], dayfirst=True, errors='coerce').dt.date
                    col_data_ref = 'Data'
                    break

        if not col_data_ref:
            # Tentar procurar em outras abas que sabemos que têm Data
            print("  'Data' não encontrada em Tratado/Original. Procurando em outras abas...")
            for sheet_name, df_temp in dfs.items():
                if "_Dia" in sheet_name or "Intervalos" in sheet_name:
                    for col in df_temp.columns:
                         if str(col).lower().strip() == 'data':
                            print(f"  Usando aba '{sheet_name}' como referência de datas.")
                            df_ref = df_temp
                            col_data_ref = col
                            break
                if col_data_ref:
                    break

        if not col_data_ref:
            print("ERRO: Coluna 'Data' não encontrada na aba de referência para identificar os dias.")
            sys.exit(1)
            
        # Converter para datetime e pegar dias únicos
        df_ref[col_data_ref] = pd.to_datetime(df_ref[col_data_ref], errors='coerce')
        datas_unicas = df_ref[col_data_ref].dropna().unique()
        
        if len(datas_unicas) == 0:
            print("ERRO: Nenhuma data válida encontrada na aba de referência.")
            sys.exit(1)

        # Filtrar datas com base no nome do arquivo (se disponível)
        dt_inicio_filtro, dt_fim_filtro = extrair_periodo_nome_arquivo(os.path.basename(arquivo_input))
        if dt_inicio_filtro and dt_fim_filtro:
            print(f"  Filtrando dias pelo período do arquivo: {dt_inicio_filtro} a {dt_fim_filtro}")
            datas_filtradas = []
            for d in datas_unicas:
                 # d é numpy.datetime64, converte para datetime.date
                 d_dt = pd.to_datetime(d).date()
                 if dt_inicio_filtro <= d_dt <= dt_fim_filtro:
                     datas_filtradas.append(d)
            
            if not datas_filtradas:
                 print(f"AVISO: Nenhuma data encontrada dentro do período {dt_inicio_filtro} a {dt_fim_filtro}.")
                 # Se filtrar tudo, talvez seja melhor não filtrar e avisar, ou sair. 
                 # O usuário quer remover dias fora. Se todos estão fora, gera vazio.
            
            datas_unicas = datas_filtradas
            
        print(f"Encontrados {len(datas_unicas)} dias únicos: {[pd.to_datetime(d).strftime('%d/%m') for d in datas_unicas]}")

        # 4. Criar diretórios de saída
        if not os.path.exists(DIRETORIO_SAIDA):
            os.makedirs(DIRETORIO_SAIDA)
        
        if not os.path.exists(DIRETORIO_XLSX):
            os.makedirs(DIRETORIO_XLSX)
            
        if not os.path.exists(DIRETORIO_JSON):
            os.makedirs(DIRETORIO_JSON)

        print(f"Diretórios de saída configurados em: {DIRETORIO_SAIDA}")
            
        # 5. Separar e Salvar
        arquivos_gerados = []
        
        for data_val in sorted(datas_unicas):
            ts = pd.to_datetime(data_val)
            data_str = ts.strftime("%d-%m-%Y")
            nome_arquivo = f"{data_str}.xlsx"
            caminho_saida = os.path.join(DIRETORIO_XLSX, nome_arquivo)
            
            print(f"\nGerando arquivos para: {data_str}")
            
            with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
                sheets_salvas = 0
                dados_dia_json = {}
                
                for sheet_name, df in dfs.items():
                    # Ignorar abas de Periodo e as abas brutas (Original/Tratado) no JSON/XLSX diário
                    if "Periodo" in sheet_name or sheet_name in ["Original", "Tratado"]:
                        # print(f"    (Ignorando aba {sheet_name} - agregação de período ou dados brutos)")
                        continue
                        
                    # Tentar filtrar por data
                    col_data_sheet = None
                    for col in df.columns:
                        if str(col).lower().strip() == 'data':
                            col_data_sheet = col
                            break
                    
                    # Se não tem 'Data', tenta 'Data Hora Local'
                    temp_col_created = False
                    if not col_data_sheet:
                        for col in df.columns:
                            if str(col).lower().strip() == 'data hora local':
                                # Criar coluna temporária para filtro
                                df['_temp_data_filter_'] = pd.to_datetime(df[col], dayfirst=True, errors='coerce').dt.date
                                col_data_sheet = '_temp_data_filter_'
                                temp_col_created = True
                                break
                    
                    df_to_save = None
                    
                    if col_data_sheet:
                        # Se tem coluna Data, filtra
                        if not temp_col_created:
                            # Garantir tipo datetime se for a coluna original
                            df[col_data_sheet] = pd.to_datetime(df[col_data_sheet], errors='coerce')
                        
                        # Filtrar (se for temp, já é date object; se for original, é datetime)
                        # ts é datetime. Precisa comparar com date se a coluna for date, ou datetime se datetime.
                        # Melhor normalizar tudo para .date() na comparação
                        
                        if temp_col_created:
                             df_filtered = df[df[col_data_sheet] == ts.date()].copy()
                             # Remove a coluna temporária
                             df_filtered = df_filtered.drop(columns=['_temp_data_filter_'])
                        else:
                             # Comparação direta de datetime (pode falhar se tiver hora). 
                             # Vamos converter para date para garantir
                             df_temp_compare = df[col_data_sheet].dt.date
                             df_filtered = df[df_temp_compare == ts.date()].copy()
                        
                        if not df_filtered.empty:
                            df_to_save = df_filtered
                        else:
                            # Se não tem dados para esse dia nesta aba, pular
                            continue
                    else:
                        # Se não tem coluna Data
                        # Pode ser uma aba de configuração ou resumo sem data.
                        # Como o usuário pediu "todas as planilhas existentes", e removemos Periodo,
                        # se sobrar algo sem data, copiamos integralmente?
                        # Risco: duplicar dados de outros dias.
                        # Decisão: Por segurança, se não tem Data e não é Periodo, avisamos e não salvamos, 
                        # a menos que seja algo conhecido como "Operadores" (tem data), "Equipamentos" (tem data).
                        # Top5 tem data. Intervalos tem data.
                        # Se não tiver data, provavelmente não deve ir para o arquivo diário específico.
                        # print(f"    (Ignorando aba {sheet_name} - sem coluna Data)")
                        continue
                    
                    if df_to_save is not None:
                        # Salvar aba no Excel (mantendo objetos datetime para formatação nativa do Excel)
                        df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheets_salvas += 1
                        
                        # Preparar DataFrame para JSON
                        df_json = df_to_save.copy()
                        
                        # 1. Formatar colunas de tempo (Início, Fim) para DD/MM/YYYY HH:MM:SS
                        for col in ["Início", "Fim"]:
                            if col in df_json.columns and pd.api.types.is_datetime64_any_dtype(df_json[col]):
                                df_json[col] = df_json[col].dt.strftime('%d/%m/%Y %H:%M:%S')

                        # 2. Remover colunas redundantes (Descrição do Equipamento, Data)
                        # O usuário solicitou remover pois já constam no nome do arquivo/aba
                        cols_to_remove = ["Descrição do Equipamento"]
                        if col_data_sheet:
                            cols_to_remove.append(col_data_sheet)
                        
                        cols_existing = [c for c in cols_to_remove if c in df_json.columns]
                        if cols_existing:
                            df_json = df_json.drop(columns=cols_existing)

                        # Formatar a coluna de data principal (caso ela NÃO tenha sido removida por algum motivo)
                        if col_data_sheet and col_data_sheet in df_json.columns:
                             if pd.api.types.is_datetime64_any_dtype(df_json[col_data_sheet]):
                                 df_json[col_data_sheet] = df_json[col_data_sheet].dt.strftime('%d/%m/%Y')

                        # Adicionar ao JSON (convertendo para dict serializável)
                        # Usar to_json e depois loads para garantir conversão correta de datas e NaNs (NaN vira null)
                        try:
                            json_str = df_json.to_json(orient='records', date_format='iso', default_handler=str)
                            dados_dia_json[sheet_name] = json.loads(json_str)
                        except Exception as e_json:
                            print(f"    AVISO: Falha ao serializar aba {sheet_name} para JSON: {e_json}")

                        # Formatação
                        worksheet = writer.sheets[sheet_name]
                        ajustar_largura_colunas(worksheet)
                        formatar_coluna_data(worksheet, df_to_save)
                
                if sheets_salvas > 0:
                    print(f"  -> Arquivo Excel salvo com {sheets_salvas} abas.")
                    arquivos_gerados.append(nome_arquivo)
                    
                    # --- GERAR JSONs ESPECÍFICOS POR TIPO DE FROTA E ORGANIZAR EM PASTAS ---
                    
                    # Função auxiliar para garantir o nome do diretório normalizado
                    def obter_nome_diretorio(nome):
                        return normalizar_texto_simples(nome)

                    # Função para normalizar texto para nomes de pasta/arquivo
                    def normalizar_texto_simples(txt):
                        if not txt: return "outros"
                        # Remove acentos e caracteres especiais
                        txt = str(txt).lower().strip()
                        # Mapeamento simples para garantir consistência com o Tratamento
                        if txt == "colhedora de cana": return "colhedora"
                        if txt == "trator transbordo": return "transbordo"
                        return txt.replace(" ", "_").replace("/", "-")

                    # Dicionário para agrupar dados por tipo de frota
                    # Estrutura: { "colhedora": { "Resumo_Dia": [...], "Operadores": [...], ... }, ... }
                    dados_por_tipo = {}

                    # Função auxiliar para adicionar dados ao agrupamento
                    def adicionar_ao_grupo(tipo, chave, dados):
                        nome_tipo = obter_nome_diretorio(tipo)
                        if nome_tipo not in dados_por_tipo:
                            dados_por_tipo[nome_tipo] = {}
                        dados_por_tipo[nome_tipo][chave] = dados

                    # Iterar sobre as abas carregadas para distribuir nos grupos
                    for key_aba, dados_lista in dados_dia_json.items():
                        
                        # 1. Abas de Resumo Diário (_Dia) -> Ex: COLHEDORA_Dia
                        if key_aba.endswith("_Dia"):
                            tipo_frota = key_aba.replace("_Dia", "")
                            
                            # Filtrar colunas principais para o Resumo
                            dados_filtrados = []
                            if isinstance(dados_lista, list):
                                for item in dados_lista:
                                    novo_item = {}
                                    # Chaves identificadoras
                                    if "Frota" in item: novo_item["Frota"] = item["Frota"]
                                    elif "Código Equipamento" in item: novo_item["Frota"] = item["Código Equipamento"]
                                    
                                    for k, v in item.items():
                                        k_lower = k.lower()
                                        if (k.startswith("Horas_") or 
                                            k.startswith("Porcentagem_") or 
                                            k.startswith("Disponibilidade_") or
                                            k.startswith("Eficiencia_") or
                                            k.startswith("Manobras_") or
                                            k.startswith("Basculamento_") or
                                            k.startswith("Velocidade_") or
                                            k.startswith("Media_") or
                                            k.startswith("Producao_") or
                                            k.startswith("Toneladas_") or
                                            k.startswith("Consumo_") or
                                            k.startswith("Uso_") or
                                            k_lower == "motor ligado" or
                                            k_lower == "motor ocioso" or
                                            "velocidade" in k_lower or
                                            "producao" in k_lower or
                                            "produção" in k_lower or
                                            "eficiencia" in k_lower or
                                            "eficiência" in k_lower or
                                            "manobras" in k_lower or
                                            "basculamento" in k_lower or
                                            "gps" in k_lower):
                                            novo_item[k] = v
                                    
                                    if novo_item:
                                        dados_filtrados.append(novo_item)
                            
                            adicionar_ao_grupo(tipo_frota, "Resumo_Dia", dados_filtrados)

                        # 2. Operadores (Separado por aba) -> Ex: Operadores_COLHEDORA
                        elif key_aba.startswith("Operadores_"):
                            tipo_frota = key_aba.replace("Operadores_", "")
                            adicionar_ao_grupo(tipo_frota, "Operadores", dados_lista)

                        # 3. Top5Ofensores (Separado por aba) -> Ex: Top5Ofensores_COLHEDORA
                        elif key_aba.startswith("Top5Ofensores_"):
                            tipo_frota = key_aba.replace("Top5Ofensores_", "")
                            adicionar_ao_grupo(tipo_frota, "Top5Ofensores", dados_lista)

                        # 4. Intervalos (Separado por aba) -> Ex: Intervalos_COLHEDORA
                        elif key_aba.startswith("Intervalos_"):
                            tipo_frota = key_aba.replace("Intervalos_", "")
                            adicionar_ao_grupo(tipo_frota, "Intervalos", dados_lista)
                        
                        # Casos genéricos (se o tratamento não separou por abas)
                        elif key_aba == "Equipamentos_Dia":
                            # Tentar separar manualmente se houver campo de descrição
                            # Se não, joga em "outros" ou "geral"
                            pass # Implementar se necessário, mas o foco é na estrutura nova
                        
                        elif key_aba == "Operadores":
                             pass

                    # Salvar os arquivos JSON agrupados e separados por tipo (Frota vs Operadores)
                    for tipo_frota, conteudo_json in dados_por_tipo.items():
                        
                        # Separar dados de Operadores
                        dados_operadores = None
                        if "Operadores" in conteudo_json:
                            dados_operadores = conteudo_json.pop("Operadores")
                        
                        dados_frota = conteudo_json # O que sobrou é frota

                        # 1. Salvar arquivo de Frota (se houver dados)
                        if dados_frota:
                            # Criar pasta específica: json/colhedora/frotas/diario
                            dir_frota = os.path.join(DIRETORIO_JSON, tipo_frota, "frotas", "diario")
                            if not os.path.exists(dir_frota):
                                os.makedirs(dir_frota)

                            # Reestruturar dados_frota para agrupar por Frota (chave principal)
                            dados_frota_agrupados = {}
                            
                            # Categorias esperadas: Resumo_Dia, Top5Ofensores, Intervalos
                            for categoria, lista_itens in dados_frota.items():
                                if not isinstance(lista_itens, list):
                                    # Se não for lista, não sabemos lidar, mantém na raiz de "Geral" ou similar?
                                    # Por segurança, vamos pular ou tratar diferente. Mas aqui espera-se listas.
                                    continue
                                
                                for item in lista_itens:
                                    # Identificar ID da Frota
                                    id_frota = None
                                    chave_frota_encontrada = None
                                    
                                    # Tentativas de encontrar a chave de frota
                                    for k in ["Frota", "Código Equipamento", "Equipamento"]:
                                        if k in item:
                                            id_frota = item[k]
                                            chave_frota_encontrada = k
                                            break
                                    
                                    if id_frota is not None:
                                        id_frota_str = str(id_frota)
                                        
                                        # Cria entrada da frota se não existir
                                        if id_frota_str not in dados_frota_agrupados:
                                            dados_frota_agrupados[id_frota_str] = {}
                                        
                                        # Cria categoria dentro da frota se não existir
                                        if categoria not in dados_frota_agrupados[id_frota_str]:
                                            dados_frota_agrupados[id_frota_str][categoria] = []
                                        
                                        # Cria cópia do item para remover a chave da frota (evitar redundância)
                                        item_limpo = item.copy()
                                        if chave_frota_encontrada:
                                            del item_limpo[chave_frota_encontrada]
                                        
                                        dados_frota_agrupados[id_frota_str][categoria].append(item_limpo)
                                    else:
                                        # Se não achou frota, joga num grupo "Geral"
                                        if "Geral" not in dados_frota_agrupados:
                                            dados_frota_agrupados["Geral"] = {}
                                        if categoria not in dados_frota_agrupados["Geral"]:
                                            dados_frota_agrupados["Geral"][categoria] = []
                                        dados_frota_agrupados["Geral"][categoria].append(item)

                            nome_arquivo_frota = f"{tipo_frota}_frota_{data_str}.json"
                            caminho_frota = os.path.join(dir_frota, nome_arquivo_frota)
                            try:
                                with open(caminho_frota, 'w', encoding='utf-8') as f:
                                    json.dump(dados_frota_agrupados, f, ensure_ascii=False, indent=2)
                                print(f"  -> JSON Frota salvo: {caminho_frota}")
                            except Exception as e_esp:
                                print(f"  -> ERRO ao salvar {nome_arquivo_frota}: {e_esp}")

                        # 2. Salvar arquivo de Operadores (se houver dados)
                        if dados_operadores:
                            # Criar pasta específica: json/colhedora/operadores/diario
                            dir_ops = os.path.join(DIRETORIO_JSON, tipo_frota, "operadores", "diario")
                            if not os.path.exists(dir_ops):
                                os.makedirs(dir_ops)

                            # Reestruturar dados_operadores para indexar por "Código - Nome"
                            dados_operadores_agrupados = {}
                            
                            # dados_operadores é uma lista de dicionários
                            if isinstance(dados_operadores, list):
                                for item in dados_operadores:
                                    # Identificar ID do Operador
                                    id_op = None
                                    chave_op_encontrada = None
                                    
                                    # Busca chave do código do operador
                                    # Baseado no arquivo lido: "Código de Operador"
                                    for k in ["Código de Operador", "Codigo Operador", "Cod Operador"]:
                                        if k in item:
                                            id_op = item[k]
                                            chave_op_encontrada = k
                                            break
                                    
                                    # Identificar Nome do Operador
                                    nome_op = "Desconhecido"
                                    chave_nome_encontrada = None
                                    for k in ["Nome", "Nome Operador", "Nome do Operador", "Operador"]:
                                        if k in item:
                                            nome_op = item[k]
                                            chave_nome_encontrada = k
                                            break
                                    
                                    if id_op is not None:
                                        # Formatar chave: "Cód - Nome"
                                        chave_final = f"{id_op} - {nome_op}"
                                        
                                        # Cria cópia para remover redundância
                                        item_limpo = item.copy()
                                        
                                        # Remover chaves usadas na composição do ID principal (se desejado remover redundância)
                                        if chave_op_encontrada:
                                            del item_limpo[chave_op_encontrada]
                                        
                                        # Opcional: Remover o nome também se já está na chave?
                                        # O usuário pediu para "não repetir em todas as entradas" no caso da frota.
                                        # Vamos manter a consistência e remover se encontrou a chave exata.
                                        if chave_nome_encontrada:
                                            del item_limpo[chave_nome_encontrada]
                                            
                                        # Como é um resumo por dia/tipo, assume-se 1 entrada por operador.
                                        dados_operadores_agrupados[chave_final] = item_limpo
                                    else:
                                        # Sem código (improvável se vier do tratamento correto)
                                        if "SemCodigo" not in dados_operadores_agrupados:
                                            dados_operadores_agrupados["SemCodigo"] = []
                                        dados_operadores_agrupados["SemCodigo"].append(item)
                            
                            nome_arquivo_ops = f"{tipo_frota}_operadores_{data_str}.json"
                            caminho_ops = os.path.join(dir_ops, nome_arquivo_ops)
                            try:
                                with open(caminho_ops, 'w', encoding='utf-8') as f:
                                    json.dump(dados_operadores_agrupados, f, ensure_ascii=False, indent=2)
                                print(f"  -> JSON Operadores salvo: {caminho_ops}")
                            except Exception as e_esp:
                                print(f"  -> ERRO ao salvar {nome_arquivo_ops}: {e_esp}")

                    # --------------------------------------------------------------------------

                else:
                    print(f"  -> AVISO: Nenhuma aba gerada para {data_str}. Arquivo não salvo.")
            
        # --- 6. Gerar JSONs de Período (Semanal/Mensal) ---
        print("\n=== GERANDO ARQUIVOS DE PERÍODO ===")
        
        # Definir string do período
        periodo_str = "periodo_desconhecido"
        if dt_inicio_filtro and dt_fim_filtro:
            periodo_str = f"{dt_inicio_filtro.strftime('%d-%m-%Y')}_{dt_fim_filtro.strftime('%d-%m-%Y')}"
        
        # Processar Frota (Periodo)
        if "Periodo_Equipamentos" in dfs:
            print(f"  Processando Periodo_Equipamentos...")
            df_p_frota = dfs["Periodo_Equipamentos"]
            
            # Tentar identificar coluna de tipo de frota
            col_tipo = None
            for c in ["Descrição do Equipamento", "Grupo", "Tipo"]:
                if c in df_p_frota.columns:
                    col_tipo = c
                    break
            
            if col_tipo:
                grupos = df_p_frota[col_tipo].dropna().unique()
                for tipo in grupos:
                    df_tipo = df_p_frota[df_p_frota[col_tipo] == tipo].copy()
                    
                    # Remover a coluna de tipo para economizar espaço
                    df_tipo = df_tipo.drop(columns=[col_tipo])
                    
                    nome_tipo_norm = normalizar_nome_pasta(tipo)
                    
                    # Preparar estrutura JSON (agrupada por ID da frota ou Geral)
                    # Usar to_json/loads para garantir serialização correta de datas e NaNs
                    json_str = df_tipo.to_json(orient='records', date_format='iso', default_handler=str)
                    records = json.loads(json_str)
                    
                    dados_frota_agrupados = {}
                    
                    for item in records:
                        # Identificar ID
                        id_frota = None
                        chave_frota_encontrada = None
                         # Tentativas de encontrar a chave de frota
                        for k in ["Frota", "Código Equipamento", "Equipamento"]:
                            if k in item:
                                id_frota = item[k]
                                chave_frota_encontrada = k
                                break
                        
                        if id_frota is not None:
                            id_frota_str = str(id_frota)
                            if id_frota_str not in dados_frota_agrupados:
                                dados_frota_agrupados[id_frota_str] = {}
                            
                            # No período, geralmente é um resumo único por frota
                            if "Resumo_Periodo" not in dados_frota_agrupados[id_frota_str]:
                                dados_frota_agrupados[id_frota_str]["Resumo_Periodo"] = []
                            
                            item_limpo = item.copy()
                            if chave_frota_encontrada:
                                del item_limpo[chave_frota_encontrada]
                            
                            dados_frota_agrupados[id_frota_str]["Resumo_Periodo"].append(item_limpo)
                        else:
                             if "Geral" not in dados_frota_agrupados:
                                 dados_frota_agrupados["Geral"] = {}
                             if "Resumo_Periodo" not in dados_frota_agrupados["Geral"]:
                                 dados_frota_agrupados["Geral"]["Resumo_Periodo"] = []
                             dados_frota_agrupados["Geral"]["Resumo_Periodo"].append(item)

                    # Salvar
                    dir_frota = os.path.join(DIRETORIO_JSON, nome_tipo_norm, "frotas", "semanal")
                    if not os.path.exists(dir_frota):
                        os.makedirs(dir_frota)
                        
                    nome_arquivo = f"{nome_tipo_norm}_frota_periodo_{periodo_str}.json"
                    caminho_arquivo = os.path.join(dir_frota, nome_arquivo)
                    
                    try:
                        with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                            json.dump(dados_frota_agrupados, f, ensure_ascii=False, indent=2)
                        print(f"    -> Salvo: {caminho_arquivo}")
                    except Exception as e:
                        print(f"    -> Erro ao salvar {nome_arquivo}: {e}")

        # Processar Operadores (Periodo)
        if "Periodo_Operadores" in dfs:
            print(f"  Processando Periodo_Operadores...")
            df_p_op = dfs["Periodo_Operadores"]
            
            col_tipo = None
            for c in ["Descrição do Equipamento", "Grupo", "Tipo"]:
                if c in df_p_op.columns:
                    col_tipo = c
                    break
            
            lista_tipos = []
            if col_tipo:
                lista_tipos = df_p_op[col_tipo].dropna().unique()
            else:
                lista_tipos = ["Geral"]

            for tipo in lista_tipos:
                if col_tipo:
                    df_tipo = df_p_op[df_p_op[col_tipo] == tipo].copy()
                    df_tipo = df_tipo.drop(columns=[col_tipo])
                    nome_tipo_norm = normalizar_nome_pasta(tipo)
                else:
                    df_tipo = df_p_op.copy()
                    nome_tipo_norm = "geral"

                json_str = df_tipo.to_json(orient='records', date_format='iso', default_handler=str)
                records = json.loads(json_str)
                dados_operadores_agrupados = {}
                
                for item in records:
                    # Identificar ID e Nome
                    id_op = None
                    chave_op_encontrada = None
                    for k in ["Código de Operador", "Codigo Operador", "Cod Operador"]:
                        if k in item:
                            id_op = item[k]
                            chave_op_encontrada = k
                            break
                    
                    nome_op = "Desconhecido"
                    chave_nome_encontrada = None
                    for k in ["Nome", "Nome Operador", "Nome do Operador", "Operador"]:
                        if k in item:
                            nome_op = item[k]
                            chave_nome_encontrada = k
                            break
                    
                    if id_op is not None:
                        chave_final = f"{id_op} - {nome_op}"
                        item_limpo = item.copy()
                        if chave_op_encontrada:
                            del item_limpo[chave_op_encontrada]
                        if chave_nome_encontrada:
                            del item_limpo[chave_nome_encontrada]
                            
                        dados_operadores_agrupados[chave_final] = item_limpo
                    else:
                        if "SemCodigo" not in dados_operadores_agrupados:
                            dados_operadores_agrupados["SemCodigo"] = []
                        dados_operadores_agrupados["SemCodigo"].append(item)
                
                # Salvar
                dir_frota = os.path.join(DIRETORIO_JSON, nome_tipo_norm, "operadores", "semanal")
                if not os.path.exists(dir_frota):
                    os.makedirs(dir_frota)
                    
                nome_arquivo = f"{nome_tipo_norm}_operadores_periodo_{periodo_str}.json"
                caminho_arquivo = os.path.join(dir_frota, nome_arquivo)
                
                try:
                    with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                        json.dump(dados_operadores_agrupados, f, ensure_ascii=False, indent=2)
                    print(f"    -> Salvo: {caminho_arquivo}")
                except Exception as e:
                    print(f"    -> Erro ao salvar {nome_arquivo}: {e}")

        print(f"\nSucesso! Arquivos Excel e JSON gerados nas pastas 'separados/xlsx' e 'separados/json'.")
            
    except Exception as e:
        print(f"\nERRO FATAL: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
