import json
import os
import glob
import re
import sys
import zipfile
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ─── Paths ──────────────────────────────────────────────────────────────────────

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
ETL_ROOT = os.path.dirname(SCRIPTS_DIR)

SOLINFTEC_JSON_DIR = os.path.join(
    ETL_ROOT, "dados", "separados", "json", "colhedora", "frotas", "diario"
)
TRATORES_JSON_DIR = os.path.join(
    ETL_ROOT, "dados", "separados", "json", "tratores", "frotas", "diario"
)
OPC_XLSX_DIR = os.path.join(ETL_ROOT, "dados", "separados", "xlsx")
CASE_DIR = os.path.join(ETL_ROOT, "dados")
LINHA_TEMPO_DIR = os.path.join(ETL_ROOT, "dados")

OUTPUT_DIR = SOLINFTEC_JSON_DIR

# Metas default (mesmas do frontend config/metas.json)
METAS_DEFAULT = {
    "eficienciaEnergetica": 85,
    "eficienciaOperacional": 60,
    "horaElevador": 5,
    "usoGPS": 90,
    "mediaVelocidade": 5,
    "manobras": 60,
    "disponibilidadeMecanica": 90,
    "motorOcioso": 15,
}


# ─── Helpers ────────────────────────────────────────────────────────────────────

def parse_date_from_filename(filename: str) -> str | None:
    """Extrai data DD-MM-YYYY do nome do arquivo."""
    m = re.search(r"(\d{2}-\d{2}-\d{4})", filename)
    return m.group(1) if m else None


def date_ddmmyyyy_to_iso(date_str: str) -> str:
    """Converte DD-MM-YYYY para YYYY-MM-DD."""
    parts = date_str.split("-")
    return f"{parts[2]}-{parts[1]}-{parts[0]}"


def time_hhmmss(dt_str: str) -> str:
    """Extrai HH:MM:SS de uma string 'DD/MM/YYYY HH:MM:SS'."""
    parts = dt_str.strip().split(" ")
    return parts[1] if len(parts) > 1 else dt_str


def parse_datetime(dt_str: str) -> datetime | None:
    """Parse 'DD/MM/YYYY HH:MM:SS' para datetime."""
    try:
        return datetime.strptime(dt_str.strip(), "%d/%m/%Y %H:%M:%S")
    except (ValueError, AttributeError):
        return None


def calc_duration_hours(start_str: str, end_str: str) -> float:
    """Calcula duração em horas entre duas strings de data."""
    s = parse_datetime(start_str)
    e = parse_datetime(end_str)
    if s and e:
        return (e - s).total_seconds() / 3600
    return 0.0


def safe_float(val, default=0.0) -> float:
    """Convert to float safely."""
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


# ─── Solinftec ──────────────────────────────────────────────────────────────────

def load_solinftec(date_str: str, base_dir: str = SOLINFTEC_JSON_DIR) -> dict | None:
    """Carrega JSON Solinftec bruto para a data (DD-MM-YYYY).
       Procura por subpastas ativamente no base_dir para evitar perder arquivos.
    """
    # Exemplo: busca em cd-diario-op, cd-diario-frotas ou qualquer lugar abaixo de json/
    # O base_dir idealmente deve ser o mais alto possível para a fonte, 
    # ou podemos subir um pouco a árvore para encontrar.
    search_dir = os.path.dirname(os.path.dirname(os.path.dirname(base_dir))) # sobe de "colhedora/frotas/diario" para "json"
    
    # Filtro adicional para não cruzar dados de trator e colhedora se não for o desejado
    # mas o nome do arquivo geralmente tem "COLHEDORA" ou "TRATORES"
    pattern = os.path.join(search_dir, "**", f"*{date_str}*.json")
    files = glob.glob(pattern, recursive=True)

    if not files:
        return None
    
    # Queremos priorizar apenas os da "família" correta (Trator ou Colhedora)
    # se base_dir apontar para trator, buscaremos 'trator', senão 'colhedora' no path ou no nome
    is_trator = "trator" in base_dir.lower()
    
    valid_files = []
    for f in files:
        f_lower = str(f).lower()
        
        # NUNCA usar arquivos de operadores como fonte para frotas
        if "operador" in f_lower:
            continue
        # Ignorar arquivos grunner (não são colhedoras nem tratores)
        if "grunner" in f_lower:
            continue
            
        if is_trator and ("trator" in f_lower or "tt" in f_lower):
            valid_files.append(f)
        elif not is_trator and ("colhedora" in f_lower or "cd" in f_lower):
            valid_files.append(f)
            
    if not valid_files:
        # Fallback: usar apenas arquivos que estejam na pasta frotas
        valid_files = [f for f in files if "operador" not in str(f).lower() and "frotas" in str(f).lower()]
    
    for latest in sorted(valid_files, key=os.path.getmtime, reverse=True):
        try:
            with open(latest, "r", encoding="utf-8") as f:
                data = json.load(f)
                
                # Se for um arquivo já consolidado (tem metadata ou coleções consolidadas), ignora
                if "metadata" in data or "eficiencia_energetica" in data:
                    continue
                
                print(f"     ℹ️ Usando fonte Solinftec: {os.path.basename(latest)}")
                return data
                
        except Exception as e:
            continue
            
    print(f"  ❌ Nenhum arquivo Solinftec válido ou não-consolidado encontrado para {date_str} em {search_dir}.")
    return None


# ─── Case IH ───────────────────────────────────────────────────────────────────

def load_case_data() -> dict:
    """
    Carrega todos os dados Case IH dos XLSX consolidados.
    Retorna dict: { 'DD/MM/YYYY': { 'frota_id': { ...campos... } } }
    """
    case_files = glob.glob(os.path.join(CASE_DIR, "Consolidado_Case_*.xlsx"))
    if not case_files:
        print("  ⚠️  Nenhum arquivo Case IH encontrado.")
        return {}

    case_data = defaultdict(lambda: defaultdict(dict))

    for cf in case_files:
        print(f"  📂 Carregando Case: {os.path.basename(cf)}")
        try:
            wb = openpyxl.load_workbook(cf, read_only=True)
        except (KeyError, zipfile.BadZipFile, OSError) as e:
            print(f"  ⚠️  Arquivo Case inválido ou corrompido, ignorando: {os.path.basename(cf)} ({e})")
            continue

        # Aba "Resumo" contém dados por frota para o período inteiro
        if "Resumo" in wb.sheetnames:
            ws = wb["Resumo"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    if frota:
                        case_data["_resumo_geral"][frota] = {
                            "horasMotor": safe_float(d.get("Total Horas Motor (Diferença)")),
                            "rpm": safe_float(d.get("RPM")),
                            "temperaturaArrefecimento": safe_float(d.get("Média Temperatura líquido de arrefecimento do motor")),
                            "temperaturaTransmissao": safe_float(d.get("Média Temperatura do óleo da transmissão")),
                            "velocidadeMedia": safe_float(d.get("Velocidade Média")),
                        }

        if "Resumo Diário" in wb.sheetnames:
            ws = wb["Resumo Diário"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    data_val = d.get("Data", "")
                    
                    if isinstance(data_val, datetime):
                        data_key = data_val.strftime("%d/%m/%Y")
                    elif isinstance(data_val, str):
                        data_key = data_val.split(" ")[0] if " " in data_val else data_val
                    else:
                        continue

                    if frota and data_key:
                        horas_motor = safe_float(d.get("Total Horas Motor (Dif)"), safe_float(d.get("Total Horas Motor (Diferença)")))
                        rpm = safe_float(d.get("RPM Médio", d.get("RPM")))
                        temp_arref = safe_float(d.get("Média Temp. Líquido Arrefecimento", d.get("Média Temperatura líquido de arrefecimento do motor")))
                        temp_trans = safe_float(d.get("Média Temp. Óleo Hidráulico", d.get("Média Temperatura do óleo da transmissão")))
                        temp_ar = safe_float(d.get("Média Temp. Ar Coletor Admissão"))
                        vel_media = safe_float(d.get("Velocidade Média"))
                        tempo_gps_ligado = safe_float(d.get("Tempo GPS Ligado"))
                        tempo_gps_desligado = safe_float(d.get("Tempo GPS Desligado"))
                        perc_gps_ligado = safe_float(d.get("% GPS Ligado"))
                        perc_gps_desligado = safe_float(d.get("% GPS Desligado"))
                        
                        # Tempos detalhados do Motor
                        motor_ocioso = safe_float(d.get("Motor Ocioso"))
                        motor_desligado = safe_float(d.get("Motor Desligado"))
                        horas_produtivas = safe_float(d.get("Horas Produtivas"))
                        tempo_registrado = safe_float(d.get("Tempo Registrado (Total)"))
                        
                        case_data[data_key][frota] = {
                            "Hora Motor Inicial": d.get("Hora Motor Inicial"),
                            "Hora Motor Final": d.get("Hora Motor Final"),
                            "Horas Motor": horas_motor,
                            "rpm": rpm,
                            "temperaturaArrefecimento": temp_arref,
                            "temperaturaTransmissao": temp_trans,
                            "temperaturaArAdmissao": temp_ar,
                            "velocidadeMedia": vel_media,
                            "motorOcioso": motor_ocioso,
                            "motorDesligado": motor_desligado,
                            "horasProdutivas": horas_produtivas,
                            "tempoRegistrado": tempo_registrado,
                            "Extras": {
                                "eficienciaEnergetica": safe_float(d.get("Eficiencia_Energetica")),
                                "eficienciaOperacional": safe_float(d.get("Eficiencia_Operacional")),
                                "usoGPS": perc_gps_ligado,
                                "horasMotor": horas_motor,
                                "velocidadeMedia": vel_media,
                                "tempoGPSLigado": tempo_gps_ligado,
                                "tempoGPSDesligado": tempo_gps_desligado,
                                "percGPSLigado": perc_gps_ligado,
                                "percGPSDesligado": perc_gps_desligado,
                            },
                        }

        # Aba "Dados" contém intervalos detalhados com coordenadas
        if "Dados" in wb.sheetnames:
            ws = wb["Dados"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    data_hora = d.get("Data Hora Local", "")
                    
                    if isinstance(data_hora, datetime):
                        data_key = data_hora.strftime("%d/%m/%Y")
                    elif isinstance(data_hora, str):
                        data_key = data_hora.split(" ")[0] if " " in data_hora else ""
                    else:
                        continue

                    if frota and data_key:
                        if "_intervalos" not in case_data[data_key].get(frota, {}):
                            if frota not in case_data[data_key]:
                                case_data[data_key][frota] = {}
                            case_data[data_key][frota].setdefault("_intervalos", [])
                        
                        case_data[data_key][frota]["_intervalos"].append({
                            "inicio": str(data_hora),
                            "duracao": safe_float(d.get("Duração")),
                            "operacao": str(d.get("Descrição da Operação", "")),
                            "grupo": str(d.get("Descrição do Grupo da Operação", "")),
                            "lat": safe_float(d.get("Latitude")),
                            "lon": safe_float(d.get("Longitude")),
                        })

        wb.close()

    return dict(case_data)


# ─── OPC (XLSX diário) ─────────────────────────────────────────────────────────

def load_opc_daily(date_str: str) -> dict | None:
    """
    Carrega dados do XLSX diário OPC para a data (DD-MM-YYYY).
    As abas relevantes são COLHEDORA_Dia, TRANSBORDO_Dia, GRUNNER_Dia.
    """
    xlsx_path = os.path.join(OPC_XLSX_DIR, f"{date_str}.xlsx")
    if not os.path.exists(xlsx_path):
        return None

    print(f"  📂 Carregando OPC: {date_str}.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    opc = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        sheet_data = []
        for row in rows[1:]:
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            sheet_data.append(d)

        opc[sheet_name] = sheet_data

    wb.close()
    return opc


# ─── Consolidação ──────────────────────────────────────────────────────────────

def consolidar_dia(
    date_str: str,
    solinftec_raw: dict,
    case_data_by_date: dict,
    opc_data: dict | None,
    include_case: bool,
) -> dict:
    """
    Consolida dados de um dia no formato do frontend.
    
    date_str: DD-MM-YYYY
    solinftec_raw: dict com frotas como chaves, cada uma com Resumo_Dia e Intervalos
    case_data_by_date: dados Case para a data (dict com frota como chave)
    opc_data: dados OPC para a data
    """
    iso_date = date_ddmmyyyy_to_iso(date_str)
    date_display = date_str.replace("-", "/")

    # Normalizar data para case lookup (DD/MM/YYYY)
    case_date_key = date_display
    case_frotas = case_data_by_date.get(case_date_key, {}) if include_case else {}

    if solinftec_raw and isinstance(solinftec_raw, dict) and "metadata" in solinftec_raw and "eficiencia_energetica" in solinftec_raw:
        resultado = dict(solinftec_raw)
        if not include_case:
            chaves_filtrar = [
                "eficiencia_energetica",
                "eficiencia_operacional",
                "horas_elevador",
                "uso_gps",
                "media_velocidade",
                "manobras_frotas",
                "motor_ocioso",
                "disponibilidade_mecanica",
                "horas_por_frota",
                "intervalos_operacao",
            ]
            for chave in chaves_filtrar:
                if isinstance(resultado.get(chave), list):
                    resultado[chave] = [item for item in resultado[chave] if item.get("fonte") != "case"]
            if isinstance(resultado.get("metadata"), dict):
                fontes = resultado["metadata"].get("fontes", [])
                if isinstance(fontes, list):
                    resultado["metadata"]["fontes"] = [f for f in fontes if f != "case"]
            if "dados_case" in resultado:
                del resultado["dados_case"]

        # ── Remover chaves obsoletas explicitamente ──
        chaves_remover = ["lavagem", "roletes", "producao", "producao_total", "producao_por_frota"]
        for chave in chaves_remover:
            if chave in resultado:
                del resultado[chave]

        if "metas" not in resultado or not isinstance(resultado.get("metas"), dict):
            resultado["metas"] = METAS_DEFAULT

        # ── Recalcular ofensores se estiverem vazios (fix: early return bypassa cálculo) ──
        if not resultado.get("ofensores"):
            op_stats = defaultdict(float)
            for intv in resultado.get("intervalos_operacao", []):
                tipo = intv.get("tipo", "")
                # "Disponível" no JSON corresponde a IMPRODUTIVA (tudo que não é produtivo/manutenção/falta de info)
                # Mas precisamos dos grupos originais. No JSON consolidado, "tipo" é simplificado.
                # Melhor abordagem: verificar se o intervalo NÃO é Produtivo e NÃO é Falta de Informação
                if tipo in ("Disponível", "Manutenção"):
                    # Usamos "tipo" como descrição genérica se não tiver nome da operação
                    # No JSON consolidado, não temos "Descrição da Operação" preservada
                    # Vamos tentar usar o campo se existir, senão usar o tipo
                    desc = intv.get("operacao", intv.get("descricao", tipo))
                    dur = safe_float(intv.get("duracaoHoras", 0))
                    if dur > 0 and desc:
                        op_stats[desc] += dur

            if op_stats:
                total_improd = sum(op_stats.values())
                ofensores_sorted = sorted(op_stats.items(), key=lambda x: x[1], reverse=True)[:5]
                resultado["ofensores"] = [
                    {
                        "nome": op,
                        "percentual": round((tempo / total_improd * 100) if total_improd > 0 else 0, 2),
                        "duracao": round(tempo, 4),
                    }
                    for i, (op, tempo) in enumerate(ofensores_sorted)
                ]

        return resultado

    # Coletar todas as frotas de todas as fontes
    all_frotas = set()
    if solinftec_raw:
        all_frotas.update(solinftec_raw.keys())
    if include_case:
        all_frotas.update(k for k in case_frotas.keys() if not k.startswith("_"))

    fontes_usadas = []
    if solinftec_raw:
        fontes_usadas.append("solinftec")
    if include_case and case_frotas:
        fontes_usadas.append("case")
    if opc_data:
        fontes_usadas.append("opc")

    # Arrays para cada seção do frontend
    eficiencia_energetica = []
    eficiencia_operacional = []
    horas_elevador = []
    uso_gps = []
    media_velocidade = []
    manobras_frotas = []
    motor_ocioso = []
    disponibilidade_mecanica = []
    intervalos_operacao = []
    horas_por_frota = []

    operation_stats = defaultdict(float)  # Para ofensores
    idx = 0

    for frota_id in sorted(all_frotas):
        idx += 1
        resumo = None
        intervalos = []
        fonte = "desconhecida"
        case_extra = case_frotas.get(frota_id, {})

        # Dados Solinftec
        if solinftec_raw and frota_id in solinftec_raw:
            frota_data = solinftec_raw[frota_id]
            resumo = frota_data.get("Resumo_Dia", [{}])[0] if frota_data.get("Resumo_Dia") else {}
            intervalos = frota_data.get("Intervalos", [])
            fonte = "solinftec"

        # ── Eficiência Energética ──
        horas_produtivas = safe_float(resumo.get("Horas_Produtivas")) if resumo else 0
        horas_motor = safe_float(resumo.get("Horas_Motor_Ligado")) if resumo else safe_float(case_extra.get("horasMotor"))
        
        ef_energetica = 0.0
        if resumo and "Eficiencia_Energetica" in resumo:
            ef_energetica = safe_float(resumo["Eficiencia_Energetica"]) * 100
        elif horas_motor > 0 and horas_produtivas > 0:
            ef_energetica = (horas_produtivas / horas_motor) * 100

        eficiencia_energetica.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": round(ef_energetica, 2),
            "horasMotor": round(horas_motor, 4),
            "horasElevador": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Eficiência Operacional ──
        horas_registradas = safe_float(resumo.get("Horas_Registradas")) if resumo else 0

        ef_operacional = 0.0
        if resumo and "Eficiencia_Operacional" in resumo:
            ef_operacional = safe_float(resumo["Eficiencia_Operacional"]) * 100
        elif horas_registradas > 0 and horas_produtivas > 0:
            ef_operacional = (horas_produtivas / horas_registradas) * 100

        eficiencia_operacional.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": round(ef_operacional, 2),
            "horasMotor": round(horas_registradas, 4),
            "horasElevador": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Horas Elevador ──
        horas_elevador.append({
            "id": idx,
            "nome": frota_id,
            "valor": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Uso GPS ──
        uso_gps_val = 0.0
        # Solinftec não fornece dados de piloto/GPS. Não inventar dados baseados em "Sem Apontamento".
        # if resumo and "Porcentagem_Sem_Apontamento" in resumo:
        #     uso_gps_val = 100 - safe_float(resumo["Porcentagem_Sem_Apontamento"])
        uso_gps.append({
            "id": idx,
            "nome": frota_id,
            "porcentagem": round(uso_gps_val, 2),
            "fonte": fonte,
        })

        # ── Média de Velocidade ──
        vel = 0.0
        if resumo and "Velocidade_Media" in resumo:
            vel = safe_float(resumo["Velocidade_Media"])
        elif resumo and "Vel_Colheita_media" in resumo:
            vel = safe_float(resumo["Vel_Colheita_media"])
        elif case_extra.get("velocidadeMedia"):
            vel = safe_float(case_extra["velocidadeMedia"])
            fonte = "case"
        media_velocidade.append({
            "id": idx,
            "nome": frota_id,
            "velocidade": round(vel, 4),
            "fonte": fonte if vel > 0 else "solinftec",
        })

        # ── Motor Ocioso ──
        motor_ocioso_pct = safe_float(resumo.get("Porcentagem_Motor_Ocioso")) if resumo else 0
        tempo_ocioso = safe_float(resumo.get("Horas_Motor_Ocioso")) if resumo else 0
        motor_ocioso.append({
            "id": idx,
            "nome": frota_id,
            "percentual": round(motor_ocioso_pct, 4),
            "tempoLigado": round(horas_motor, 4),
            "tempoOcioso": round(tempo_ocioso, 4),
            "fonte": fonte,
        })

        # ── Disponibilidade Mecânica ──
        disp_mec = 100.0
        horas_manutencao = 0.0
        if resumo and "Disponibilidade_Mecanica" in resumo:
            disp_mec = safe_float(resumo["Disponibilidade_Mecanica"]) * 100
            horas_manutencao = safe_float(resumo.get("Horas_Manutencao", 0))
        disponibilidade_mecanica.append({
            "id": idx,
            "nome": frota_id,
            "disponibilidade": round(disp_mec, 2),
            "horasMotor": round(horas_motor, 4),
            "tempoManutencao": round(horas_manutencao, 4),
            "fonte": fonte,
        })

        # ── Horas por Frota ──
        horas_por_frota.append({
            "id": idx,
            "nome": frota_id,
            "frota": frota_id,
            "horas": round(horas_registradas, 2),
            "fonte": fonte,
        })


        # ── Manobras ──
        qtd_manobras = 0
        tempo_total_manobras = 0.0
        tempo_medio_manobras = 0.0

        if resumo:
            qtd_manobras = int(safe_float(resumo.get("Quantidade_Manobras", 0)))
            tempo_total_manobras = safe_float(resumo.get("Tempo_Total_Manobras_h", 0))
            tempo_medio_manobras = safe_float(resumo.get("Tempo_Medio_Manobras_min", 0))

        # Calcular hh:mm do tempo total e médio
        total_h = int(tempo_total_manobras)
        total_m = int((tempo_total_manobras - total_h) * 60)
        total_s = int(((tempo_total_manobras - total_h) * 60 - total_m) * 60)
        medio_m = int(tempo_medio_manobras)
        medio_s = int((tempo_medio_manobras - medio_m) * 60)

        manobras_frotas.append({
            "Frota": frota_id,
            "Tempo Total": round(tempo_total_manobras, 4),
            "Tempo Médio": round(tempo_medio_manobras / 60 if tempo_medio_manobras > 0 else 0, 6),
            "Intervalos Válidos": qtd_manobras,
            "Tempo Total (hh:mm)": f"{total_h:02d}:{total_m:02d}:{total_s:02d}",
            "Tempo Médio (hh:mm)": f"00:{medio_m:02d}:{medio_s:02d}",
            "fonte": fonte,
        })

        # ── Intervalos de Operação (Gantt) ──
        for intv in intervalos:
            start_str = intv.get("Início", "")
            end_str = intv.get("Fim", "")
            dur = calc_duration_hours(start_str, end_str)

            grupo = str(intv.get("Descrição do Grupo da Operação", intv.get("Grupo", ""))).strip().upper()
            descricao = str(intv.get("Descrição da Operação", intv.get("operacao", ""))).strip()

            tipo = "Disponível"
            if grupo == "PRODUTIVA":
                tipo = "Produtivo"
            elif grupo == "MANUTENÇÃO":
                tipo = "Manutenção"
            if descricao == "SEM APONTAMENTO":
                tipo = "Falta de Informação"

            intervalos_operacao.append({
                "equipamento": frota_id,
                "tipo": tipo,
                "inicio": time_hhmmss(start_str),
                "duracaoHoras": round(dur, 6),
                "fonte": "solinftec",
            })

            # Agregar ofensores (Fallback se não vier pronto)
            if grupo in ("IMPRODUTIVA", "MANUTENÇÃO", "MANUTENCAO"):
                if descricao:
                    operation_stats[descricao] += dur

        # Intervalos Case (se houver)
        case_intervals = case_extra.get("_intervalos", [])
        for ci in case_intervals:
            dur = safe_float(ci.get("duracao", 0))
            grupo = ci.get("grupo", "")
            operacao = ci.get("operacao", "")
            
            tipo = "Disponível"
            if "PRODUTIVA" in grupo.upper():
                tipo = "Produtivo"
            elif "MANUTENÇÃO" in grupo.upper() or "MANUTENCAO" in grupo.upper():
                tipo = "Manutenção"

            intervalos_operacao.append({
                "equipamento": frota_id,
                "tipo": tipo,
                "inicio": ci.get("inicio", ""),
                "duracaoHoras": round(dur, 6),
                "fonte": "case",
            })

            if "IMPRODUTIVA" in grupo.upper() or "MANUTENÇÃO" in grupo.upper():
                operation_stats[operacao] += dur

    # ── Ofensores (Top 5) ──
    ofensores_list = []
    top5_agg = defaultdict(float)
    found_source = False

    # 1. Tentar carregar do OPC (XLSX Diário processado) - Fonte Prioritária
    if opc_data:
        # Tenta nomes possíveis para a aba de ofensores da colhedora
        sheet = opc_data.get("Top5Ofensores_COLHEDORA") or opc_data.get("Top5Ofensores_COLHEDORA_DE_CANA")
        if isinstance(sheet, list) and sheet:
            for row in sheet:
                desc = str(row.get("Descrição da Operação", "")).strip()
                dur = safe_float(row.get("Duracao_Improd_h"))
                if desc and dur > 0:
                    top5_agg[desc] += dur
            if top5_agg:
                found_source = True

    # 2. Se não achou no OPC, tentar do Solinftec Geral (Legado/Fallback)
    if not found_source and solinftec_raw:
        geral = solinftec_raw.get("Geral")
        if isinstance(geral, dict):
            for item in geral.get("Top5Ofensores", []):
                desc = str(item.get("Descrição da Operação", "")).strip()
                dur = safe_float(item.get("Duracao_Improd_h"))
                if desc and dur > 0:
                    top5_agg[desc] += dur
            if top5_agg:
                found_source = True

    # Gerar lista final
    if top5_agg:
        total_improd = sum(top5_agg.values())
        ofensores_sorted = sorted(top5_agg.items(), key=lambda x: x[1], reverse=True)[:5]
        ofensores_list = [
            {
                "nome": op,
                "percentual": round((tempo / total_improd * 100) if total_improd > 0 else 0, 2),
                "duracao": round(tempo, 4),
            }
            for i, (op, tempo) in enumerate(ofensores_sorted)
        ]
    else:
        # Último caso: calcular dos intervalos (apenas se não achou nada pronto)
        # Evitar "Disponível" como ofensor
        valid_stats = {k: v for k, v in operation_stats.items() if k not in ["Disponível", "Manutenção", "Falta de Informação", ""]}
        if valid_stats:
             total_improd = sum(valid_stats.values())
             ofensores = sorted(valid_stats.items(), key=lambda x: x[1], reverse=True)[:5]
             for i, (op, tempo) in enumerate(ofensores):
                ofensores_list.append({
                    "nome": op,
                    "percentual": round((tempo / total_improd * 100) if total_improd > 0 else 0, 2),
                    "duracao": round(tempo, 4),
                })

    # ── Dados Case Extra (para seção complementar) ──
    # ── JSON Unificado ──
    resultado = {
        "metadata": {
            "date": iso_date,
            "type": "cd_diario_novo",
            "frente": "frente5",
            "generated_at": datetime.now().isoformat(),
            "fontes": fontes_usadas,
        },
        "metas": METAS_DEFAULT,
        "eficiencia_energetica": eficiencia_energetica,
        "eficiencia_operacional": eficiencia_operacional,
        "horas_elevador": horas_elevador,
        "uso_gps": uso_gps,
        "media_velocidade": media_velocidade,
        "manobras_frotas": manobras_frotas,
        "motor_ocioso": motor_ocioso,
        "disponibilidade_mecanica": disponibilidade_mecanica,
        "horas_por_frota": horas_por_frota,
        "intervalos_operacao": intervalos_operacao,
        "ofensores": ofensores_list,
        "imagens": {
            "mapaGPS": "",
            "areaTrabalhada": "",
        },
    }

    if include_case:
        dados_case = {}
        for frota_id, case_info in case_frotas.items():
            if frota_id.startswith("_"):
                continue
            dados_case[frota_id] = {
                k: v for k, v in case_info.items() if not k.startswith("_")
            }
        resultado["dados_case"] = dados_case

    return resultado


def consolidar_tratores_case(
    date_str: str,
    case_data_by_date: dict,
    solinftec_data: dict | None = None
) -> dict:
    iso_date = date_ddmmyyyy_to_iso(date_str)
    
    # Identificar frotas
    case_date_key = date_str.replace("-", "/")
    case_frotas = case_data_by_date.get(case_date_key, {})
    case_frotas_ids = set(k for k in case_frotas.keys() if not k.startswith("_"))
    
    solinftec_ids = set(solinftec_data.keys()) if solinftec_data else set()
    
    all_frotas = sorted(list(case_frotas_ids | solinftec_ids))
    
    eficiencia_energetica = []
    eficiencia_operacional = []
    uso_gps = []
    media_velocidade = []
    manobras_frotas = []
    motor_ocioso = []
    disponibilidade_mecanica = []
    horas_por_frota = []
    intervalos_operacao = []
    falta_apontamento = []
    temperaturas = []
    detalhes_motor_case = []
    detalhes_gps_case = []
    transbordo = []
    velocidades_detalhadas = []
    horas_improdutivas = []
    horas_auxiliar = []
    horas_climatico = []
    ofensores_list = []
    operation_stats = defaultdict(float)
    
    idx = 0
    fontes_usadas = []
    if case_frotas: fontes_usadas.append("case")
    if solinftec_data: fontes_usadas.append("solinftec")

    for frota_id in all_frotas:
        idx += 1
        
        # Dados Solinftec
        sol_data = solinftec_data.get(frota_id, {}) if solinftec_data else {}
        resumo_list = sol_data.get("Resumo_Dia", [])
        resumo = resumo_list[0] if resumo_list else {}
        
        # Dados Case
        case_info = case_frotas.get(frota_id, {})
        case_extra = case_info.get("Extras", {})
        
        # Prioridade Solinftec > Case
        fonte_atual = "solinftec" if sol_data else ("case" if case_info else "none")
        has_solinftec = bool(sol_data)
        has_case = bool(case_info)

        # --- Métricas ---
        
        # Eficiência Energética
        val_ee = safe_float(resumo.get("Eficiencia_Energetica"))
        if not val_ee and case_info:
            val_ee = safe_float(case_extra.get("eficienciaEnergetica", 0))
        val_ee_pct = round(val_ee * 100, 2) if val_ee <= 1 else round(val_ee, 2)
        
        horas_motor = safe_float(resumo.get("Horas_Motor_Ligado"))
        if not horas_motor and case_info:
             horas_motor = safe_float(case_info.get("Horas Motor", 0))

        horas_produtivas = safe_float(resumo.get("Horas_Produtivas", 0))
        
        eficiencia_energetica.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": val_ee_pct,
            "horasMotor": round(horas_motor, 4),
            "fonte": fonte_atual,
        })
        
        # Eficiência Operacional
        val_eo = safe_float(resumo.get("Eficiencia_Operacional"))
        eficiencia_operacional.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": round(val_eo * 100, 2) if val_eo <= 1 else round(val_eo, 2), # Se vier 0.15 virar 15%?
            "horasMotor": round(horas_motor, 4),
            "fonte": fonte_atual,
        })
        
        # Uso GPS (Case only usually)
        val_gps = safe_float(case_extra.get("usoGPS", 0))
        uso_gps.append({
            "id": idx,
            "nome": frota_id,
            "porcentagem": round(val_gps, 2),
            "fonte": "case" if val_gps > 0 else fonte_atual,
        })
        
        # Velocidade Média
        val_vel = safe_float(resumo.get("Velocidade_Media"))
        if not val_vel and case_info:
            val_vel = safe_float(case_extra.get("velocidadeMedia", 0))
            
        media_velocidade.append({
            "id": idx,
            "nome": frota_id,
            "velocidade": round(val_vel, 2),
            "fonte": fonte_atual,
        })
        
        # Motor Ocioso
        val_ocioso_pct = safe_float(resumo.get("Porcentagem_Motor_Ocioso"))
        val_ocioso_horas = safe_float(resumo.get("Horas_Motor_Ocioso"))
        motor_ocioso.append({
            "id": idx,
            "nome": frota_id,
            "percentual": round(val_ocioso_pct, 2),
            "tempoLigado": round(horas_motor, 4),
            "tempoOcioso": round(val_ocioso_horas, 4),
            "fonte": fonte_atual,
        })
        
        # Disponibilidade Mecânica
        val_disp = safe_float(resumo.get("Disponibilidade_Mecanica"))
        horas_manut = safe_float(resumo.get("Horas_Manutencao"))
        disponibilidade_mecanica.append({
            "id": idx,
            "nome": frota_id,
            "disponibilidade": round(val_disp * 100, 2) if val_disp <= 1 else round(val_disp, 2),
            "horasMotor": round(horas_motor, 4),
            "tempoManutencao": round(horas_manut, 4),
            "fonte": fonte_atual,
        })
        
        # Horas Totais (Registradas)
        if has_solinftec:
            horas_total = safe_float(resumo.get("Horas_Registradas"))
        elif has_case:
            horas_total = safe_float(case_info.get("Horas Motor", case_info.get("horasMotor", 0)))
        else:
            horas_total = 0
            
        horas_por_frota.append({
            "id": idx,
            "nome": frota_id,
            "frota": frota_id,
            "horas": round(horas_total, 2),
            "fonte": fonte_atual,
        })
        
        # Manobras
        qtd_manobras = safe_float(resumo.get("Quantidade_Manobras", 0)) if has_solinftec else 0
        tempo_manobras_h = safe_float(resumo.get("Tempo_Total_Manobras_h", 0)) if has_solinftec else 0
        tempo_medio_man_min = safe_float(resumo.get("Tempo_Medio_Manobras_min", 0)) if has_solinftec else 0
        
        # Converter tempo total para hh:mm:ss
        total_sec = int(tempo_manobras_h * 3600)
        hh = total_sec // 3600
        mm = (total_sec % 3600) // 60
        ss = total_sec % 60
        tempo_total_str = f"{hh:02}:{mm:02}:{ss:02}"
        
        # Converter tempo medio para hh:mm:ss
        med_sec = int(tempo_medio_man_min * 60)
        mhh = med_sec // 3600
        mmm = (med_sec % 3600) // 60
        mss = med_sec % 60
        tempo_med_str = f"{mhh:02}:{mmm:02}:{mss:02}"

        if qtd_manobras > 0 or has_solinftec:
            manobras_frotas.append({
                "Frota": frota_id,
                "Tempo Total": round(tempo_manobras_h, 4),
                "Tempo Médio": round(tempo_medio_man_min, 4),
                "Intervalos Válidos": int(qtd_manobras),
                "Tempo Total (hh:mm)": tempo_total_str,
                "Tempo Médio (hh:mm)": tempo_med_str,
                "fonte": fonte_atual,
            })

        # Falta de Apontamento (Sem Apontamento)
        if has_solinftec:
            pct_sem = safe_float(resumo.get("Porcentagem_Sem_Apontamento", 0))
            hrs_sem = safe_float(resumo.get("Tempo_Sem_Apontamento_h", 0))
            falta_apontamento.append({
                "id": idx,
                "nome": frota_id,
                "percentual": round(pct_sem, 2),
                "horasSemApontar": round(hrs_sem, 4),
                "tempoLigado": round(horas_motor, 4),
                "tempoOcioso": round(hrs_sem, 4), # Frontend usa isso pra barra
                "fonte": "solinftec"
            })
            
        # Transbordo
        if has_solinftec:
            qtd_transb = safe_float(resumo.get("Quantidade_Transbordos", 0))
            tempo_transb = safe_float(resumo.get("Tempo_Total_Transbordo_h", 0))
            if qtd_transb > 0 or tempo_transb > 0:
                transbordo.append({
                    "id": idx,
                    "nome": frota_id,
                    "quantidade": int(qtd_transb),
                    "tempoTotal": round(tempo_transb, 4),
                    "fonte": "solinftec"
                })

        # Velocidades Detalhadas (Tratores)
        if has_solinftec:
            vel_colheita = safe_float(resumo.get("Vel_Colheita_media", 0))
            vel_vazio = safe_float(resumo.get("Vel_Desl_Vazio_media", 0))
            vel_carregado = safe_float(resumo.get("Vel_Desl_Carregado_media", 0))
            
            if vel_colheita > 0 or vel_vazio > 0 or vel_carregado > 0:
                velocidades_detalhadas.append({
                    "id": idx,
                    "nome": frota_id,
                    "colheita": round(vel_colheita, 2),
                    "vazio": round(vel_vazio, 2),
                    "carregado": round(vel_carregado, 2),
                    "fonte": "solinftec"
                })

        # Outras horas (Improdutivas, Auxiliares, Climático)
        if has_solinftec:
            h_impro = safe_float(resumo.get("Horas_Improdutivas", 0))
            h_aux = safe_float(resumo.get("Horas_Auxiliar", 0))
            h_clim = safe_float(resumo.get("Horas_Climático", 0))
            
            if h_impro > 0:
                horas_improdutivas.append({
                    "id": idx, "nome": frota_id, "horas": round(h_impro, 4), "fonte": "solinftec"
                })
            if h_aux > 0:
                horas_auxiliar.append({
                    "id": idx, "nome": frota_id, "horas": round(h_aux, 4), "fonte": "solinftec"
                })
            if h_clim > 0:
                horas_climatico.append({
                    "id": idx, "nome": frota_id, "horas": round(h_clim, 4), "fonte": "solinftec"
                })

        # ----------------------------------------------------
        # NOVAS MÉTRICAS ESPECÍFICAS DA CASE IH
        # ----------------------------------------------------
        if case_info:
            temperaturas.append({
                "id": idx,
                "nome": frota_id,
                "arrefecimento": safe_float(case_info.get("temperaturaArrefecimento", 0)),
                "oleoHidraulico": safe_float(case_info.get("temperaturaTransmissao", 0)),
                "arAdmissao": safe_float(case_info.get("temperaturaArAdmissao", 0)),
                "fonte": "case"
            })
            
            detalhes_motor_case.append({
                "id": idx,
                "nome": frota_id,
                "motorOcioso": safe_float(case_info.get("motorOcioso", 0)),
                "motorDesligado": safe_float(case_info.get("motorDesligado", 0)),
                "horasProdutivas": safe_float(case_info.get("horasProdutivas", 0)),
                "tempoRegistrado": safe_float(case_info.get("tempoRegistrado", 0)),
                "horasMotor": safe_float(case_info.get("Horas Motor", case_info.get("horasMotor", 0))),
                "fonte": "case"
            })
            
            det_extras = case_info.get("Extras", {})
            detalhes_gps_case.append({
                "id": idx,
                "nome": frota_id,
                "tempoLigado": safe_float(det_extras.get("tempoGPSLigado", 0)),
                "tempoDesligado": safe_float(det_extras.get("tempoGPSDesligado", 0)),
                "percLigado": safe_float(det_extras.get("percGPSLigado", 0)),
                "percDesligado": safe_float(det_extras.get("percGPSDesligado", 0)),
                "fonte": "case"
            })
        # --- Intervalos ---
        # Solinftec Intervalos
        intervalos = sol_data.get("Intervalos", [])
        if intervalos:
            for interval in intervalos:
                grupo = interval.get("Grupo", "")
                operacao = interval.get("Descrição da Operação", "")
                inicio = interval.get("Início", "")
                fim = interval.get("Fim", "")
                
                # Calcular duracao
                dur_h = calc_duration_hours(inicio, fim)
                
                tipo = "Disponível"
                if "PRODUTIVA" in grupo.upper():
                    tipo = "Produtivo"
                elif "MANUTENÇÃO" in grupo.upper() or "MANUTENCAO" in grupo.upper():
                    tipo = "Manutenção"
                
                intervalos_operacao.append({
                    "equipamento": frota_id,
                    "tipo": tipo,
                    "inicio": inicio,
                    "duracaoHoras": round(dur_h, 6),
                    "fonte": "solinftec",
                })
                
                # Ofensores (Improdutiva/Manutenção)
                if "IMPRODUTIVA" in grupo.upper() or "MANUTENÇÃO" in grupo.upper() or "DISPONIVEL" in grupo.upper():
                    # Solinftec 'DISPONIVEL' often includes idle time or waits, check if it should count as offensor
                    # Usually "FALTA DE ..." is an offensor.
                    # For now, let's include everything not PRODUTIVA as potential offensor?
                    # Or just IMPRODUTIVA/MANUTENCAO?
                    # In Solinftec, 'DISPONIVEL' with 'FALTA DE...' is bad.
                    # Let's count everything that is NOT Produtiva.
                    if "PRODUTIVA" not in grupo.upper():
                        operation_stats[operacao] += dur_h

        elif case_info:
             # Fallback Case Intervalos
            case_intervals = case_extra.get("_intervalos", [])
            for ci in case_intervals:
                dur = safe_float(ci.get("duracao", 0))
                grupo = ci.get("grupo", "")
                operacao = ci.get("operacao", "")

                tipo = "Disponível"
                if "PRODUTIVA" in grupo.upper():
                    tipo = "Produtivo"
                elif "MANUTENÇÃO" in grupo.upper() or "MANUTENCAO" in grupo.upper():
                    tipo = "Manutenção"

                intervalos_operacao.append({
                    "equipamento": frota_id,
                    "tipo": tipo,
                    "inicio": ci.get("inicio", ""),
                    "duracaoHoras": round(dur, 6),
                    "fonte": "case",
                })

                if "IMPRODUTIVA" in grupo.upper() or "MANUTENÇÃO" in grupo.upper():
                    operation_stats[operacao] += dur

    # Consolidar Ofensores Globais
    total_improd = sum(operation_stats.values())
    ofensores = sorted(operation_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    ofensores_list = []
    for i, (op, tempo) in enumerate(ofensores):
        ofensores_list.append({
            "id": str(i),
            "tempo": round(tempo, 4),
            "operacao": op,
            "porcentagem": round((tempo / total_improd * 100) if total_improd > 0 else 0, 2),
        })
    
    dados_case = {}
    for frota_id, case_info in case_frotas.items():
        if frota_id.startswith("_"):
            continue
        dados_case[frota_id] = {
            k: v for k, v in case_info.items() if not k.startswith("_")
        }

    resultado = {
        "metadata": {
            "date": iso_date,
            "type": "tt_diario_novo",
            "frente": "frente5", # TODO: Detectar frente se possível
            "generated_at": datetime.now().isoformat(),
            "fontes": fontes_usadas,
        },
        "metas": METAS_DEFAULT,
        "eficiencia_energetica": eficiencia_energetica,
        "eficiencia_operacional": eficiencia_operacional,
        "uso_gps": uso_gps,
        "media_velocidade": media_velocidade,
        "manobras_frotas": manobras_frotas,
        "motor_ocioso": motor_ocioso,
        "disponibilidade_mecanica": disponibilidade_mecanica,
        "horas_por_frota": horas_por_frota,
        "intervalos_operacao": intervalos_operacao,
        "falta_apontamento": falta_apontamento,
        "temperaturas": temperaturas,
        "detalhes_motor_case": detalhes_motor_case,
        "detalhes_gps_case": detalhes_gps_case,
        "transbordo": transbordo,
        "velocidades_detalhadas": velocidades_detalhadas,
        "horas_improdutivas": horas_improdutivas,
        "horas_auxiliar": horas_auxiliar,
        "horas_climatico": horas_climatico,
        "ofensores": ofensores_list,
        "imagens": {
            "mapaGPS": "",
            "areaTrabalhada": "",
        },
        "dados_case": dados_case,
    }

    return resultado


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  🔄 Consolidação de JSON Unificado por Dia")
    print("=" * 60)

    # 1. Listar datas disponíveis (dos JSONs Solinftec existentes)
    base_json_dir = os.path.join(ETL_ROOT, "dados", "separados", "json")
    solinftec_files = sorted(glob.glob(os.path.join(base_json_dir, "**", "*.json"), recursive=True))
    
    date_to_dir = {}
    dates = []
    for sf in solinftec_files:
        d = parse_date_from_filename(os.path.basename(sf))
        if d:
            if d not in dates:
                dates.append(d)
            # Guarda a pasta base de onde achou esse arquivo para usar no load_solinftec
            date_to_dir[d] = os.path.dirname(sf)

    if not dates:
        print("  ❌ Nenhum JSON Solinftec encontrado.")
        return

    # 1.1 Respeitar intervalo configurado em utils/config_automacao.json (se existir)
    cfg_path = os.path.join(ETL_ROOT, "utils", "config_automacao.json")
    allowed_range = None
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            params = cfg.get("automacao", {}).get("parametros", {})
            str_ini = params.get("data_inicial") or params.get("data_inicio")
            str_fim = params.get("data_final") or params.get("data_fim")
            if str_ini and str_fim:
                dt_ini = datetime.strptime(str_ini, "%d/%m/%Y").date()
                dt_fim = datetime.strptime(str_fim, "%d/%m/%Y").date()
                allowed_range = (dt_ini, dt_fim)
                filtered = []
                for d in dates:
                    # d está em DD-MM-YYYY
                    try:
                        dd, mm, yyyy = d.split("-")
                        dt = datetime(int(yyyy), int(mm), int(dd)).date()
                        if dt_ini <= dt <= dt_fim:
                            filtered.append(d)
                    except Exception:
                        continue
                dates = filtered
        except Exception as e:
            print(f"  ⚠️ Erro ao aplicar filtro de datas do config_automacao.json: {e}")

    if not dates:
        print("  ❌ Nenhuma data dentro do intervalo configurado.")
        return

    print(f"\n  📅 Datas encontradas (após filtro): {dates}")

    # 2. Carregar dados Case (uma vez para todas as datas)
    print("\n  📦 Carregando dados Case IH...")
    case_data = load_case_data()
    print(f"     Datas Case disponíveis: {[k for k in case_data.keys() if not k.startswith('_')]}")

    # 3. FASE 1: Carregar TODOS os dados brutos Solinftec ANTES de escrever qualquer saída
    #    Isso evita o problema de sobrescrever o arquivo fonte com o consolidado
    print(f"\n  📥 FASE 1: Carregando dados brutos Solinftec...")
    raw_data_cache = {}  # { date_str: { 'colhedora': dict|None, 'tratores': dict|None } }
    
    for date_str in dates:
        solinftec_colhedora = load_solinftec(date_str, base_dir=SOLINFTEC_JSON_DIR)
        solinftec_tratores = load_solinftec(date_str, base_dir=TRATORES_JSON_DIR)
        raw_data_cache[date_str] = {
            'colhedora': solinftec_colhedora,
            'tratores': solinftec_tratores,
        }
        n_col = len(solinftec_colhedora) if solinftec_colhedora else 0
        n_trt = len(solinftec_tratores) if solinftec_tratores else 0
        print(f"     {date_str}: Colhedoras={n_col} frotas, Tratores={n_trt} frotas")

    # 4. FASE 2: Processar e salvar consolidados
    print(f"\n  📤 FASE 2: Consolidando e salvando...")
    for date_str in dates:
        print(f"\n  {'─' * 50}")
        print(f"  📅 Processando {date_str}...")
        try:
            dd, mm, yyyy = date_str.split("-")
            dt = datetime(int(yyyy), int(mm), int(dd)).date()
            if allowed_range and not (allowed_range[0] <= dt <= allowed_range[1]):
                print(f"     ⚠️ Pulando data fora do intervalo configurado: {date_str}")
                continue
        except Exception:
            pass

        # Usar dados já carregados do cache (não do disco!)
        cached = raw_data_cache.get(date_str, {})
        solinftec_raw = cached.get('colhedora') or {}
            
        n_frotas_sol = len(solinftec_raw)
        print(f"     Solinftec (Colhedoras): {n_frotas_sol} frotas")

        # Case
        case_date_key = date_str.replace("-", "/")
        case_frotas = case_data.get(case_date_key, {})
        n_frotas_case = len([k for k in case_frotas if not k.startswith("_")])
        print(f"     Case IH: {n_frotas_case} frotas")

        # OPC
        opc_data = load_opc_daily(date_str)
        if opc_data:
            print(f"     OPC: {list(opc_data.keys())}")
        else:
            print(f"     OPC: sem dados")

        # Consolidar Colhedoras
        resultado = consolidar_dia(date_str, solinftec_raw, case_data, opc_data, False)

        # Salvar
        output_name = f"colhedora_frota_{date_str}.json"
        output_path = os.path.join(OUTPUT_DIR, output_name)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        n_total = len(resultado.get("eficiencia_energetica", []))
        n_intervalos = len(resultado.get("intervalos_operacao", []))
        n_ofensores = len(resultado.get("ofensores", []))
        print(f"     ✅ Salvo: {output_name}")
        print(f"        {n_total} frotas, {n_intervalos} intervalos, {n_ofensores} ofensores")
        print(f"        Fontes: {resultado['metadata']['fontes']}")

        # Tratores (usando dados do cache)
        solinftec_tratores = cached.get('tratores')
        resultado_tratores = consolidar_tratores_case(date_str, case_data, solinftec_tratores)
        
        os.makedirs(TRATORES_JSON_DIR, exist_ok=True)
        output_name_tratores = f"tratores_frota_{date_str}.json"
        output_path_tratores = os.path.join(TRATORES_JSON_DIR, output_name_tratores)
        with open(output_path_tratores, "w", encoding="utf-8") as f:
            json.dump(resultado_tratores, f, ensure_ascii=False, indent=2)

        n_total_tratores = len(resultado_tratores.get("eficiencia_energetica", []))
        n_intervalos_tratores = len(resultado_tratores.get("intervalos_operacao", []))
        n_ofensores_tratores = len(resultado_tratores.get("ofensores", []))
        print(f"     ✅ Salvo: {output_name_tratores}")
        print(f"        {n_total_tratores} frotas, {n_intervalos_tratores} intervalos, {n_ofensores_tratores} ofensores")
        print(f"        Fontes: {resultado_tratores['metadata']['fontes']}")

    print(f"\n{'=' * 60}")
    print(f"  ✅ Consolidação concluída! {len(dates)} dias processados.")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
