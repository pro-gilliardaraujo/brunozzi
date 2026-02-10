#!/usr/bin/env python3
"""Script para mapear headers de todas as fontes de dados."""
import json, os, glob
import openpyxl

SCRIPTS = os.path.dirname(os.path.abspath(__file__))
ETL_ROOT = os.path.dirname(SCRIPTS)  # automacao_etl/

# 1. Case IH - Consolidado
print("=" * 70)
print("CASE IH - Consolidado")
print("=" * 70)
case_path = os.path.join(ETL_ROOT, "dados", "Consolidado_Case_05_11-10-2025.xlsx")
print(f"Path: {case_path}")
print(f"Exists: {os.path.exists(case_path)}")
wb = openpyxl.load_workbook(case_path, read_only=True)
print(f"Abas: {wb.sheetnames}")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=4))
    if not rows:
        print(f"\n  [{sheet_name}] VAZIA")
        continue
    headers = [c.value for c in rows[0]]
    print(f"\n  [{sheet_name}] Headers ({len(headers)} cols):")
    for h in headers:
        print(f"    - {h}")
    if len(rows) > 1:
        vals = {str(headers[j]): rows[1][j].value for j in range(min(len(headers), len(rows[1])))}
        print(f"    Amostra: {vals}")
wb.close()

# 2. XLSX Diário (separados/xlsx)
print("\n" + "=" * 70)
print("XLSX DIÁRIO (separados/xlsx/05-10-2025.xlsx)")
print("=" * 70)
xlsx_path = os.path.join(SCRIPTS, "dados", "separados", "xlsx", "05-10-2025.xlsx")
wb2 = openpyxl.load_workbook(xlsx_path, read_only=True)
print(f"Abas: {wb2.sheetnames}")
for sheet_name in wb2.sheetnames[:5]:
    ws = wb2[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=2))
    if not rows:
        continue
    headers = [c.value for c in rows[0]]
    print(f"\n  [{sheet_name}] Headers ({len(headers)} cols):")
    for h in headers[:20]:
        print(f"    - {h}")
    if len(headers) > 20:
        print(f"    ...+{len(headers)-20} mais")
wb2.close()

# 3. Solinftec JSON - Resumo_Dia completo
print("\n" + "=" * 70)
print("SOLINFTEC JSON - colhedora_frota_05-10-2025.json")
print("=" * 70)
json_path = os.path.join(SCRIPTS, "dados", "separados", "json", "colhedora", "frotas", "diario", "colhedora_frota_05-10-2025.json")
with open(json_path, "r", encoding="utf-8") as f:
    d = json.load(f)
frotas = list(d.keys())
print(f"Frotas: {frotas}")
first = d[frotas[0]]
print(f"Seções por frota: {list(first.keys())}")

for k in first.keys():
    v = first[k]
    if isinstance(v, list):
        print(f"\n  [{k}] ({len(v)} items)")
        if v and isinstance(v[0], dict):
            print(f"    Campos: {list(v[0].keys())}")
    elif isinstance(v, dict):
        print(f"\n  [{k}] (dict, {len(v)} keys): {list(v.keys())[:10]}")
    else:
        print(f"\n  [{k}] = {v}")

# 4. Verificar dados OPC
print("\n" + "=" * 70)
print("DADOS OPC / JOHN DEERE")
print("=" * 70)
opc_files = []
for root, dirs, files in os.walk(ETL_ROOT):
    for fn in files:
        if "opc" in fn.lower() or "john" in fn.lower() or "deere" in fn.lower():
            opc_files.append(os.path.join(root, fn))
if opc_files:
    for p in opc_files:
        print(f"  {os.path.relpath(p, ETL_ROOT)} ({os.path.getsize(p):,} bytes)")
else:
    print("  Nenhum arquivo OPC encontrado")

# 5. Verificar Linha do Tempo
print("\n" + "=" * 70)  
print("LINHA DO TEMPO (tratado)")
print("=" * 70)
lt_path = os.path.join(SCRIPTS, "dados", "Linha_do_tempo-05-10-2025_11-10-2025_tratado.xlsx")
wb3 = openpyxl.load_workbook(lt_path, read_only=True)
print(f"Abas: {wb3.sheetnames}")
for sheet_name in wb3.sheetnames[:3]:
    ws = wb3[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=2))
    if not rows:
        continue
    headers = [c.value for c in rows[0]]
    print(f"\n  [{sheet_name}] Headers ({len(headers)} cols):")
    for h in headers[:20]:
        print(f"    - {h}")
    if len(headers) > 20:
        print(f"    ...+{len(headers)-20} mais")
wb3.close()

# 6. Frontend - o que o componente espera
print("\n" + "=" * 70)
print("FRONTEND - Labels/campos que o relatório espera (page.tsx)")
print("=" * 70)
print("  Baseado na análise do componente cd-diario-frotas/page.tsx:")
print("  Seções de dados esperadas por frota:")
print("    - Eficiência Energética (Eficiencia_Energetica, meta)")
print("    - Eficiência Operacional (Eficiencia_Operacional, meta)")
print("    - Horas Elevador")
print("    - Uso GPS")
print("    - Mapa GPS / Área Trabalhada")
print("    - Média de Velocidade (Vel_Colheita_media)")
print("    - Manobras (Quantidade_Manobras, Tempo_Medio_Manobras_min)")
print("    - Lavagem")
print("    - Roletes")
print("    - Motor Ocioso (Porcentagem_Motor_Ocioso)")
print("    - Top 5 Ofensores")
print("    - Disponibilidade Mecânica (Disponibilidade_Mecanica)")
print("    - Intervalos de Operação (Início, Fim, Grupo, Descrição)")
print("    - Resumo (cards + tabela)")
