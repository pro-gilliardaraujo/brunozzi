#!/usr/bin/env python3
"""
6_ConsolidarJSON.py

Consolida dados de múltiplas fontes (Solinftec, Case IH, OPC) 
em um único JSON por dia, no formato que o frontend espera.

Fontes:
  - Solinftec: scripts/dados/separados/json/colhedora/frotas/diario/*.json
  - Case IH:  dados/Consolidado_Case_*.xlsx
  - OPC:      scripts/dados/separados/xlsx/*.xlsx (Linha do tempo tratado)

Saída:
  - scripts/dados/separados/json/colhedora/frotas/diario/*.json (sobrescreve)
"""

import json
import os
import glob
import re
import sys
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ─── Paths ──────────────────────────────────────────────────────────────────────

SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
ETL_ROOT = os.path.dirname(SCRIPTS_DIR)

SOLINFTEC_JSON_DIR = os.path.join(
    SCRIPTS_DIR, "dados", "separados", "json", "colhedora", "frotas", "diario"
)
OPC_XLSX_DIR = os.path.join(SCRIPTS_DIR, "dados", "separados", "xlsx")
CASE_DIR = os.path.join(ETL_ROOT, "dados")
LINHA_TEMPO_DIR = os.path.join(SCRIPTS_DIR, "dados")

# Saída (mesmo diretório do Solinftec — sobrescreve com formato unificado)
OUTPUT_DIR = SOLINFTEC_JSON_DIR

# Metas default (mesmas do frontend config/metas.json)
METAS_DEFAULT = {
    "eficienciaEnergetica": 85,
    "eficienciaOperacional": 60,
    "horaElevador": 5,
    "usoGPS": 90,
    "mediaVelocidade": 5,
    "manobras": 60,
    "producao": 1000,
    "disponibilidadeMecanica": 90,
    "motorOcioso": 15,
}


# ─── Helpers ────────────────────────────────────────────────────────────────────

def parse_date_from_filename(filename: str) -> str | None:
    """Extrai data DD-MM-YYYY do nome do arquivo."""
    m = re.search(r"(\d{2}-\d{2}-\d{4})", filename)
    return m.group(1) if m else None


def date_ddmmyyyy_to_iso(date_str: str) -> str:
    """Converte DD-MM-YYYY para YYYY-MM-DD."""
    parts = date_str.split("-")
    return f"{parts[2]}-{parts[1]}-{parts[0]}"


def time_hhmmss(dt_str: str) -> str:
    """Extrai HH:MM:SS de uma string 'DD/MM/YYYY HH:MM:SS'."""
    parts = dt_str.strip().split(" ")
    return parts[1] if len(parts) > 1 else dt_str


def parse_datetime(dt_str: str) -> datetime | None:
    """Parse 'DD/MM/YYYY HH:MM:SS' para datetime."""
    try:
        return datetime.strptime(dt_str.strip(), "%d/%m/%Y %H:%M:%S")
    except (ValueError, AttributeError):
        return None


def calc_duration_hours(start_str: str, end_str: str) -> float:
    """Calcula duração em horas entre duas strings de data."""
    s = parse_datetime(start_str)
    e = parse_datetime(end_str)
    if s and e:
        return (e - s).total_seconds() / 3600
    return 0.0


def safe_float(val, default=0.0) -> float:
    """Convert to float safely."""
    try:
        return float(val) if val is not None else default
    except (ValueError, TypeError):
        return default


# ─── Solinftec ──────────────────────────────────────────────────────────────────

def load_solinftec(date_str: str) -> dict | None:
    """Carrega JSON Solinftec bruto para a data (DD-MM-YYYY)."""
    pattern = os.path.join(SOLINFTEC_JSON_DIR, f"*{date_str}*.json")
    files = glob.glob(pattern)
    if not files:
        return None
    with open(files[0], "r", encoding="utf-8") as f:
        return json.load(f)


# ─── Case IH ───────────────────────────────────────────────────────────────────

def load_case_data() -> dict:
    """
    Carrega todos os dados Case IH dos XLSX consolidados.
    Retorna dict: { 'DD/MM/YYYY': { 'frota_id': { ...campos... } } }
    """
    case_files = glob.glob(os.path.join(CASE_DIR, "Consolidado_Case_*.xlsx"))
    if not case_files:
        print("  ⚠️  Nenhum arquivo Case IH encontrado.")
        return {}

    case_data = defaultdict(lambda: defaultdict(dict))

    for cf in case_files:
        print(f"  📂 Carregando Case: {os.path.basename(cf)}")
        wb = openpyxl.load_workbook(cf, read_only=True)

        # Aba "Resumo" contém dados por frota para o período inteiro
        if "Resumo" in wb.sheetnames:
            ws = wb["Resumo"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    if frota:
                        case_data["_resumo_geral"][frota] = {
                            "horasMotor": safe_float(d.get("Total Horas Motor (Diferença)")),
                            "rpm": safe_float(d.get("RPM")),
                            "temperaturaArrefecimento": safe_float(d.get("Média Temperatura líquido de arrefecimento do motor")),
                            "temperaturaTransmissao": safe_float(d.get("Média Temperatura do óleo da transmissão")),
                            "velocidadeMedia": safe_float(d.get("Velocidade Média")),
                        }

        # Aba "Resumo Diário" contém dados por frota POR DIA
        if "Resumo Diário" in wb.sheetnames:
            ws = wb["Resumo Diário"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    data_val = d.get("Data", "")
                    
                    # Normalizar data
                    if isinstance(data_val, datetime):
                        data_key = data_val.strftime("%d/%m/%Y")
                    elif isinstance(data_val, str):
                        data_key = data_val.split(" ")[0] if " " in data_val else data_val
                    else:
                        continue

                    if frota and data_key:
                        case_data[data_key][frota] = {
                            "horasMotor": safe_float(d.get("Total Horas Motor (Diferença)")),
                            "rpm": safe_float(d.get("RPM")),
                            "temperaturaArrefecimento": safe_float(d.get("Média Temperatura líquido de arrefecimento do motor")),
                            "temperaturaTransmissao": safe_float(d.get("Média Temperatura do óleo da transmissão")),
                            "velocidadeMedia": safe_float(d.get("Velocidade Média")),
                        }

        # Aba "Dados" contém intervalos detalhados com coordenadas
        if "Dados" in wb.sheetnames:
            ws = wb["Dados"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else "" for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    frota = str(d.get("Frota", "")).strip()
                    data_hora = d.get("Data Hora Local", "")
                    
                    if isinstance(data_hora, datetime):
                        data_key = data_hora.strftime("%d/%m/%Y")
                    elif isinstance(data_hora, str):
                        data_key = data_hora.split(" ")[0] if " " in data_hora else ""
                    else:
                        continue

                    if frota and data_key:
                        if "_intervalos" not in case_data[data_key].get(frota, {}):
                            if frota not in case_data[data_key]:
                                case_data[data_key][frota] = {}
                            case_data[data_key][frota].setdefault("_intervalos", [])
                        
                        case_data[data_key][frota]["_intervalos"].append({
                            "inicio": str(data_hora),
                            "duracao": safe_float(d.get("Duração")),
                            "operacao": str(d.get("Descrição da Operação", "")),
                            "grupo": str(d.get("Descrição do Grupo da Operação", "")),
                            "lat": safe_float(d.get("Latitude")),
                            "lon": safe_float(d.get("Longitude")),
                        })

        wb.close()

    return dict(case_data)


# ─── OPC (XLSX diário) ─────────────────────────────────────────────────────────

def load_opc_daily(date_str: str) -> dict | None:
    """
    Carrega dados do XLSX diário OPC para a data (DD-MM-YYYY).
    As abas relevantes são COLHEDORA_Dia, TRANSBORDO_Dia, GRUNNER_Dia.
    """
    xlsx_path = os.path.join(OPC_XLSX_DIR, f"{date_str}.xlsx")
    if not os.path.exists(xlsx_path):
        return None

    print(f"  📂 Carregando OPC: {date_str}.xlsx")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    opc = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        sheet_data = []
        for row in rows[1:]:
            d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            sheet_data.append(d)

        opc[sheet_name] = sheet_data

    wb.close()
    return opc


# ─── Consolidação ──────────────────────────────────────────────────────────────

def consolidar_dia(
    date_str: str,
    solinftec_raw: dict,
    case_data_by_date: dict,
    opc_data: dict | None,
) -> dict:
    """
    Consolida dados de um dia no formato do frontend.
    
    date_str: DD-MM-YYYY
    solinftec_raw: dict com frotas como chaves, cada uma com Resumo_Dia e Intervalos
    case_data_by_date: dados Case para a data (dict com frota como chave)
    opc_data: dados OPC para a data
    """
    iso_date = date_ddmmyyyy_to_iso(date_str)
    date_display = date_str.replace("-", "/")

    # Normalizar data para case lookup (DD/MM/YYYY)
    case_date_key = date_display
    case_frotas = case_data_by_date.get(case_date_key, {})

    # Coletar todas as frotas de todas as fontes
    all_frotas = set()
    if solinftec_raw:
        all_frotas.update(solinftec_raw.keys())
    all_frotas.update(k for k in case_frotas.keys() if not k.startswith("_"))

    fontes_usadas = []
    if solinftec_raw:
        fontes_usadas.append("solinftec")
    if case_frotas:
        fontes_usadas.append("case")
    if opc_data:
        fontes_usadas.append("opc")

    # Arrays para cada seção do frontend
    eficiencia_energetica = []
    eficiencia_operacional = []
    horas_elevador = []
    uso_gps = []
    media_velocidade = []
    manobras_frotas = []
    motor_ocioso = []
    disponibilidade_mecanica = []
    intervalos_operacao = []
    horas_por_frota = []
    producao_por_frota = []

    operation_stats = defaultdict(float)  # Para ofensores
    total_producao = 0.0
    idx = 0

    for frota_id in sorted(all_frotas):
        idx += 1
        resumo = None
        intervalos = []
        fonte = "desconhecida"
        case_extra = case_frotas.get(frota_id, {})

        # Dados Solinftec
        if solinftec_raw and frota_id in solinftec_raw:
            frota_data = solinftec_raw[frota_id]
            resumo = frota_data.get("Resumo_Dia", [{}])[0] if frota_data.get("Resumo_Dia") else {}
            intervalos = frota_data.get("Intervalos", [])
            fonte = "solinftec"

        # ── Eficiência Energética ──
        horas_produtivas = safe_float(resumo.get("Horas_Produtivas")) if resumo else 0
        horas_motor = safe_float(resumo.get("Horas_Motor_Ligado")) if resumo else safe_float(case_extra.get("horasMotor"))
        
        ef_energetica = 0.0
        if resumo and "Eficiencia_Energetica" in resumo:
            ef_energetica = safe_float(resumo["Eficiencia_Energetica"]) * 100
        elif horas_motor > 0 and horas_produtivas > 0:
            ef_energetica = (horas_produtivas / horas_motor) * 100

        eficiencia_energetica.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": round(ef_energetica, 2),
            "horasMotor": round(horas_motor, 4),
            "horasElevador": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Eficiência Operacional ──
        horas_registradas = safe_float(resumo.get("Horas_Registradas")) if resumo else 0

        ef_operacional = 0.0
        if resumo and "Eficiencia_Operacional" in resumo:
            ef_operacional = safe_float(resumo["Eficiencia_Operacional"]) * 100
        elif horas_registradas > 0 and horas_produtivas > 0:
            ef_operacional = (horas_produtivas / horas_registradas) * 100

        eficiencia_operacional.append({
            "id": idx,
            "nome": frota_id,
            "eficiencia": round(ef_operacional, 2),
            "horasMotor": round(horas_registradas, 4),
            "horasElevador": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Horas Elevador ──
        horas_elevador.append({
            "id": idx,
            "nome": frota_id,
            "valor": round(horas_produtivas, 4),
            "fonte": fonte,
        })

        # ── Uso GPS ──
        uso_gps_val = 0.0
        if resumo and "Porcentagem_Sem_Apontamento" in resumo:
            uso_gps_val = 100 - safe_float(resumo["Porcentagem_Sem_Apontamento"])
        uso_gps.append({
            "id": idx,
            "nome": frota_id,
            "porcentagem": round(uso_gps_val, 2),
            "fonte": fonte,
        })

        # ── Média de Velocidade ──
        vel = 0.0
        if resumo and "Vel_Colheita_media" in resumo:
            vel = safe_float(resumo["Vel_Colheita_media"])
        elif case_extra.get("velocidadeMedia"):
            vel = safe_float(case_extra["velocidadeMedia"])
            fonte = "case"
        media_velocidade.append({
            "id": idx,
            "nome": frota_id,
            "velocidade": round(vel, 4),
            "fonte": fonte if vel > 0 else "solinftec",
        })

        # ── Motor Ocioso ──
        motor_ocioso_pct = safe_float(resumo.get("Porcentagem_Motor_Ocioso")) if resumo else 0
        tempo_ocioso = safe_float(resumo.get("Horas_Motor_Ocioso")) if resumo else 0
        motor_ocioso.append({
            "id": idx,
            "nome": frota_id,
            "percentual": round(motor_ocioso_pct, 4),
            "tempoLigado": round(horas_motor, 4),
            "tempoOcioso": round(tempo_ocioso, 4),
            "fonte": fonte,
        })

        # ── Disponibilidade Mecânica ──
        disp_mec = 100.0
        horas_manutencao = 0.0
        if resumo and "Disponibilidade_Mecanica" in resumo:
            disp_mec = safe_float(resumo["Disponibilidade_Mecanica"]) * 100
            horas_manutencao = safe_float(resumo.get("Horas_Manutencao", 0))
        disponibilidade_mecanica.append({
            "id": idx,
            "nome": frota_id,
            "disponibilidade": round(disp_mec, 2),
            "horasMotor": round(horas_motor, 4),
            "tempoManutencao": round(horas_manutencao, 4),
            "fonte": fonte,
        })

        # ── Horas por Frota ──
        horas_por_frota.append({
            "id": idx,
            "nome": frota_id,
            "frota": frota_id,
            "horas": round(horas_registradas, 2),
            "fonte": fonte,
        })

        # ── Produção por Frota ──
        prod_frota = safe_float(resumo.get("Frotas_no_dia", 0)) if resumo else 0
        producao_por_frota.append({
            "id": idx,
            "nome": frota_id,
            "valor": round(prod_frota, 2),
            "fonte": fonte,
        })
        total_producao += prod_frota

        # ── Manobras ──
        qtd_manobras = 0
        tempo_total_manobras = 0.0
        tempo_medio_manobras = 0.0

        if resumo:
            qtd_manobras = int(safe_float(resumo.get("Quantidade_Manobras", 0)))
            tempo_total_manobras = safe_float(resumo.get("Tempo_Total_Manobras_h", 0))
            tempo_medio_manobras = safe_float(resumo.get("Tempo_Medio_Manobras_min", 0))

        # Calcular hh:mm do tempo total e médio
        total_h = int(tempo_total_manobras)
        total_m = int((tempo_total_manobras - total_h) * 60)
        total_s = int(((tempo_total_manobras - total_h) * 60 - total_m) * 60)
        medio_m = int(tempo_medio_manobras)
        medio_s = int((tempo_medio_manobras - medio_m) * 60)

        manobras_frotas.append({
            "Frota": frota_id,
            "Tempo Total": round(tempo_total_manobras, 4),
            "Tempo Médio": round(tempo_medio_manobras / 60 if tempo_medio_manobras > 0 else 0, 6),
            "Intervalos Válidos": qtd_manobras,
            "Tempo Total (hh:mm)": f"{total_h:02d}:{total_m:02d}:{total_s:02d}",
            "Tempo Médio (hh:mm)": f"00:{medio_m:02d}:{medio_s:02d}",
            "fonte": fonte,
        })

        # ── Intervalos de Operação (Gantt) ──
        for intv in intervalos:
            start_str = intv.get("Início", "")
            end_str = intv.get("Fim", "")
            dur = calc_duration_hours(start_str, end_str)

            grupo = intv.get("Grupo", "")
            descricao = intv.get("Descrição da Operação", "")

            tipo = "Disponível"
            if grupo == "PRODUTIVA":
                tipo = "Produtivo"
            elif grupo == "MANUTENÇÃO":
                tipo = "Manutenção"
            if descricao == "SEM APONTAMENTO":
                tipo = "Falta de Informação"

            intervalos_operacao.append({
                "equipamento": frota_id,
                "tipo": tipo,
                "inicio": time_hhmmss(start_str),
                "duracaoHoras": round(dur, 6),
                "fonte": "solinftec",
            })

            # Agregar ofensores
            if grupo in ("IMPRODUTIVA", "MANUTENÇÃO"):
                operation_stats[descricao] += dur

        # Intervalos Case (se houver)
        case_intervals = case_extra.get("_intervalos", [])
        for ci in case_intervals:
            dur = safe_float(ci.get("duracao", 0))
            grupo = ci.get("grupo", "")
            operacao = ci.get("operacao", "")
            
            tipo = "Disponível"
            if "PRODUTIVA" in grupo.upper():
                tipo = "Produtivo"
            elif "MANUTENÇÃO" in grupo.upper() or "MANUTENCAO" in grupo.upper():
                tipo = "Manutenção"

            intervalos_operacao.append({
                "equipamento": frota_id,
                "tipo": tipo,
                "inicio": ci.get("inicio", ""),
                "duracaoHoras": round(dur, 6),
                "fonte": "case",
            })

            if "IMPRODUTIVA" in grupo.upper() or "MANUTENÇÃO" in grupo.upper():
                operation_stats[operacao] += dur

    # ── Ofensores (Top 5) ──
    total_improd = sum(operation_stats.values())
    ofensores = sorted(operation_stats.items(), key=lambda x: x[1], reverse=True)[:5]
    ofensores_list = []
    for i, (op, tempo) in enumerate(ofensores):
        ofensores_list.append({
            "id": str(i),
            "tempo": round(tempo, 4),
            "operacao": op,
            "porcentagem": round((tempo / total_improd * 100) if total_improd > 0 else 0, 2),
        })

    # ── Lavagem e Roletes (dos Intervalos Solinftec) ──
    lavagem = []
    roletes = []
    if solinftec_raw:
        for frota_id, frota_data in solinftec_raw.items():
            for intv in frota_data.get("Intervalos", []):
                descricao = intv.get("Descrição da Operação", "").upper()
                if "LAVAGEM" in descricao:
                    dur = calc_duration_hours(intv["Início"], intv["Fim"])
                    lavagem.append({
                        "Data": date_display,
                        "Equipamento": frota_id,
                        "Início": time_hhmmss(intv["Início"]),
                        "Fim": time_hhmmss(intv["Fim"]),
                        "Duração (horas)": round(dur, 6),
                        "Intervalo": "Intervalo 1",
                        "Tempo Total do Dia": round(dur, 6),
                    })
                elif "ROLETE" in descricao or "GIRO DO ROLETE" in descricao:
                    dur = calc_duration_hours(intv["Início"], intv["Fim"])
                    roletes.append({
                        "Data": date_display,
                        "Equipamento": frota_id,
                        "Início": time_hhmmss(intv["Início"]),
                        "Fim": time_hhmmss(intv["Fim"]),
                        "Duração (horas)": round(dur, 6),
                        "Intervalo": "Intervalo 1",
                        "Tempo Total do Dia": round(dur, 6),
                    })

    # Agrupar lavagem/roletes por equipamento para calcular Tempo Total do Dia
    for lista in [lavagem, roletes]:
        equip_totals = defaultdict(float)
        for item in lista:
            equip_totals[item["Equipamento"]] += item["Duração (horas)"]
        for item in lista:
            item["Tempo Total do Dia"] = round(equip_totals[item["Equipamento"]], 6)

    # ── Dados Case Extra (para seção complementar) ──
    dados_case = {}
    for frota_id, case_info in case_frotas.items():
        if frota_id.startswith("_"):
            continue
        dados_case[frota_id] = {
            k: v for k, v in case_info.items() if not k.startswith("_")
        }

    # ── JSON Unificado ──
    resultado = {
        "metadata": {
            "date": iso_date,
            "type": "cd_diario_novo",
            "frente": "frente5",
            "generated_at": datetime.now().isoformat(),
            "fontes": fontes_usadas,
        },
        "metas": METAS_DEFAULT,
        "eficiencia_energetica": eficiencia_energetica,
        "eficiencia_operacional": eficiencia_operacional,
        "horas_elevador": horas_elevador,
        "uso_gps": uso_gps,
        "media_velocidade": media_velocidade,
        "manobras_frotas": manobras_frotas,
        "motor_ocioso": motor_ocioso,
        "disponibilidade_mecanica": disponibilidade_mecanica,
        "horas_por_frota": horas_por_frota,
        "intervalos_operacao": intervalos_operacao,
        "ofensores": ofensores_list,
        "lavagem": lavagem,
        "roletes": roletes,
        "producao": round(total_producao, 2),
        "producao_total": [{"valor": round(total_producao, 2)}],
        "producao_por_frota": producao_por_frota,
        "imagens": {
            "mapaGPS": "",
            "areaTrabalhada": "",
        },
        "dados_case": dados_case,
    }

    return resultado


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  🔄 Consolidação de JSON Unificado por Dia")
    print("=" * 60)

    # 1. Listar datas disponíveis (dos JSONs Solinftec existentes)
    solinftec_files = sorted(glob.glob(os.path.join(SOLINFTEC_JSON_DIR, "*.json")))
    dates = []
    for sf in solinftec_files:
        d = parse_date_from_filename(os.path.basename(sf))
        if d:
            dates.append(d)

    if not dates:
        print("  ❌ Nenhum JSON Solinftec encontrado.")
        return

    print(f"\n  📅 Datas encontradas: {dates}")

    # 2. Carregar dados Case (uma vez para todas as datas)
    print("\n  📦 Carregando dados Case IH...")
    case_data = load_case_data()
    print(f"     Datas Case disponíveis: {[k for k in case_data.keys() if not k.startswith('_')]}")

    # 3. Processar cada data
    for date_str in dates:
        print(f"\n  {'─' * 50}")
        print(f"  📅 Processando {date_str}...")

        # Solinftec
        solinftec_raw = load_solinftec(date_str)
        n_frotas_sol = len(solinftec_raw) if solinftec_raw else 0
        print(f"     Solinftec: {n_frotas_sol} frotas")

        # Case
        case_date_key = date_str.replace("-", "/")
        case_frotas = case_data.get(case_date_key, {})
        n_frotas_case = len([k for k in case_frotas if not k.startswith("_")])
        print(f"     Case IH: {n_frotas_case} frotas")

        # OPC
        opc_data = load_opc_daily(date_str)
        if opc_data:
            print(f"     OPC: {list(opc_data.keys())}")
        else:
            print(f"     OPC: sem dados")

        # Consolidar
        resultado = consolidar_dia(date_str, solinftec_raw, case_data, opc_data)

        # Salvar
        output_name = f"colhedora_frota_{date_str}.json"
        output_path = os.path.join(OUTPUT_DIR, output_name)
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        n_total = len(resultado.get("eficiencia_energetica", []))
        n_intervalos = len(resultado.get("intervalos_operacao", []))
        n_ofensores = len(resultado.get("ofensores", []))
        print(f"     ✅ Salvo: {output_name}")
        print(f"        {n_total} frotas, {n_intervalos} intervalos, {n_ofensores} ofensores")
        print(f"        Fontes: {resultado['metadata']['fontes']}")

    print(f"\n{'=' * 60}")
    print(f"  ✅ Consolidação concluída! {len(dates)} dias processados.")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
