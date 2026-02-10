import os, json
import openpyxl

ETL_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
case_path = os.path.join(ETL_ROOT, "dados", "Consolidado_Case_05_11-10-2025.xlsx")
wb = openpyxl.load_workbook(case_path, read_only=True)

for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    print(f"\n=== {sn} ({len(rows)} sample rows) ===")
    if rows:
        headers = list(rows[0])
        print(f"Headers: {headers}")
        if len(rows) > 1:
            print(f"Row 1: {list(rows[1])}")
        if len(rows) > 2:
            print(f"Row 2: {list(rows[2])}")
wb.close()
