import os
import sys
import shutil
import warnings
import unicodedata
import zipfile
import re
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

# Suprimir avisos específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- CONFIGURAÇÕES GERAIS ---
# Diretório onde estão os arquivos Excel originais
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DIRETORIO_ENTRADA = os.path.join(BASE_DIR, "dados")

# Colunas a serem removidas na etapa 1
COLUNAS_PARA_REMOVER = [
    "Descrição Regional",
    "Descrição da Unidade",
    "Descrição do Grupo de Equipamento",
    "Código da Fazenda",
    "Código da Zona",
    "Código do Talhão",
    "Descrição da Fazenda",
    "Horímetro/Odometro Secundário"
]
# ----------------------------

def validar_diretorio(caminho):
    """Verifica se o diretório de entrada existe."""
    if not os.path.exists(caminho):
        print(f"ERRO: O diretório configurado não existe: {caminho}")
        return False
    return True

def extrair_periodo_nome_arquivo(nome_arquivo):
    """
    Extrai datas de início e fim do nome do arquivo.
    Padrão esperado: ...DD-MM-YYYY_DD-MM-YYYY...
    """
    match = re.search(r"(\d{2}-\d{2}-\d{4})_(\d{2}-\d{2}-\d{4})", nome_arquivo)
    if match:
        d1_str, d2_str = match.groups()
        try:
            d1 = pd.to_datetime(d1_str, format="%d-%m-%Y").date()
            d2 = pd.to_datetime(d2_str, format="%d-%m-%Y").date()
            return d1, d2
        except:
            pass
    return None, None

def obter_arquivos_xlsx(diretorio):
    """Retorna uma lista de arquivos .xlsx no diretório."""
    arquivos = [
        os.path.join(diretorio, f) 
        for f in os.listdir(diretorio) 
        if f.lower().endswith(".xlsx") and not f.startswith("~$") # Ignora arquivos temporários do Excel
    ]
    return arquivos

def obter_arquivos_zip(diretorio):
    arquivos = [
        os.path.join(diretorio, f)
        for f in os.listdir(diretorio)
        if f.lower().endswith(".zip") and not f.startswith("~$") and "linha_do_tempo" in f.lower()
    ]
    return arquivos

def extrair_zips(diretorio, arquivos_zip):
    for arquivo_zip in arquivos_zip:
        try:
            with zipfile.ZipFile(arquivo_zip, "r") as zf:
                membros = [
                    n for n in zf.namelist()
                    if n.lower().endswith(".xls") or n.lower().endswith(".xlsx")
                ]
                for membro in membros:
                    zf.extract(membro, diretorio)
            print(f"ZIP extraído: {os.path.basename(arquivo_zip)}")
        except Exception as e:
            print(f"ERRO ao extrair ZIP {os.path.basename(arquivo_zip)}: {e}")

def ajustar_largura_colunas(worksheet):
    """
    Ajusta a largura das colunas de uma worksheet baseada no conteúdo.
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        
        # Define um tamanho mínimo e adiciona um pouco de folga
        adjusted_width = (max_length + 2) * 1.2
        # Limita a largura máxima para não ficar excessivamente larga se houver textos gigantes
        if adjusted_width > 100:
            adjusted_width = 100
        worksheet.column_dimensions[column_letter].width = adjusted_width

def formatar_coluna_data(ws):
    """Percorre a primeira linha para achar 'Data' e aplica formatação."""
    col_data_idx = None
    for cell in ws[1]:
        if cell.value == "Data":
            col_data_idx = cell.column
            break
    if col_data_idx is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_data_idx)
            if cell.value:
                cell.number_format = "dd/mm/yyyy"

def normalizar_texto(valor):
    if valor is None:
        return ""
    texto = str(valor)
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.lower()

def normalizar_numero_serie(serie):
    serie_texto = serie.astype(str).str.replace("\u00a0", " ", regex=False)
    serie_texto = serie_texto.str.replace(" ", "", regex=False).str.replace(",", ".", regex=False)
    return pd.to_numeric(serie_texto, errors="coerce").fillna(0)

def formatar_nome_grupo(grupo_norm):
    return str(grupo_norm).strip().title().replace(" ", "_")

def encontrar_coluna_horas_manut(df):
    for col in df.columns:
        col_norm = normalizar_texto(col).replace("_", "")
        if col_norm.startswith("horasmanut"):
            return col
    return None

def ordenar_colunas(df, cols_base, cols_ordem):
    cols = [c for c in cols_base if c in df.columns]
    for c in cols_ordem:
        if c in df.columns and c not in cols:
            cols.append(c)
    for c in df.columns:
        if c not in cols:
            cols.append(c)
    return df[cols]

def tratar_arquivo(caminho_arquivo):
    """
    Abre o arquivo Excel, remove colunas especificadas e salva em uma nova aba.
    """
    print(f"\nIniciando processamento: {os.path.basename(caminho_arquivo)}")
    
    try:
        # Carrega a planilha original (primeira aba)
        # Usamos engine='openpyxl' para garantir compatibilidade
        df_original = pd.read_excel(caminho_arquivo, sheet_name=0, engine="openpyxl")
        
        # Identifica quais colunas da lista realmente existem no arquivo
        colunas_existentes = [col for col in COLUNAS_PARA_REMOVER if col in df_original.columns]
        
        # Filtro por período extraído do nome do arquivo (Aplicar ao df_original também para garantir consistência)
        dt_inicio_filtro, dt_fim_filtro = extrair_periodo_nome_arquivo(os.path.basename(caminho_arquivo))
        if dt_inicio_filtro and dt_fim_filtro:
            print(f"  Filtrando dados pelo período: {dt_inicio_filtro} a {dt_fim_filtro}")
            
            # Precisamos identificar a coluna de data no df_original
            col_data_orig = None
            for col in df_original.columns:
                if "Data" in str(col) or "data" in str(col).lower():
                    # Tenta validar se é data
                    try:
                        # Verifica primeiras linhas não nulas
                        sample = df_original[col].dropna().head()
                        if not sample.empty:
                            pd.to_datetime(sample, dayfirst=True, errors='raise')
                            col_data_orig = col
                            break
                    except:
                        continue
            
            if col_data_orig:
                df_original["_temp_data_"] = pd.to_datetime(df_original[col_data_orig], dayfirst=True, errors="coerce").dt.date
                df_original = df_original[
                    (df_original["_temp_data_"] >= dt_inicio_filtro) & 
                    (df_original["_temp_data_"] <= dt_fim_filtro)
                ].drop(columns=["_temp_data_"])
            else:
                 print("  AVISO: Não foi possível filtrar df_original por data (coluna não identificada).")

        if not colunas_existentes:
            print("  AVISO: Nenhuma das colunas configuradas para remoção foi encontrada.")
            df_tratado = df_original.copy()
        else:
            print(f"  Removendo {len(colunas_existentes)} colunas...")
            df_tratado = df_original.drop(columns=colunas_existentes)

        if "Descrição do Equipamento" in df_tratado.columns:
            desc_series = df_tratado["Descrição do Equipamento"].astype(str)
            mask_colhedora_cana = desc_series.str.strip().str.upper() == "COLHEDORA DE CANA"
            df_tratado.loc[mask_colhedora_cana, "Descrição do Equipamento"] = "COLHEDORA"
            
            mask_trator_transbordo = desc_series.str.strip().str.upper() == "TRATOR TRANSBORDO"
            df_tratado.loc[mask_trator_transbordo, "Descrição do Equipamento"] = "TRATORES"

        df_calc = df_tratado.copy()

        try:
            df_calc["Data"] = pd.to_datetime(df_calc["Data Hora Local"], dayfirst=True, errors="coerce").dt.date
            
            # Filtro já aplicado no df_original, mas garantimos aqui caso df_calc tenha sido criado antes ou precise de reforço
            if dt_inicio_filtro and dt_fim_filtro:
                 df_calc = df_calc.dropna(subset=["Data"])
                 df_calc = df_calc[
                     (df_calc["Data"] >= dt_inicio_filtro) & 
                     (df_calc["Data"] <= dt_fim_filtro)
                 ].copy()

            df_calc["Hora Inicial"] = df_calc["Hora Inicial"].astype(str)
            df_calc["Hora Final"] = df_calc["Hora Final"].astype(str)

            inicio_str = df_calc["Data Hora Local"].astype(str) + " " + df_calc["Hora Inicial"]
            fim_str = df_calc["Data Hora Local"].astype(str) + " " + df_calc["Hora Final"]

            df_calc["dt_inicial"] = pd.to_datetime(inicio_str, dayfirst=True, errors="coerce")
            df_calc["dt_final"] = pd.to_datetime(fim_str, dayfirst=True, errors="coerce")

            df_calc["Duracao_min"] = (df_calc["dt_final"] - df_calc["dt_inicial"]).dt.total_seconds() / 60

            mask_negativo = df_calc["Duracao_min"] < 0
            df_calc.loc[mask_negativo, "Duracao_min"] += 1440
        except Exception as e:
            print(f"  AVISO: Não foi possível calcular duração (erro de formato de data): {e}")
            df_calc["Duracao_min"] = 0

        if "Duracao_min" in df_calc.columns:
            df_calc["Duracao_h"] = df_calc["Duracao_min"] / 60
            if "Hora Final" in df_tratado.columns:
                try:
                    idx = df_tratado.columns.get_loc("Hora Final") + 1
                    df_tratado.insert(idx, "Duração", df_calc["Duracao_h"])
                except Exception:
                    df_tratado["Duração"] = df_calc["Duracao_h"]

        df_intervalos_cols = []
        for col in [
            "Data",
            "Código Equipamento",
            "Descrição do Equipamento",
            "Código de Operador",
            "Nome",
            "Descrição da Operação",
            "Descrição do Grupo da Operação",
            "Velocidade Média",
        ]:
            if col in df_calc.columns:
                df_intervalos_cols.append(col)

        df_intervalos = df_calc[df_intervalos_cols].copy()
        if "dt_inicial" in df_calc.columns:
            df_intervalos["Início"] = df_calc["dt_inicial"]
        if "dt_final" in df_calc.columns:
            df_intervalos["Fim"] = df_calc["dt_final"]
        if "Duracao_min" in df_calc.columns:
            df_intervalos["Duração (min)"] = df_calc["Duracao_min"]

        grupo_col = "Descrição do Grupo da Operação"
        op_col = "Descrição da Operação"

        df_calc["dur_total"] = df_calc["Duracao_min"]
        grupo_series = df_calc[grupo_col] if grupo_col in df_calc.columns else pd.Series("", index=df_calc.index)
        grupo_norm = grupo_series.astype(str).str.strip().str.upper()
        df_calc["dur_prod"] = np.where(grupo_norm == "PRODUTIVA", df_calc["Duracao_min"], 0)
        df_calc["dur_improd"] = np.where(grupo_norm == "IMPRODUTIVA", df_calc["Duracao_min"], 0)
        grupos_extra = sorted([g for g in grupo_norm.dropna().unique() if g and g not in ["PRODUTIVA", "IMPRODUTIVA"]])
        dur_grupo_cols = []
        horas_grupo_cols = []
        for grupo in grupos_extra:
            col_dur = f"dur_grupo_{grupo}"
            df_calc[col_dur] = np.where(grupo_norm == grupo, df_calc["Duracao_min"], 0)
            dur_grupo_cols.append(col_dur)
            horas_grupo_cols.append(f"Horas_{formatar_nome_grupo(grupo)}")

        df_calc["dur_manobra"] = np.where(df_calc.get(op_col, "") == "MANOBRA", df_calc["Duracao_min"], 0)
        df_calc["dur_transbordo"] = np.where(df_calc.get(op_col, "") == "TRANSBORDANDO CANA", df_calc["Duracao_min"], 0)
        df_calc["dur_sem_apont"] = np.where(df_calc.get(op_col, "") == "SEM APONTAMENTO", df_calc["Duracao_min"], 0)

        df_calc["dur_colheita"] = np.where(
            df_calc.get(op_col, "").isin(["CARREGANDO CANA", "COLHENDO CANA"]),
            df_calc["Duracao_min"],
            0,
        )
        df_calc["dur_vazio"] = np.where(df_calc.get(op_col, "") == "DESL VAZIO", df_calc["Duracao_min"], 0)
        df_calc["dur_carregado"] = np.where(df_calc.get(op_col, "") == "DESL CARREGADO", df_calc["Duracao_min"], 0)

        vm_series = df_calc.get("Velocidade Média")
        vm = normalizar_numero_serie(vm_series) if vm_series is not None else pd.Series(0, index=df_calc.index)

        colunas_normalizadas = {col: normalizar_texto(col) for col in df_calc.columns}
        horimetro_inicial_col = next(
            (col for col, nome in colunas_normalizadas.items() if ("horimetro" in nome or "odometro" in nome) and "inicial" in nome),
            None,
        )
        horimetro_final_col = next(
            (col for col, nome in colunas_normalizadas.items() if ("horimetro" in nome or "odometro" in nome) and "final" in nome),
            None,
        )

        if horimetro_inicial_col and horimetro_final_col:
            horimetro_inicial = normalizar_numero_serie(df_calc[horimetro_inicial_col])
            horimetro_final = normalizar_numero_serie(df_calc[horimetro_final_col])
            horimetro_diff = horimetro_final - horimetro_inicial
            mask_horimetro = horimetro_diff != 0
        else:
            mask_horimetro = pd.Series(False, index=df_calc.index)
            horimetro_diff = pd.Series(0, index=df_calc.index)

        mask_motor_ocioso = (
            mask_horimetro
            & (vm == 0)
            & (grupo_norm == "IMPRODUTIVA")
        )
        df_calc["dur_motor_ocioso"] = np.where(mask_motor_ocioso, df_calc["Duracao_min"], 0)
        df_calc["dur_motor_ocioso_h"] = df_calc["dur_motor_ocioso"] / 60
        df_calc["dur_motor_ligado"] = np.where(horimetro_diff > 0, df_calc["Duracao_min"], 0)
        df_calc["dur_motor_ligado_h"] = df_calc["dur_motor_ligado"] / 60

        if "Dif_Horimetro" not in df_tratado.columns:
            dif_horimetro_series = horimetro_diff
            if "Velocidade Média" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Velocidade Média")
                df_tratado.insert(idx, "Dif_Horimetro", dif_horimetro_series)
            else:
                df_tratado["Dif_Horimetro"] = dif_horimetro_series

        if "Motor Ligado" not in df_tratado.columns:
            motor_ligado_series = df_calc.get("dur_motor_ligado_h", 0)
            if "Dif_Horimetro" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Dif_Horimetro") + 1
                df_tratado.insert(idx, "Motor Ligado", motor_ligado_series)
            elif "Velocidade Média" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Velocidade Média") + 1
                df_tratado.insert(idx, "Motor Ligado", motor_ligado_series)
            else:
                df_tratado["Motor Ligado"] = motor_ligado_series

        if "Motor Ocioso" not in df_tratado.columns:
            motor_ocioso_series = df_calc.get("dur_motor_ocioso_h", 0)
            if "Motor Ligado" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Motor Ligado") + 1
                df_tratado.insert(idx, "Motor Ocioso", motor_ocioso_series)
            elif "Dif_Horimetro" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Dif_Horimetro") + 1
                df_tratado.insert(idx, "Motor Ocioso", motor_ocioso_series)
            elif "Velocidade Média" in df_tratado.columns:
                idx = df_tratado.columns.get_loc("Velocidade Média") + 1
                df_tratado.insert(idx, "Motor Ocioso", motor_ocioso_series)
            else:
                df_tratado["Motor Ocioso"] = motor_ocioso_series

        df_calc["vel_colheita_x_min"] = np.where(df_calc["dur_colheita"] > 0, vm * df_calc["dur_colheita"], 0)
        df_calc["vel_vazio_x_min"] = np.where(df_calc["dur_vazio"] > 0, vm * df_calc["dur_vazio"], 0)
        df_calc["vel_carregado_x_min"] = np.where(df_calc["dur_carregado"] > 0, vm * df_calc["dur_carregado"], 0)

        df_calc["cnt_manobra"] = np.where(df_calc.get(op_col, "") == "MANOBRA", 1, 0)
        df_calc["cnt_transbordo"] = np.where(df_calc.get(op_col, "") == "TRANSBORDANDO CANA", 1, 0)

        col_data = "Data"
        col_equip = "Código Equipamento"
        col_equip_desc = "Descrição do Equipamento"
        col_op_cod = "Código de Operador"
        col_op_nome = "Nome"

        # --- 3.Dia_Frota ---
        df_dia_frota = pd.DataFrame()
        if all(c in df_calc.columns for c in [col_data, col_equip]):
            group_cols_frota = [c for c in [col_data, col_equip, col_equip_desc] if c in df_calc.columns]
            agg_frota = {
                "dur_total_min": ("dur_total", "sum"),
                "dur_prod_min": ("dur_prod", "sum"),
                "dur_improd_min": ("dur_improd", "sum"),
                "dur_manobra_min": ("dur_manobra", "sum"),
                "dur_transbordo_min": ("dur_transbordo", "sum"),
                "dur_sem_apont_min": ("dur_sem_apont", "sum"),
                "dur_colheita_min": ("dur_colheita", "sum"),
                "dur_vazio_min": ("dur_vazio", "sum"),
                "dur_carregado_min": ("dur_carregado", "sum"),
                "dur_motor_ocioso_min": ("dur_motor_ocioso", "sum"),
                "dur_motor_ligado_min": ("dur_motor_ligado", "sum"),
                "vel_colheita_x_min": ("vel_colheita_x_min", "sum"),
                "vel_vazio_x_min": ("vel_vazio_x_min", "sum"),
                "vel_carregado_x_min": ("vel_carregado_x_min", "sum"),
                "cnt_manobra": ("cnt_manobra", "sum"),
                "cnt_transbordo": ("cnt_transbordo", "sum"),
            }
            for col_dur in dur_grupo_cols:
                agg_frota[col_dur] = (col_dur, "sum")
            df_dia_frota = df_calc.groupby(group_cols_frota).agg(**agg_frota).reset_index()

            # Converter para horas e renomear
            df_dia_frota["Horas_Registradas"] = df_dia_frota["dur_total_min"] / 60
            df_dia_frota["Horas_Produtivas"] = df_dia_frota["dur_prod_min"] / 60
            df_dia_frota["Horas_Improdutivas"] = df_dia_frota["dur_improd_min"] / 60
            df_dia_frota["Horas_Motor_Ocioso"] = df_dia_frota["dur_motor_ocioso_min"] / 60
            df_dia_frota["Horas_Motor_Ligado"] = df_dia_frota["dur_motor_ligado_min"] / 60
            for col_dur, col_horas in zip(dur_grupo_cols, horas_grupo_cols):
                df_dia_frota[col_horas] = df_dia_frota[col_dur] / 60
            df_dia_frota["Tempo_Sem_Apontamento_h"] = df_dia_frota["dur_sem_apont_min"] / 60
            
            df_dia_frota["Porcentagem_Motor_Ligado"] = np.where(
                df_dia_frota["Horas_Registradas"] > 0,
                (df_dia_frota["Horas_Motor_Ligado"] / df_dia_frota["Horas_Registradas"]) * 100,
                0,
            )
            df_dia_frota["Porcentagem_Motor_Ocioso"] = np.where(
                df_dia_frota["Horas_Registradas"] > 0,
                (df_dia_frota["Horas_Motor_Ocioso"] / df_dia_frota["Horas_Registradas"]) * 100,
                0,
            )
            df_dia_frota["Porcentagem_Sem_Apontamento"] = np.where(
                df_dia_frota["Horas_Registradas"] > 0,
                (df_dia_frota["Tempo_Sem_Apontamento_h"] / df_dia_frota["Horas_Registradas"]) * 100,
                0,
            )
            
            # Manter métricas de manobra/transbordo em minutos (geralmente são curtos) ou converter se desejado            
            df_dia_frota["Tempo_Manobra_total_h"] = df_dia_frota["dur_manobra_min"] / 60
            df_dia_frota["Tempo_Manobra_medio_min"] = np.where(
                df_dia_frota["cnt_manobra"] > 0,
                df_dia_frota["dur_manobra_min"] / df_dia_frota["cnt_manobra"],
                0,
            )

            df_dia_frota["Tempo_Transbordo_total_h"] = df_dia_frota["dur_transbordo_min"] / 60
            df_dia_frota["Tempo_Transbordo_medio_min"] = np.where(
                df_dia_frota["cnt_transbordo"] > 0,
                df_dia_frota["dur_transbordo_min"] / df_dia_frota["cnt_transbordo"],
                0,
            )

            df_dia_frota.rename(
                columns={
                    "cnt_manobra": "Quantidade_Manobras",
                    "cnt_transbordo": "Quantidade_Transbordos",
                    "Tempo_Manobra_total_h": "Tempo_Total_Manobras_h",
                    "Tempo_Manobra_medio_min": "Tempo_Medio_Manobras_min",
                    "Tempo_Transbordo_total_h": "Tempo_Total_Transbordo_h",
                    "Tempo_Transbordo_medio_min": "Tempo_Medio_Transbordo_min",
                },
                inplace=True,
            )

            col_manut_frota = encontrar_coluna_horas_manut(df_dia_frota)
            if col_manut_frota:
                df_dia_frota["Disponibilidade_Mecanica"] = np.where(
                    df_dia_frota["Horas_Registradas"] > 0,
                    1 - (df_dia_frota[col_manut_frota] / df_dia_frota["Horas_Registradas"]),
                    0,
                )
            df_dia_frota["Eficiencia_Energetica"] = np.where(
                df_dia_frota["Horas_Motor_Ligado"] > 0,
                df_dia_frota["Horas_Produtivas"] / df_dia_frota["Horas_Motor_Ligado"],
                0,
            )
            df_dia_frota["Eficiencia_Operacional"] = np.where(
                df_dia_frota["Horas_Registradas"] > 0,
                df_dia_frota["Horas_Produtivas"] / df_dia_frota["Horas_Registradas"],
                0,
            )

            df_dia_frota["Vel_Colheita_media"] = np.where(
                df_dia_frota["dur_colheita_min"] > 0,
                df_dia_frota["vel_colheita_x_min"] / df_dia_frota["dur_colheita_min"],
                np.nan,
            )
            df_dia_frota["Vel_Desl_Vazio_media"] = np.where(
                df_dia_frota["dur_vazio_min"] > 0,
                df_dia_frota["vel_vazio_x_min"] / df_dia_frota["dur_vazio_min"],
                np.nan,
            )
            df_dia_frota["Vel_Desl_Carregado_media"] = np.where(
                df_dia_frota["dur_carregado_min"] > 0,
                df_dia_frota["vel_carregado_x_min"] / df_dia_frota["dur_carregado_min"],
                np.nan,
            )
            
            # Remover colunas auxiliares em minutos que não são mais necessárias
            cols_drop = [
                "dur_total_min", "dur_prod_min", "dur_improd_min", "dur_motor_ocioso_min", "dur_motor_ligado_min", "dur_sem_apont_min",
                "dur_manobra_min", "dur_transbordo_min", "dur_colheita_min",
                "dur_vazio_min", "dur_carregado_min", "vel_colheita_x_min",
                "vel_vazio_x_min", "vel_carregado_x_min"
            ]
            cols_drop.extend(dur_grupo_cols)
            df_dia_frota.drop(columns=cols_drop, inplace=True, errors="ignore")
            horas_ordem = [
                "Horas_Registradas",
                "Horas_Produtivas",
                "Horas_Improdutivas",
            ]
            for col in horas_grupo_cols:
                horas_ordem.append(col)
                if col_manut_frota and col == col_manut_frota:
                    horas_ordem.append("Disponibilidade_Mecanica")
            
            horas_ordem.extend([
                "Horas_Motor_Ligado",
                "Porcentagem_Motor_Ligado",
                "Horas_Motor_Ocioso",
                "Porcentagem_Motor_Ocioso",
                "Tempo_Sem_Apontamento_h",
                "Porcentagem_Sem_Apontamento",
                "Eficiencia_Energetica",
                "Eficiencia_Operacional",
                "Tempo_Total_Manobras_h",
                "Quantidade_Manobras",
                "Tempo_Medio_Manobras_min",
                "Tempo_Total_Transbordo_h",
                "Quantidade_Transbordos",
                "Tempo_Medio_Transbordo_min",
            ])
            # Caso Disponibilidade_Mecanica não tenha sido inserida (coluna de manutenção não encontrada nos grupos)
            if "Disponibilidade_Mecanica" in df_dia_frota.columns and "Disponibilidade_Mecanica" not in horas_ordem:
                horas_ordem.append("Disponibilidade_Mecanica")
            df_dia_frota = ordenar_colunas(df_dia_frota, group_cols_frota, horas_ordem)

        # --- 4.Dia_Operador ---
        df_dia_operador = pd.DataFrame()
        if all(c in df_calc.columns for c in [col_data, col_op_cod, col_op_nome]):
            group_cols_op = [col_data, col_op_cod, col_op_nome]
            if col_equip_desc in df_calc.columns:
                group_cols_op.append(col_equip_desc)
            
            agg_operador = {
                "dur_total_min": ("dur_total", "sum"),
                "dur_prod_min": ("dur_prod", "sum"),
                "dur_improd_min": ("dur_improd", "sum"),
                "dur_manobra_min": ("dur_manobra", "sum"),
                "dur_transbordo_min": ("dur_transbordo", "sum"),
                "dur_sem_apont_min": ("dur_sem_apont", "sum"),
                "dur_colheita_min": ("dur_colheita", "sum"),
                "dur_vazio_min": ("dur_vazio", "sum"),
                "dur_carregado_min": ("dur_carregado", "sum"),
                "dur_motor_ocioso_min": ("dur_motor_ocioso", "sum"),
                "dur_motor_ligado_min": ("dur_motor_ligado", "sum"),
                "vel_colheita_x_min": ("vel_colheita_x_min", "sum"),
                "vel_vazio_x_min": ("vel_vazio_x_min", "sum"),
                "vel_carregado_x_min": ("vel_carregado_x_min", "sum"),
                "cnt_manobra": ("cnt_manobra", "sum"),
                "cnt_transbordo": ("cnt_transbordo", "sum"),
            }
            for col_dur in dur_grupo_cols:
                agg_operador[col_dur] = (col_dur, "sum")
            df_dia_operador = df_calc.groupby(group_cols_op).agg(**agg_operador).reset_index()

            if col_equip in df_calc.columns:
                frotas = (
                    df_calc.groupby(group_cols_op)[col_equip]
                    .agg(lambda x: ", ".join(str(v) for v in sorted(set(x.dropna()))))
                    .reset_index()
                )
                df_dia_operador = df_dia_operador.merge(frotas, on=group_cols_op, how="left")
                df_dia_operador.rename(columns={col_equip: "Frotas_no_dia"}, inplace=True)

            df_dia_operador["Horas_Registradas"] = df_dia_operador["dur_total_min"] / 60
            df_dia_operador["Horas_Produtivas"] = df_dia_operador["dur_prod_min"] / 60
            df_dia_operador["Horas_Improdutivas"] = df_dia_operador["dur_improd_min"] / 60
            df_dia_operador["Horas_Motor_Ocioso"] = df_dia_operador["dur_motor_ocioso_min"] / 60
            df_dia_operador["Horas_Motor_Ligado"] = df_dia_operador["dur_motor_ligado_min"] / 60
            for col_dur, col_horas in zip(dur_grupo_cols, horas_grupo_cols):
                df_dia_operador[col_horas] = df_dia_operador[col_dur] / 60
            df_dia_operador["Tempo_Sem_Apontamento_h"] = df_dia_operador["dur_sem_apont_min"] / 60
            
            df_dia_operador["Porcentagem_Motor_Ligado"] = np.where(
                df_dia_operador["Horas_Registradas"] > 0,
                (df_dia_operador["Horas_Motor_Ligado"] / df_dia_operador["Horas_Registradas"]) * 100,
                0,
            )
            df_dia_operador["Porcentagem_Motor_Ocioso"] = np.where(
                df_dia_operador["Horas_Registradas"] > 0,
                (df_dia_operador["Horas_Motor_Ocioso"] / df_dia_operador["Horas_Registradas"]) * 100,
                0,
            )
            df_dia_operador["Porcentagem_Sem_Apontamento"] = np.where(
                df_dia_operador["Horas_Registradas"] > 0,
                (df_dia_operador["Tempo_Sem_Apontamento_h"] / df_dia_operador["Horas_Registradas"]) * 100,
                0,
            )

            df_dia_operador["Tempo_Manobra_total_h"] = df_dia_operador["dur_manobra_min"] / 60
            df_dia_operador["Tempo_Manobra_medio_min"] = np.where(
                df_dia_operador["cnt_manobra"] > 0,
                df_dia_operador["dur_manobra_min"] / df_dia_operador["cnt_manobra"],
                0,
            )

            df_dia_operador["Tempo_Transbordo_total_h"] = df_dia_operador["dur_transbordo_min"] / 60
            df_dia_operador["Tempo_Transbordo_medio_min"] = np.where(
                df_dia_operador["cnt_transbordo"] > 0,
                df_dia_operador["dur_transbordo_min"] / df_dia_operador["cnt_transbordo"],
                0,
            )

            df_dia_operador.rename(
                columns={
                    "cnt_manobra": "Quantidade_Manobras",
                    "cnt_transbordo": "Quantidade_Transbordos",
                    "Tempo_Manobra_total_h": "Tempo_Total_Manobras_h",
                    "Tempo_Manobra_medio_min": "Tempo_Medio_Manobras_min",
                    "Tempo_Transbordo_total_h": "Tempo_Total_Transbordo_h",
                    "Tempo_Transbordo_medio_min": "Tempo_Medio_Transbordo_min",
                },
                inplace=True,
            )

            # col_manut_operador = encontrar_coluna_horas_manut(df_dia_operador)
            # if col_manut_operador:
            #     df_dia_operador["Disponibilidade_Mecanica"] = np.where(
            #         df_dia_operador["Horas_Registradas"] > 0,
            #         df_dia_operador[col_manut_operador] / df_dia_operador["Horas_Registradas"],
            #         0,
            #     )
            df_dia_operador["Eficiencia_Energetica"] = np.where(
                df_dia_operador["Horas_Motor_Ligado"] > 0,
                df_dia_operador["Horas_Produtivas"] / df_dia_operador["Horas_Motor_Ligado"],
                0,
            )
            df_dia_operador["Eficiencia_Operacional"] = np.where(
                df_dia_operador["Horas_Registradas"] > 0,
                df_dia_operador["Horas_Produtivas"] / df_dia_operador["Horas_Registradas"],
                0,
            )

            df_dia_operador["Vel_Colheita_media"] = np.where(
                df_dia_operador["dur_colheita_min"] > 0,
                df_dia_operador["vel_colheita_x_min"] / df_dia_operador["dur_colheita_min"],
                np.nan,
            )
            df_dia_operador["Vel_Desl_Vazio_media"] = np.where(
                df_dia_operador["dur_vazio_min"] > 0,
                df_dia_operador["vel_vazio_x_min"] / df_dia_operador["dur_vazio_min"],
                np.nan,
            )
            df_dia_operador["Vel_Desl_Carregado_media"] = np.where(
                df_dia_operador["dur_carregado_min"] > 0,
                df_dia_operador["vel_carregado_x_min"] / df_dia_operador["dur_carregado_min"],
                np.nan,
            )

            cols_drop = [
                "dur_total_min", "dur_prod_min", "dur_improd_min", "dur_motor_ocioso_min", "dur_motor_ligado_min", "dur_sem_apont_min",
                "dur_manobra_min", "dur_transbordo_min", "dur_colheita_min",
                "dur_vazio_min", "dur_carregado_min", "vel_colheita_x_min",
                "vel_vazio_x_min", "vel_carregado_x_min"
            ]
            cols_drop.extend(dur_grupo_cols)
            df_dia_operador.drop(columns=cols_drop, inplace=True, errors="ignore")
            horas_ordem = [
                "Horas_Registradas",
                "Horas_Produtivas",
                "Horas_Improdutivas",
                *horas_grupo_cols,
                "Horas_Motor_Ligado",
                "Porcentagem_Motor_Ligado",
                "Horas_Motor_Ocioso",
                "Porcentagem_Motor_Ocioso",
                "Tempo_Sem_Apontamento_h",
                "Porcentagem_Sem_Apontamento",
                # "Disponibilidade_Mecanica",
                "Eficiencia_Energetica",
                "Eficiencia_Operacional",
                "Tempo_Total_Manobras_h",
                "Quantidade_Manobras",
                "Tempo_Medio_Manobras_min",
                "Tempo_Total_Transbordo_h",
                "Quantidade_Transbordos",
                "Tempo_Medio_Transbordo_min",
            ]
            df_dia_operador = ordenar_colunas(df_dia_operador, group_cols_op, horas_ordem)

        df_periodo_frota = pd.DataFrame()
        if not df_dia_frota.empty and col_data in df_dia_frota.columns:
            group_cols = [c for c in [col_equip, col_equip_desc] if c in df_dia_frota.columns]
            if group_cols:
                cols_totais = [
                    "Horas_Registradas",
                    "Horas_Produtivas",
                    "Horas_Improdutivas",
                    *horas_grupo_cols,
                    "Horas_Motor_Ligado",
                    "Horas_Motor_Ocioso",
                    "Tempo_Sem_Apontamento_h",
                ]
                agg_periodo = {
                    f"{col}_total": (col, "sum")
                    for col in cols_totais
                    if col in df_dia_frota.columns
                }
                agg_periodo["Dias_com_dados"] = (col_data, "nunique")
                df_periodo_frota = df_dia_frota.groupby(group_cols).agg(**agg_periodo).reset_index()
                df_periodo_frota["Horas_media_por_dia"] = np.where(
                    df_periodo_frota["Dias_com_dados"] > 0,
                    df_periodo_frota["Horas_Registradas_total"] / df_periodo_frota["Dias_com_dados"],
                    0,
                )
                df_periodo_frota["Horas_Motor_Ocioso_media_por_dia"] = np.where(
                    df_periodo_frota["Dias_com_dados"] > 0,
                    df_periodo_frota["Horas_Motor_Ocioso_total"] / df_periodo_frota["Dias_com_dados"],
                    0,
                )
                col_manut_periodo_frota = encontrar_coluna_horas_manut(df_periodo_frota)
                if col_manut_periodo_frota:
                    df_periodo_frota["Disponibilidade_Mecanica"] = np.where(
                        df_periodo_frota["Horas_Registradas_total"] > 0,
                        1 - (df_periodo_frota[col_manut_periodo_frota] / df_periodo_frota["Horas_Registradas_total"]),
                        0,
                    )
                df_periodo_frota["Eficiencia_Energetica"] = np.where(
                    df_periodo_frota["Horas_Motor_Ligado_total"] > 0,
                    df_periodo_frota["Horas_Produtivas_total"] / df_periodo_frota["Horas_Motor_Ligado_total"],
                    0,
                )
                df_periodo_frota["Eficiencia_Operacional"] = np.where(
                    df_periodo_frota["Horas_Registradas_total"] > 0,
                    df_periodo_frota["Horas_Produtivas_total"] / df_periodo_frota["Horas_Registradas_total"],
                    0,
                )

                df_periodo_frota["Porcentagem_Motor_Ligado"] = np.where(
                    df_periodo_frota["Horas_Registradas_total"] > 0,
                    (df_periodo_frota["Horas_Motor_Ligado_total"] / df_periodo_frota["Horas_Registradas_total"]) * 100,
                    0,
                )
                df_periodo_frota["Porcentagem_Motor_Ocioso"] = np.where(
                    df_periodo_frota["Horas_Registradas_total"] > 0,
                    (df_periodo_frota["Horas_Motor_Ocioso_total"] / df_periodo_frota["Horas_Registradas_total"]) * 100,
                    0,
                )
                df_periodo_frota["Porcentagem_Sem_Apontamento"] = np.where(
                    df_periodo_frota["Horas_Registradas_total"] > 0,
                    (df_periodo_frota["Tempo_Sem_Apontamento_h_total"] / df_periodo_frota["Horas_Registradas_total"]) * 100,
                    0,
                )

                horas_total_ordem = [
                    "Horas_Registradas_total",
                    "Horas_Produtivas_total",
                    "Horas_Improdutivas_total",
                ]
                for col in horas_grupo_cols:
                    col_total = f"{col}_total"
                    horas_total_ordem.append(col_total)
                    if col_manut_periodo_frota and col_total == col_manut_periodo_frota:
                        horas_total_ordem.append("Disponibilidade_Mecanica")
                
                horas_total_ordem.extend([
                    "Horas_Motor_Ligado_total",
                    "Porcentagem_Motor_Ligado",
                    "Horas_Motor_Ocioso_total",
                    "Porcentagem_Motor_Ocioso",
                    "Tempo_Sem_Apontamento_h_total",
                    "Porcentagem_Sem_Apontamento",
                    "Eficiencia_Energetica",
                    "Eficiencia_Operacional",
                ])
                if "Disponibilidade_Mecanica" in df_periodo_frota.columns and "Disponibilidade_Mecanica" not in horas_total_ordem:
                    horas_total_ordem.append("Disponibilidade_Mecanica")
                df_periodo_frota = ordenar_colunas(df_periodo_frota, group_cols, horas_total_ordem)

        df_periodo_operador = pd.DataFrame()
        if not df_dia_operador.empty and col_data in df_dia_operador.columns:
            group_cols = [col_op_cod, col_op_nome]
            if col_equip_desc in df_dia_operador.columns:
                group_cols.append(col_equip_desc)

            cols_totais = [
                "Horas_Registradas",
                "Horas_Produtivas",
                "Horas_Improdutivas",
                *horas_grupo_cols,
                "Horas_Motor_Ligado",
                "Horas_Motor_Ocioso",
                "Tempo_Sem_Apontamento_h",
            ]
            agg_periodo = {
                f"{col}_total": (col, "sum")
                for col in cols_totais
                if col in df_dia_operador.columns
            }
            agg_periodo["Dias_com_dados"] = (col_data, "nunique")
            df_periodo_operador = df_dia_operador.groupby(group_cols).agg(**agg_periodo).reset_index()
            df_periodo_operador["Horas_media_por_dia"] = np.where(
                df_periodo_operador["Dias_com_dados"] > 0,
                df_periodo_operador["Horas_Registradas_total"] / df_periodo_operador["Dias_com_dados"],
                0,
            )
            df_periodo_operador["Horas_Motor_Ocioso_media_por_dia"] = np.where(
                df_periodo_operador["Dias_com_dados"] > 0,
                df_periodo_operador["Horas_Motor_Ocioso_total"] / df_periodo_operador["Dias_com_dados"],
                0,
            )
            # col_manut_periodo_operador = encontrar_coluna_horas_manut(df_periodo_operador)
            # if col_manut_periodo_operador:
            #     df_periodo_operador["Disponibilidade_Mecanica"] = np.where(
            #         df_periodo_operador["Horas_Registradas_total"] > 0,
            #         df_periodo_operador[col_manut_periodo_operador] / df_periodo_operador["Horas_Registradas_total"],
            #         0,
            #     )
            df_periodo_operador["Eficiencia_Energetica"] = np.where(
                df_periodo_operador["Horas_Motor_Ligado_total"] > 0,
                df_periodo_operador["Horas_Produtivas_total"] / df_periodo_operador["Horas_Motor_Ligado_total"],
                0,
            )
            df_periodo_operador["Eficiencia_Operacional"] = np.where(
                df_periodo_operador["Horas_Registradas_total"] > 0,
                df_periodo_operador["Horas_Produtivas_total"] / df_periodo_operador["Horas_Registradas_total"],
                0,
            )

            df_periodo_operador["Porcentagem_Motor_Ligado"] = np.where(
                df_periodo_operador["Horas_Registradas_total"] > 0,
                (df_periodo_operador["Horas_Motor_Ligado_total"] / df_periodo_operador["Horas_Registradas_total"]) * 100,
                0,
            )
            df_periodo_operador["Porcentagem_Motor_Ocioso"] = np.where(
                df_periodo_operador["Horas_Registradas_total"] > 0,
                (df_periodo_operador["Horas_Motor_Ocioso_total"] / df_periodo_operador["Horas_Registradas_total"]) * 100,
                0,
            )
            df_periodo_operador["Porcentagem_Sem_Apontamento"] = np.where(
                df_periodo_operador["Horas_Registradas_total"] > 0,
                (df_periodo_operador["Tempo_Sem_Apontamento_h_total"] / df_periodo_operador["Horas_Registradas_total"]) * 100,
                0,
            )

            horas_total_ordem = [
                "Horas_Registradas_total",
                "Horas_Produtivas_total",
                "Horas_Improdutivas_total",
                *[f"{col}_total" for col in horas_grupo_cols],
                "Horas_Motor_Ligado_total",
                "Porcentagem_Motor_Ligado",
                "Horas_Motor_Ocioso_total",
                "Porcentagem_Motor_Ocioso",
                "Tempo_Sem_Apontamento_h_total",
                "Porcentagem_Sem_Apontamento",
                # "Disponibilidade_Mecanica",
                "Eficiencia_Energetica",
                "Eficiencia_Operacional",
            ]
            df_periodo_operador = ordenar_colunas(df_periodo_operador, group_cols, horas_total_ordem)

        # --- 3.Top5Ofensores ---
        df_top5_ofensores = pd.DataFrame()
        # Verificar colunas necessárias
        if all(c in df_calc.columns for c in [col_equip_desc, col_data, grupo_col, op_col]):
            # Filtrar apenas o que é improdutivo (Ofensores)
            df_improd_only = df_calc[df_calc["dur_improd"] > 0].copy()
            
            if not df_improd_only.empty:
                # 1. Calcular Total de Horas por (Equipamento, Data) - usando df_calc completo (todo o tempo)
                # Isso servirá de denominador para a porcentagem
                df_total_dia = df_calc.groupby([col_equip_desc, col_data])["Duracao_min"].sum().reset_index()
                df_total_dia.rename(columns={"Duracao_min": "Total_Horas_Dia_min"}, inplace=True)
                
                # 2. Agrupar por (Equipamento, Data, Operação) - usando apenas improdutivos
                # Incluímos grupo_col apenas para referência
                df_grp = df_improd_only.groupby([col_equip_desc, col_data, op_col, grupo_col]).agg(
                    Duracao_Improd_min=("dur_improd", "sum")
                ).reset_index()
                
                # 3. Merge com o Tempo Total do Dia
                df_grp = df_grp.merge(df_total_dia, on=[col_equip_desc, col_data], how="left")
                
                # 4. Calcular Porcentagem
                df_grp["Porcentagem_Improdutiva"] = np.where(
                    df_grp["Total_Horas_Dia_min"] > 0,
                    (df_grp["Duracao_Improd_min"] / df_grp["Total_Horas_Dia_min"]) * 100,
                    0
                )
                
                # 5. Ordenar por Duração Improdutiva (Desc) e pegar Top 5
                df_grp = df_grp.sort_values([col_equip_desc, col_data, "Duracao_Improd_min"], ascending=[True, True, False])
                df_top5_ofensores = df_grp.groupby([col_equip_desc, col_data]).head(5).reset_index(drop=True)
                
                # 6. Converter para Horas
                df_top5_ofensores["Duracao_Improd_h"] = df_top5_ofensores["Duracao_Improd_min"] / 60
                df_top5_ofensores["Total_Horas_Dia_h"] = df_top5_ofensores["Total_Horas_Dia_min"] / 60
                
                # 7. Selecionar e Renomear Colunas
                cols_top5 = [col_equip_desc, col_data, op_col, "Duracao_Improd_h", "Total_Horas_Dia_h", "Porcentagem_Improdutiva"]
                df_top5_ofensores = df_top5_ofensores[cols_top5]

        df_frota_intervalos = pd.DataFrame()
        req_cols_gantt = [col_data, col_equip, col_equip_desc, "dt_inicial", "dt_final", "Descrição do Grupo da Operação"]
        
        if all(c in df_calc.columns for c in [col_data, col_equip, col_equip_desc]):
            cols_to_use = [c for c in req_cols_gantt if c in df_calc.columns]
            if "Descrição da Operação" in df_calc.columns:
                cols_to_use.append("Descrição da Operação")
            
            df_frota_intervalos = df_calc[cols_to_use].copy()
            
            rename_map = {
                "dt_inicial": "Início",
                "dt_final": "Fim",
                "Descrição do Grupo da Operação": "Grupo",
                col_equip: "Frota"
            }
            df_frota_intervalos.rename(columns=rename_map, inplace=True)
            
            if "Grupo" in df_frota_intervalos.columns:
                s = df_frota_intervalos["Grupo"].astype(str).str.upper().str.strip()
                
                cond_prod = s.str.contains("PRODUTIVA", na=False) & ~s.str.contains("IMPRODUTIVA", na=False)
                cond_manut = s.str.contains("MANUTEN", na=False)
                
                df_frota_intervalos["Grupo"] = np.where(cond_prod, "PRODUTIVA",
                                               np.where(cond_manut, "MANUTENCAO", "DISPONIVEL"))
            
            sort_cols = ["Frota"]
            if "Início" in df_frota_intervalos.columns:
                sort_cols.append("Início")
            
            df_frota_intervalos.sort_values(sort_cols, inplace=True)

        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            nome_aba_original = "Original"
            df_original.to_excel(writer, sheet_name=nome_aba_original, index=False)
            ajustar_largura_colunas(writer.sheets[nome_aba_original])

            nome_aba_tratado = "Tratado"
            df_tratado.to_excel(writer, sheet_name=nome_aba_tratado, index=False)
            ajustar_largura_colunas(writer.sheets[nome_aba_tratado])

            if not df_dia_frota.empty:
                tipo_col = col_equip_desc
                if tipo_col in df_dia_frota.columns:
                    tipos = [t for t in df_dia_frota[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_dia_frota[df_dia_frota[tipo_col] == tipo].copy()
                        
                        # Remover colunas que estão totalmente vazias ou zeradas para este tipo
                        cols_validas = [c for c in df_tipo.columns if not (pd.api.types.is_numeric_dtype(df_tipo[c]) and (df_tipo[c].sum() == 0))]
                        df_tipo = df_tipo[cols_validas]

                        # Remover Vel_Colheita_media para TRANSBORDO (antigo TRATOR TRANSBORDO)
                        if "TRANSBORDO" in str(tipo).upper():
                             if "Vel_Colheita_media" in df_tipo.columns:
                                 df_tipo = df_tipo.drop(columns=["Vel_Colheita_media"])
                        
                        # Remove a coluna de descrição do equipamento pois já está separada por aba
                        if tipo_col in df_tipo.columns:
                            df_tipo = df_tipo.drop(columns=[tipo_col])

                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        sufixo = "_Dia"
                        max_len = 31 - len(sufixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{safe_tipo}{sufixo}"
                        
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ajustar_largura_colunas(writer.sheets[nome_aba])
                        formatar_coluna_data(writer.sheets[nome_aba])
                else:
                    df_dia_frota.to_excel(writer, sheet_name="Equipamentos_Dia", index=False)
                    ajustar_largura_colunas(writer.sheets["Equipamentos_Dia"])
                    formatar_coluna_data(writer.sheets["Equipamentos_Dia"])

            if not df_dia_operador.empty:
                tipo_col = col_equip_desc
                if tipo_col in df_dia_operador.columns:
                    tipos = [t for t in df_dia_operador[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_dia_operador[df_dia_operador[tipo_col] == tipo].copy()
                        
                        cols_validas = [c for c in df_tipo.columns if not (pd.api.types.is_numeric_dtype(df_tipo[c]) and (df_tipo[c].sum() == 0))]
                        df_tipo = df_tipo[cols_validas]

                        if "TRANSBORDO" in str(tipo).upper():
                             if "Vel_Colheita_media" in df_tipo.columns:
                                 df_tipo = df_tipo.drop(columns=["Vel_Colheita_media"])

                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        prefixo = "Operadores_"
                        max_len = 31 - len(prefixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{prefixo}{safe_tipo}"
                        
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ajustar_largura_colunas(writer.sheets[nome_aba])
                        formatar_coluna_data(writer.sheets[nome_aba])
                else:
                    df_dia_operador.to_excel(writer, sheet_name="Operadores", index=False)
                    ajustar_largura_colunas(writer.sheets["Operadores"])
                    formatar_coluna_data(writer.sheets["Operadores"])

            if not df_periodo_frota.empty:
                df_periodo_frota.to_excel(writer, sheet_name="Periodo_Equipamentos", index=False)
                ajustar_largura_colunas(writer.sheets["Periodo_Equipamentos"])

            if not df_periodo_operador.empty:
                df_periodo_operador.to_excel(writer, sheet_name="Periodo_Operadores", index=False)
                ajustar_largura_colunas(writer.sheets["Periodo_Operadores"])

            if not df_top5_ofensores.empty:
                tipo_col = col_equip_desc
                if tipo_col in df_top5_ofensores.columns:
                    tipos = [t for t in df_top5_ofensores[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_top5_ofensores[df_top5_ofensores[tipo_col] == tipo].copy()
                        
                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        prefixo = "Top5Ofensores_"
                        max_len = 31 - len(prefixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{prefixo}{safe_tipo}"
                        
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ajustar_largura_colunas(writer.sheets[nome_aba])
                        formatar_coluna_data(writer.sheets[nome_aba])
                else:
                    df_top5_ofensores.to_excel(writer, sheet_name="Top5_Ofensores", index=False)
                    ajustar_largura_colunas(writer.sheets["Top5_Ofensores"])
                    formatar_coluna_data(writer.sheets["Top5_Ofensores"])

            if not df_frota_intervalos.empty:
                tipo_col = col_equip_desc
                if tipo_col in df_frota_intervalos.columns:
                    tipos = [t for t in df_frota_intervalos[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_frota_intervalos[df_frota_intervalos[tipo_col] == tipo].copy()
                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        prefixo = "Intervalos_"
                        max_len = 31 - len(prefixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{prefixo}{safe_tipo}"
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ws2 = writer.sheets[nome_aba]
                        ajustar_largura_colunas(ws2)

                        col_data_idx = None
                        for cell in ws2[1]:
                            if cell.value == col_data:
                                col_data_idx = cell.column
                                break
                        if col_data_idx is not None:
                            for row in range(2, ws2.max_row + 1):
                                cell = ws2.cell(row=row, column=col_data_idx)
                                if cell.value:
                                    cell.number_format = "dd/mm/yyyy"
                        
                        for col_name in ["Início", "Fim"]:
                            col_idx = None
                            for cell in ws2[1]:
                                if cell.value == col_name:
                                    col_idx = cell.column
                                    break
                            if col_idx is not None:
                                for row in range(2, ws2.max_row + 1):
                                    cell = ws2.cell(row=row, column=col_idx)
                                    if cell.value:
                                        cell.number_format = "dd/mm/yyyy hh:mm:ss"
                else:
                    df_frota_intervalos.to_excel(writer, sheet_name="Intervalos_Geral", index=False)
                    ajustar_largura_colunas(writer.sheets["Intervalos_Geral"])

            if "Plan1" in writer.book.sheetnames:
                del writer.book["Plan1"]

            for sheet_name in writer.book.sheetnames:
                ajustar_largura_colunas(writer.book[sheet_name])

        print(f"  Processamento finalizado para {os.path.basename(caminho_arquivo)}")
        
    except Exception as e:
        print(f"  ERRO ao processar arquivo: {e}")

def main():
    print("=== INICIANDO TRATAMENTO DE DADOS ===")
    print(f"Diretório alvo: {DIRETORIO_ENTRADA}")
    
    if not validar_diretorio(DIRETORIO_ENTRADA):
        sys.exit(1)
        
    arquivos = obter_arquivos_xlsx(DIRETORIO_ENTRADA)

    if not arquivos:
        arquivos_zip = obter_arquivos_zip(DIRETORIO_ENTRADA)
        if arquivos_zip:
            extrair_zips(DIRETORIO_ENTRADA, arquivos_zip)
            arquivos = obter_arquivos_xlsx(DIRETORIO_ENTRADA)

    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado na pasta dados.")
        sys.exit(0)
        
    print(f"Encontrados {len(arquivos)} arquivos para processar.")
    
    for arquivo in arquivos:
        try:
            dir_name = os.path.dirname(arquivo)
            base = os.path.splitext(os.path.basename(arquivo))[0]
            ext = os.path.splitext(arquivo)[1]
            novo_nome = f"{base}_tratado{ext}"
            novo_caminho = os.path.join(dir_name, novo_nome)
            if os.path.exists(novo_caminho):
                os.remove(novo_caminho)
            shutil.copy2(arquivo, novo_caminho)
            print(f"Gerada cópia para tratamento: {os.path.basename(novo_caminho)}")
            tratar_arquivo(novo_caminho)
        except Exception as e:
            print(f"ERRO ao copiar/tratar arquivo {os.path.basename(arquivo)}: {e}")
        
    print("\n=== TRATAMENTO CONCLUÍDO ===")

if __name__ == "__main__":
    main()
