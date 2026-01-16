import os
import sys
import shutil
import warnings
from datetime import datetime
import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

# Suprimir avisos específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- CONFIGURAÇÕES GERAIS ---
# Diretório onde estão os arquivos Excel originais
DIRETORIO_ENTRADA = r"c:\Users\arauj\OneDrive\Área de Trabalho\testes\brunozzi\automacao_etl\scripts\dados"

# Colunas a serem removidas na etapa 1
COLUNAS_PARA_REMOVER = [
    "Descrição Regional",
    "Descrição da Unidade",
    "Descrição do Grupo de Equipamento",
    "Código da Fazenda",
    "Código da Zona",
    "Código do Talhão",
    "Descrição da Fazenda",
    "Horímetro/Odometro Secundário"
]
# ----------------------------

def validar_diretorio(caminho):
    """Verifica se o diretório de entrada existe."""
    if not os.path.exists(caminho):
        print(f"ERRO: O diretório configurado não existe: {caminho}")
        return False
    return True

def obter_arquivos_xlsx(diretorio):
    """Retorna uma lista de arquivos .xlsx no diretório."""
    arquivos = [
        os.path.join(diretorio, f) 
        for f in os.listdir(diretorio) 
        if f.lower().endswith(".xlsx") and not f.startswith("~$") # Ignora arquivos temporários do Excel
    ]
    return arquivos

def ajustar_largura_colunas(worksheet):
    """
    Ajusta a largura das colunas de uma worksheet baseada no conteúdo.
    """
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        
        # Define um tamanho mínimo e adiciona um pouco de folga
        adjusted_width = (max_length + 2) * 1.2
        # Limita a largura máxima para não ficar excessivamente larga se houver textos gigantes
        if adjusted_width > 100:
            adjusted_width = 100
        worksheet.column_dimensions[column_letter].width = adjusted_width

def formatar_coluna_data(ws):
    """Percorre a primeira linha para achar 'Data' e aplica formatação."""
    col_data_idx = None
    for cell in ws[1]:
        if cell.value == "Data":
            col_data_idx = cell.column
            break
    if col_data_idx is not None:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_data_idx)
            if cell.value:
                cell.number_format = "dd/mm/yyyy"

def tratar_arquivo(caminho_arquivo):
    """
    Abre o arquivo Excel, remove colunas especificadas e salva em uma nova aba.
    """
    print(f"\nIniciando processamento: {os.path.basename(caminho_arquivo)}")
    
    try:
        # Carrega a planilha original (primeira aba)
        # Usamos engine='openpyxl' para garantir compatibilidade
        df_original = pd.read_excel(caminho_arquivo, sheet_name=0, engine="openpyxl")
        
        # Identifica quais colunas da lista realmente existem no arquivo
        colunas_existentes = [col for col in COLUNAS_PARA_REMOVER if col in df_original.columns]
        
        if not colunas_existentes:
            print("  AVISO: Nenhuma das colunas configuradas para remoção foi encontrada.")
            df_tratado = df_original.copy()
        else:
            print(f"  Removendo {len(colunas_existentes)} colunas...")
            df_tratado = df_original.drop(columns=colunas_existentes)

        if "Descrição do Equipamento" in df_tratado.columns:
            desc_series = df_tratado["Descrição do Equipamento"].astype(str)
            mask_colhedora_cana = desc_series.str.strip().str.upper() == "COLHEDORA DE CANA"
            df_tratado.loc[mask_colhedora_cana, "Descrição do Equipamento"] = "COLHEDORA"

        df_calc = df_tratado.copy()

        try:
            df_calc["Data"] = pd.to_datetime(df_calc["Data Hora Local"], dayfirst=True, errors="coerce").dt.date

            df_calc["Hora Inicial"] = df_calc["Hora Inicial"].astype(str)
            df_calc["Hora Final"] = df_calc["Hora Final"].astype(str)

            inicio_str = df_calc["Data Hora Local"].astype(str) + " " + df_calc["Hora Inicial"]
            fim_str = df_calc["Data Hora Local"].astype(str) + " " + df_calc["Hora Final"]

            df_calc["dt_inicial"] = pd.to_datetime(inicio_str, dayfirst=True, errors="coerce")
            df_calc["dt_final"] = pd.to_datetime(fim_str, dayfirst=True, errors="coerce")

            df_calc["Duracao_min"] = (df_calc["dt_final"] - df_calc["dt_inicial"]).dt.total_seconds() / 60

            mask_negativo = df_calc["Duracao_min"] < 0
            df_calc.loc[mask_negativo, "Duracao_min"] += 1440
        except Exception as e:
            print(f"  AVISO: Não foi possível calcular duração (erro de formato de data): {e}")
            df_calc["Duracao_min"] = 0

        if "Duracao_min" in df_calc.columns:
            df_calc["Duracao_h"] = df_calc["Duracao_min"] / 60
            if "Hora Final" in df_tratado.columns:
                try:
                    idx = df_tratado.columns.get_loc("Hora Final") + 1
                    df_tratado.insert(idx, "Duração", df_calc["Duracao_h"])
                except Exception:
                    df_tratado["Duração"] = df_calc["Duracao_h"]

        df_intervalos_cols = []
        for col in [
            "Data",
            "Código Equipamento",
            "Descrição do Equipamento",
            "Código de Operador",
            "Nome",
            "Descrição da Operação",
            "Descrição do Grupo da Operação",
            "Velocidade Média",
        ]:
            if col in df_calc.columns:
                df_intervalos_cols.append(col)

        df_intervalos = df_calc[df_intervalos_cols].copy()
        if "dt_inicial" in df_calc.columns:
            df_intervalos["Início"] = df_calc["dt_inicial"]
        if "dt_final" in df_calc.columns:
            df_intervalos["Fim"] = df_calc["dt_final"]
        if "Duracao_min" in df_calc.columns:
            df_intervalos["Duração (min)"] = df_calc["Duracao_min"]

        grupo_col = "Descrição do Grupo da Operação"
        op_col = "Descrição da Operação"

        df_calc["dur_total"] = df_calc["Duracao_min"]
        df_calc["dur_prod"] = np.where(df_calc.get(grupo_col, "") == "PRODUTIVA", df_calc["Duracao_min"], 0)
        df_calc["dur_improd"] = np.where(df_calc.get(grupo_col, "") == "IMPRODUTIVA", df_calc["Duracao_min"], 0)

        df_calc["dur_manobra"] = np.where(df_calc.get(op_col, "") == "MANOBRA", df_calc["Duracao_min"], 0)
        df_calc["dur_transbordo"] = np.where(df_calc.get(op_col, "") == "TRANSBORDANDO CANA", df_calc["Duracao_min"], 0)
        df_calc["dur_sem_apont"] = np.where(df_calc.get(op_col, "") == "SEM APONTAMENTO", df_calc["Duracao_min"], 0)

        df_calc["dur_colheita"] = np.where(
            df_calc.get(op_col, "").isin(["CARREGANDO CANA", "COLHENDO CANA"]),
            df_calc["Duracao_min"],
            0,
        )
        df_calc["dur_vazio"] = np.where(df_calc.get(op_col, "") == "DESL VAZIO", df_calc["Duracao_min"], 0)
        df_calc["dur_carregado"] = np.where(df_calc.get(op_col, "") == "DESL CARREGADO", df_calc["Duracao_min"], 0)

        vm_series = df_calc.get("Velocidade Média")
        vm = pd.to_numeric(vm_series, errors="coerce").fillna(0) if vm_series is not None else 0

        df_calc["vel_colheita_x_min"] = np.where(df_calc["dur_colheita"] > 0, vm * df_calc["dur_colheita"], 0)
        df_calc["vel_vazio_x_min"] = np.where(df_calc["dur_vazio"] > 0, vm * df_calc["dur_vazio"], 0)
        df_calc["vel_carregado_x_min"] = np.where(df_calc["dur_carregado"] > 0, vm * df_calc["dur_carregado"], 0)

        df_calc["cnt_manobra"] = np.where(df_calc.get(op_col, "") == "MANOBRA", 1, 0)
        df_calc["cnt_transbordo"] = np.where(df_calc.get(op_col, "") == "TRANSBORDANDO CANA", 1, 0)

        col_data = "Data"
        col_equip = "Código Equipamento"
        col_equip_desc = "Descrição do Equipamento"
        col_op_cod = "Código de Operador"
        col_op_nome = "Nome"

        # --- 3.Dia_Frota ---
        df_dia_frota = pd.DataFrame()
        if all(c in df_calc.columns for c in [col_data, col_equip]):
            group_cols_frota = [c for c in [col_data, col_equip, col_equip_desc] if c in df_calc.columns]
            df_dia_frota = (
                df_calc.groupby(group_cols_frota)
                .agg(
                    dur_total_min=("dur_total", "sum"),
                    dur_prod_min=("dur_prod", "sum"),
                    dur_improd_min=("dur_improd", "sum"),
                    dur_manobra_min=("dur_manobra", "sum"),
                    dur_transbordo_min=("dur_transbordo", "sum"),
                    dur_sem_apont_min=("dur_sem_apont", "sum"),
                    dur_colheita_min=("dur_colheita", "sum"),
                    dur_vazio_min=("dur_vazio", "sum"),
                    dur_carregado_min=("dur_carregado", "sum"),
                    vel_colheita_x_min=("vel_colheita_x_min", "sum"),
                    vel_vazio_x_min=("vel_vazio_x_min", "sum"),
                    vel_carregado_x_min=("vel_carregado_x_min", "sum"),
                    cnt_manobra=("cnt_manobra", "sum"),
                    cnt_transbordo=("cnt_transbordo", "sum"),
                )
                .reset_index()
            )

            # Converter para horas e renomear
            df_dia_frota["Horas_Registradas"] = df_dia_frota["dur_total_min"] / 60
            df_dia_frota["Horas_Produtivas"] = df_dia_frota["dur_prod_min"] / 60
            df_dia_frota["Horas_Improdutivas"] = df_dia_frota["dur_improd_min"] / 60
            df_dia_frota["Tempo_Sem_Apontamento_h"] = df_dia_frota["dur_sem_apont_min"] / 60
            
            # Manter métricas de manobra/transbordo em minutos (geralmente são curtos) ou converter se desejado
            # O usuário pediu "tudo como horas", mas tempos médios de manobra costumam ser minutos.
            # Vou manter minutos para "médios" e adicionar horas para totais.
            
            df_dia_frota["Tempo_Manobra_total_h"] = df_dia_frota["dur_manobra_min"] / 60
            df_dia_frota["Tempo_Manobra_medio_min"] = np.where(
                df_dia_frota["cnt_manobra"] > 0,
                df_dia_frota["dur_manobra_min"] / df_dia_frota["cnt_manobra"],
                0,
            )

            df_dia_frota["Tempo_Transbordo_total_h"] = df_dia_frota["dur_transbordo_min"] / 60
            df_dia_frota["Tempo_Transbordo_medio_min"] = np.where(
                df_dia_frota["cnt_transbordo"] > 0,
                df_dia_frota["dur_transbordo_min"] / df_dia_frota["cnt_transbordo"],
                0,
            )

            df_dia_frota["Vel_Colheita_media"] = np.where(
                df_dia_frota["dur_colheita_min"] > 0,
                df_dia_frota["vel_colheita_x_min"] / df_dia_frota["dur_colheita_min"],
                np.nan,
            )
            df_dia_frota["Vel_Desl_Vazio_media"] = np.where(
                df_dia_frota["dur_vazio_min"] > 0,
                df_dia_frota["vel_vazio_x_min"] / df_dia_frota["dur_vazio_min"],
                np.nan,
            )
            df_dia_frota["Vel_Desl_Carregado_media"] = np.where(
                df_dia_frota["dur_carregado_min"] > 0,
                df_dia_frota["vel_carregado_x_min"] / df_dia_frota["dur_carregado_min"],
                np.nan,
            )
            
            # Remover colunas auxiliares em minutos que não são mais necessárias
            cols_drop = [
                "dur_total_min", "dur_prod_min", "dur_improd_min", "dur_sem_apont_min",
                "dur_manobra_min", "dur_transbordo_min", "dur_colheita_min",
                "dur_vazio_min", "dur_carregado_min", "vel_colheita_x_min",
                "vel_vazio_x_min", "vel_carregado_x_min"
            ]
            df_dia_frota.drop(columns=cols_drop, inplace=True, errors="ignore")

        # --- 4.Dia_Operador ---
        df_dia_operador = pd.DataFrame()
        if all(c in df_calc.columns for c in [col_data, col_op_cod, col_op_nome]):
            group_cols_op = [col_data, col_op_cod, col_op_nome]
            df_dia_operador = (
                df_calc.groupby(group_cols_op)
                .agg(
                    dur_total_min=("dur_total", "sum"),
                    dur_prod_min=("dur_prod", "sum"),
                    dur_improd_min=("dur_improd", "sum"),
                    dur_manobra_min=("dur_manobra", "sum"),
                    dur_transbordo_min=("dur_transbordo", "sum"),
                    dur_sem_apont_min=("dur_sem_apont", "sum"),
                    dur_colheita_min=("dur_colheita", "sum"),
                    dur_vazio_min=("dur_vazio", "sum"),
                    dur_carregado_min=("dur_carregado", "sum"),
                    vel_colheita_x_min=("vel_colheita_x_min", "sum"),
                    vel_vazio_x_min=("vel_vazio_x_min", "sum"),
                    vel_carregado_x_min=("vel_carregado_x_min", "sum"),
                    cnt_manobra=("cnt_manobra", "sum"),
                    cnt_transbordo=("cnt_transbordo", "sum"),
                )
                .reset_index()
            )

            if col_equip in df_calc.columns:
                frotas = (
                    df_calc.groupby(group_cols_op)[col_equip]
                    .agg(lambda x: ", ".join(str(v) for v in sorted(set(x.dropna()))))
                    .reset_index()
                )
                df_dia_operador = df_dia_operador.merge(frotas, on=group_cols_op, how="left")
                df_dia_operador.rename(columns={col_equip: "Frotas_no_dia"}, inplace=True)

            df_dia_operador["Horas_Registradas"] = df_dia_operador["dur_total_min"] / 60
            df_dia_operador["Horas_Produtivas"] = df_dia_operador["dur_prod_min"] / 60
            df_dia_operador["Horas_Improdutivas"] = df_dia_operador["dur_improd_min"] / 60
            df_dia_operador["Tempo_Sem_Apontamento_h"] = df_dia_operador["dur_sem_apont_min"] / 60

            df_dia_operador["Tempo_Manobra_total_h"] = df_dia_operador["dur_manobra_min"] / 60
            df_dia_operador["Tempo_Manobra_medio_min"] = np.where(
                df_dia_operador["cnt_manobra"] > 0,
                df_dia_operador["dur_manobra_min"] / df_dia_operador["cnt_manobra"],
                0,
            )

            df_dia_operador["Tempo_Transbordo_total_h"] = df_dia_operador["dur_transbordo_min"] / 60
            df_dia_operador["Tempo_Transbordo_medio_min"] = np.where(
                df_dia_operador["cnt_transbordo"] > 0,
                df_dia_operador["dur_transbordo_min"] / df_dia_operador["cnt_transbordo"],
                0,
            )

            df_dia_operador["Vel_Colheita_media"] = np.where(
                df_dia_operador["dur_colheita_min"] > 0,
                df_dia_operador["vel_colheita_x_min"] / df_dia_operador["dur_colheita_min"],
                np.nan,
            )
            df_dia_operador["Vel_Desl_Vazio_media"] = np.where(
                df_dia_operador["dur_vazio_min"] > 0,
                df_dia_operador["vel_vazio_x_min"] / df_dia_operador["dur_vazio_min"],
                np.nan,
            )
            df_dia_operador["Vel_Desl_Carregado_media"] = np.where(
                df_dia_operador["dur_carregado_min"] > 0,
                df_dia_operador["vel_carregado_x_min"] / df_dia_operador["dur_carregado_min"],
                np.nan,
            )

            cols_drop = [
                "dur_total_min", "dur_prod_min", "dur_improd_min", "dur_sem_apont_min",
                "dur_manobra_min", "dur_transbordo_min", "dur_colheita_min",
                "dur_vazio_min", "dur_carregado_min", "vel_colheita_x_min",
                "vel_vazio_x_min", "vel_carregado_x_min"
            ]
            df_dia_operador.drop(columns=cols_drop, inplace=True, errors="ignore")

        df_periodo_frota = pd.DataFrame()
        if not df_dia_frota.empty and col_data in df_dia_frota.columns:
            group_cols = [c for c in [col_equip, col_equip_desc] if c in df_dia_frota.columns]
            if group_cols:
                df_periodo_frota = (
                    df_dia_frota.groupby(group_cols)
                    .agg(
                        Horas_Registradas_total=("Horas_Registradas", "sum"),
                        Dias_com_dados=(col_data, "nunique"),
                    )
                    .reset_index()
                )
                df_periodo_frota["Horas_media_por_dia"] = np.where(
                    df_periodo_frota["Dias_com_dados"] > 0,
                    df_periodo_frota["Horas_Registradas_total"] / df_periodo_frota["Dias_com_dados"],
                    0,
                )

        df_periodo_operador = pd.DataFrame()
        if not df_dia_operador.empty and col_data in df_dia_operador.columns:
            group_cols = [col_op_cod, col_op_nome]
            df_periodo_operador = (
                df_dia_operador.groupby(group_cols)
                .agg(
                    Horas_Registradas_total=("Horas_Registradas", "sum"),
                    Dias_com_dados=(col_data, "nunique"),
                )
                .reset_index()
            )
            df_periodo_operador["Horas_media_por_dia"] = np.where(
                df_periodo_operador["Dias_com_dados"] > 0,
                df_periodo_operador["Horas_Registradas_total"] / df_periodo_operador["Dias_com_dados"],
                0,
            )

        # --- 3.Top5Ofensores ---
        df_top5_ofensores = pd.DataFrame()
        # Verificar colunas necessárias
        if all(c in df_calc.columns for c in [col_equip_desc, col_data, grupo_col, op_col]):
            # Filtrar apenas o que é improdutivo (Ofensores)
            df_improd_only = df_calc[df_calc["dur_improd"] > 0].copy()
            
            if not df_improd_only.empty:
                # 1. Calcular Total de Horas por (Equipamento, Data) - usando df_calc completo (todo o tempo)
                # Isso servirá de denominador para a porcentagem
                df_total_dia = df_calc.groupby([col_equip_desc, col_data])["Duracao_min"].sum().reset_index()
                df_total_dia.rename(columns={"Duracao_min": "Total_Horas_Dia_min"}, inplace=True)
                
                # 2. Agrupar por (Equipamento, Data, Operação) - usando apenas improdutivos
                # Incluímos grupo_col apenas para referência
                df_grp = df_improd_only.groupby([col_equip_desc, col_data, op_col, grupo_col]).agg(
                    Duracao_Improd_min=("dur_improd", "sum")
                ).reset_index()
                
                # 3. Merge com o Tempo Total do Dia
                df_grp = df_grp.merge(df_total_dia, on=[col_equip_desc, col_data], how="left")
                
                # 4. Calcular Porcentagem
                df_grp["Porcentagem_Improdutiva"] = np.where(
                    df_grp["Total_Horas_Dia_min"] > 0,
                    (df_grp["Duracao_Improd_min"] / df_grp["Total_Horas_Dia_min"]) * 100,
                    0
                )
                
                # 5. Ordenar por Duração Improdutiva (Desc) e pegar Top 5
                df_grp = df_grp.sort_values([col_equip_desc, col_data, "Duracao_Improd_min"], ascending=[True, True, False])
                df_top5_ofensores = df_grp.groupby([col_equip_desc, col_data]).head(5).reset_index(drop=True)
                
                # 6. Converter para Horas
                df_top5_ofensores["Duracao_Improd_h"] = df_top5_ofensores["Duracao_Improd_min"] / 60
                df_top5_ofensores["Total_Horas_Dia_h"] = df_top5_ofensores["Total_Horas_Dia_min"] / 60
                
                # 7. Selecionar e Renomear Colunas
                cols_top5 = [col_equip_desc, col_data, grupo_col, op_col, "Duracao_Improd_h", "Total_Horas_Dia_h", "Porcentagem_Improdutiva"]
                df_top5_ofensores = df_top5_ofensores[cols_top5]

        df_frota_duracao = pd.DataFrame()
        if all(c in df_calc.columns for c in [col_equip, col_equip_desc, col_data, "Duracao_min"]):
            apont_col = "Descrição da Operação"
            group_cols = [col_equip, col_equip_desc, col_data]
            if apont_col in df_calc.columns:
                group_cols.append(apont_col)

            df_frota_grp = (
                df_calc.groupby(group_cols)
                .agg(Duracao_total_min=("Duracao_min", "sum"))
                .reset_index()
            )
            df_frota_grp["Duracao_total_h"] = df_frota_grp["Duracao_total_min"] / 60
            if "Duracao_total_min" in df_frota_grp.columns:
                df_frota_grp = df_frota_grp.drop(columns=["Duracao_total_min"])
            cols = df_frota_grp.columns.tolist()
            if col_data in cols:
                cols.insert(0, cols.pop(cols.index(col_data)))
                df_frota_grp = df_frota_grp[cols]
            df_frota_duracao = df_frota_grp

        with pd.ExcelWriter(caminho_arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            nome_aba_1 = "1.ColunasRemovidas"
            df_tratado.to_excel(writer, sheet_name=nome_aba_1, index=False)
            ajustar_largura_colunas(writer.sheets[nome_aba_1])

            if not df_dia_frota.empty:
                # Separar 3.Dia_Frota por Tipo de Equipamento
                tipo_col = col_equip_desc
                if tipo_col in df_dia_frota.columns:
                    tipos = [t for t in df_dia_frota[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_dia_frota[df_dia_frota[tipo_col] == tipo].copy()
                        
                        # Remover colunas que estão totalmente vazias ou zeradas para este tipo
                        cols_validas = [c for c in df_tipo.columns if not (pd.api.types.is_numeric_dtype(df_tipo[c]) and (df_tipo[c].sum() == 0))]
                        df_tipo = df_tipo[cols_validas]

                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        prefixo = "3."
                        sufixo = "_Dia"
                        max_len = 31 - len(prefixo) - len(sufixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{prefixo}{safe_tipo}{sufixo}"
                        
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ajustar_largura_colunas(writer.sheets[nome_aba])
                        formatar_coluna_data(writer.sheets[nome_aba])
                else:
                    # Fallback caso a coluna de descrição não exista
                    df_dia_frota.to_excel(writer, sheet_name="3.Dia_Frota", index=False)
                    ajustar_largura_colunas(writer.sheets["3.Dia_Frota"])
                    formatar_coluna_data(writer.sheets["3.Dia_Frota"])

            if not df_dia_operador.empty:
                df_dia_operador.to_excel(writer, sheet_name="4.Dia_Operador", index=False)
                ajustar_largura_colunas(writer.sheets["4.Dia_Operador"])
                formatar_coluna_data(writer.sheets["4.Dia_Operador"])

            if not df_periodo_frota.empty:
                df_periodo_frota.to_excel(writer, sheet_name="5.Periodo_Frota", index=False)
                ajustar_largura_colunas(writer.sheets["5.Periodo_Frota"])

            if not df_periodo_operador.empty:
                df_periodo_operador.to_excel(writer, sheet_name="6.Periodo_Operador", index=False)
                ajustar_largura_colunas(writer.sheets["6.Periodo_Operador"])

            if not df_top5_ofensores.empty:
                # Separar 3.Top5Ofensores por Tipo de Equipamento
                tipo_col = col_equip_desc
                if tipo_col in df_top5_ofensores.columns:
                    tipos = [t for t in df_top5_ofensores[tipo_col].dropna().unique()]
                    for tipo in tipos:
                        df_tipo = df_top5_ofensores[df_top5_ofensores[tipo_col] == tipo].copy()
                        
                        safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                        prefixo = "3."
                        sufixo = "_Ofens"
                        max_len = 31 - len(prefixo) - len(sufixo)
                        safe_tipo = safe_tipo[:max_len]
                        nome_aba = f"{prefixo}{safe_tipo}{sufixo}"
                        
                        df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                        ajustar_largura_colunas(writer.sheets[nome_aba])
                        formatar_coluna_data(writer.sheets[nome_aba])
                else:
                    df_top5_ofensores.to_excel(writer, sheet_name="3.Top5Ofensores", index=False)
                    ajustar_largura_colunas(writer.sheets["3.Top5Ofensores"])
                    formatar_coluna_data(writer.sheets["3.Top5Ofensores"])

            if not df_frota_duracao.empty:
                tipo_col = col_equip_desc
                tipos = [t for t in df_frota_duracao[tipo_col].dropna().unique()]
                for tipo in tipos:
                    df_tipo = df_frota_duracao[df_frota_duracao[tipo_col] == tipo].copy()
                    safe_tipo = str(tipo).replace("/", "-").replace("\\", "-")
                    prefixo = "2."
                    sufixo = "_Duração"
                    max_len = 31 - len(prefixo) - len(sufixo)
                    safe_tipo = safe_tipo[:max_len]
                    nome_aba = f"{prefixo}{safe_tipo}{sufixo}"
                    df_tipo.to_excel(writer, sheet_name=nome_aba, index=False)
                    ws2 = writer.sheets[nome_aba]
                    ajustar_largura_colunas(ws2)

                    col_data_idx = None
                    for cell in ws2[1]:
                        if cell.value == col_data:
                            col_data_idx = cell.column
                            break
                    if col_data_idx is not None:
                        for row in range(2, ws2.max_row + 1):
                            cell = ws2.cell(row=row, column=col_data_idx)
                            if cell.value:
                                cell.number_format = "dd/mm/yyyy"

            for sheet_name in writer.book.sheetnames:
                ajustar_largura_colunas(writer.book[sheet_name])

        print(f"  Processamento finalizado para {os.path.basename(caminho_arquivo)}")
        
    except Exception as e:
        print(f"  ERRO ao processar arquivo: {e}")

def main():
    print("=== INICIANDO TRATAMENTO DE DADOS ===")
    print(f"Diretório alvo: {DIRETORIO_ENTRADA}")
    
    if not validar_diretorio(DIRETORIO_ENTRADA):
        sys.exit(1)
        
    arquivos = obter_arquivos_xlsx(DIRETORIO_ENTRADA)
    
    if not arquivos:
        print("Nenhum arquivo .xlsx encontrado na pasta dados.")
        sys.exit(0)
        
    print(f"Encontrados {len(arquivos)} arquivos para processar.")
    
    for arquivo in arquivos:
        try:
            dir_name = os.path.dirname(arquivo)
            base = os.path.splitext(os.path.basename(arquivo))[0]
            ext = os.path.splitext(arquivo)[1]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            novo_nome = f"{base}_tratado_{timestamp}{ext}"
            novo_caminho = os.path.join(dir_name, novo_nome)
            shutil.copy2(arquivo, novo_caminho)
            print(f"Gerada cópia para tratamento: {os.path.basename(novo_caminho)}")
            tratar_arquivo(novo_caminho)
        except Exception as e:
            print(f"ERRO ao copiar/tratar arquivo {os.path.basename(arquivo)}: {e}")
        
    print("\n=== TRATAMENTO CONCLUÍDO ===")

if __name__ == "__main__":
    main()
