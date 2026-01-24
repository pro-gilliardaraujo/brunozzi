import os
import sys
import pandas as pd
import glob
import warnings
import re
import json
from datetime import datetime
from openpyxl.utils import get_column_letter

# Suprimir avisos específicos do openpyxl
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# --- CONFIGURAÇÕES ---
# Usa o diretório de trabalho atual como base para construir o caminho absoluto
BASE_CWD = os.getcwd()
DIRETORIO_DADOS = os.path.join(BASE_CWD, "automacao_etl", "scripts", "dados")
DIRETORIO_SAIDA = os.path.join(DIRETORIO_DADOS, "separados")
DIRETORIO_XLSX = os.path.join(DIRETORIO_SAIDA, "xlsx")
DIRETORIO_JSON = os.path.join(DIRETORIO_SAIDA, "json")

def extrair_periodo_nome_arquivo(nome_arquivo):
    """
    Extrai datas de início e fim do nome do arquivo.
    Padrão esperado: ...DD-MM-YYYY_DD-MM-YYYY...
    """
    match = re.search(r"(\d{2}-\d{2}-\d{4})_(\d{2}-\d{2}-\d{4})", nome_arquivo)
    if match:
        d1_str, d2_str = match.groups()
        try:
            d1 = pd.to_datetime(d1_str, format="%d-%m-%Y").date()
            d2 = pd.to_datetime(d2_str, format="%d-%m-%Y").date()
            return d1, d2
        except:
            pass
    return None, None

def ajustar_largura_colunas(worksheet):
    """Ajusta a largura das colunas automaticamente."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

def formatar_coluna_data(worksheet, df):
    """Aplica formatação de data DD/MM/YYYY nas colunas de data."""
    col_idx_map = {name: i+1 for i, name in enumerate(df.columns)}
    
    # Procura colunas que parecem ser data
    for col_name in df.columns:
        if "Data" in col_name or "Início" in col_name or "Fim" in col_name:
            col_idx = col_idx_map[col_name]
            col_letter = get_column_letter(col_idx)
            
            # Aplica formato na coluna inteira (exceto cabeçalho)
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row}"]
                cell.number_format = 'DD/MM/YYYY'

def encontrar_ultimo_tratado():
    """Encontra o arquivo tratado mais recente na pasta dados."""
    if not os.path.exists(DIRETORIO_DADOS):
        print(f"ERRO: Diretório de dados não encontrado: {DIRETORIO_DADOS}")
        return None

    # Tenta padrões diferentes para garantir que encontre o arquivo
    padroes = ["*tratado*.xlsx", "*_tratado.xlsx"]
    arquivos = []
    
    for p in padroes:
        caminho_padrao = os.path.join(DIRETORIO_DADOS, p)
        encontrados = glob.glob(caminho_padrao)
        arquivos.extend(encontrados)
    
    # Remove duplicatas
    arquivos = list(set(arquivos))
    
    if not arquivos:
        return None
    
    # Ordena por data de modificação (mais recente primeiro)
    arquivos.sort(key=os.path.getmtime, reverse=True)
    return arquivos[0]

def main():
    print("=== INICIANDO SEPARAÇÃO COMPLETA POR DIA ===")
    
    # 1. Encontrar o arquivo
    arquivo_input = encontrar_ultimo_tratado()
    
    if not arquivo_input:
        print(f"ERRO: Nenhum arquivo tratado encontrado em: {DIRETORIO_DADOS}")
        print("Execute o script 2_Tratamento.py primeiro.")
        sys.exit(1)
        
    print(f"Lendo arquivo base: {os.path.basename(arquivo_input)}")
    
    try:
        # 2. Ler todas as abas do arquivo Excel
        xl = pd.ExcelFile(arquivo_input, engine='openpyxl')
        sheet_names = xl.sheet_names
        print(f"Abas encontradas: {sheet_names}")
        
        # Carregar todas as abas em um dicionário de DataFrames
        dfs = {}
        for sheet in sheet_names:
            print(f"  Carregando aba: {sheet}")
            dfs[sheet] = pd.read_excel(arquivo_input, sheet_name=sheet, engine='openpyxl')
        
        # 3. Identificar dias únicos (usando a aba 'Tratado' ou 'Original' como referência)
        # Preferência por 'Tratado' pois já passou pelo filtro de datas do 2_Tratamento.py
        df_ref = None
        if 'Tratado' in dfs:
            df_ref = dfs['Tratado']
        elif 'Original' in dfs:
            df_ref = dfs['Original']
        else:
            # Se não tiver Tratado nem Original, pega a primeira
            df_ref = dfs[sheet_names[0]]
            
        # Encontrar coluna de data na referência
        col_data_ref = None
        for col in df_ref.columns:
            if str(col).lower().strip() == 'data':
                col_data_ref = col
                break
        
        # Se não achou 'Data', tenta criar a partir de 'Data Hora Local'
        if not col_data_ref:
            print("  Coluna 'Data' explícita não encontrada na referência. Tentando derivar de 'Data Hora Local'...")
            for col in df_ref.columns:
                if str(col).lower().strip() == 'data hora local':
                    # Criar coluna Data no df_ref
                    df_ref['Data'] = pd.to_datetime(df_ref[col], dayfirst=True, errors='coerce').dt.date
                    col_data_ref = 'Data'
                    break

        if not col_data_ref:
            # Tentar procurar em outras abas que sabemos que têm Data
            print("  'Data' não encontrada em Tratado/Original. Procurando em outras abas...")
            for sheet_name, df_temp in dfs.items():
                if "_Dia" in sheet_name or "Intervalos" in sheet_name:
                    for col in df_temp.columns:
                         if str(col).lower().strip() == 'data':
                            print(f"  Usando aba '{sheet_name}' como referência de datas.")
                            df_ref = df_temp
                            col_data_ref = col
                            break
                if col_data_ref:
                    break

        if not col_data_ref:
            print("ERRO: Coluna 'Data' não encontrada na aba de referência para identificar os dias.")
            sys.exit(1)
            
        # Converter para datetime e pegar dias únicos
        df_ref[col_data_ref] = pd.to_datetime(df_ref[col_data_ref], errors='coerce')
        datas_unicas = df_ref[col_data_ref].dropna().unique()
        
        if len(datas_unicas) == 0:
            print("ERRO: Nenhuma data válida encontrada na aba de referência.")
            sys.exit(1)

        # Filtrar datas com base no nome do arquivo (se disponível)
        dt_inicio_filtro, dt_fim_filtro = extrair_periodo_nome_arquivo(os.path.basename(arquivo_input))
        if dt_inicio_filtro and dt_fim_filtro:
            print(f"  Filtrando dias pelo período do arquivo: {dt_inicio_filtro} a {dt_fim_filtro}")
            datas_filtradas = []
            for d in datas_unicas:
                 # d é numpy.datetime64, converte para datetime.date
                 d_dt = pd.to_datetime(d).date()
                 if dt_inicio_filtro <= d_dt <= dt_fim_filtro:
                     datas_filtradas.append(d)
            
            if not datas_filtradas:
                 print(f"AVISO: Nenhuma data encontrada dentro do período {dt_inicio_filtro} a {dt_fim_filtro}.")
                 # Se filtrar tudo, talvez seja melhor não filtrar e avisar, ou sair. 
                 # O usuário quer remover dias fora. Se todos estão fora, gera vazio.
            
            datas_unicas = datas_filtradas
            
        print(f"Encontrados {len(datas_unicas)} dias únicos: {[pd.to_datetime(d).strftime('%d/%m') for d in datas_unicas]}")

        # 4. Criar diretórios de saída
        if not os.path.exists(DIRETORIO_SAIDA):
            os.makedirs(DIRETORIO_SAIDA)
        
        if not os.path.exists(DIRETORIO_XLSX):
            os.makedirs(DIRETORIO_XLSX)
            
        if not os.path.exists(DIRETORIO_JSON):
            os.makedirs(DIRETORIO_JSON)

        print(f"Diretórios de saída configurados em: {DIRETORIO_SAIDA}")
            
        # 5. Separar e Salvar
        arquivos_gerados = []
        
        for data_val in sorted(datas_unicas):
            ts = pd.to_datetime(data_val)
            data_str = ts.strftime("%d-%m-%Y")
            nome_arquivo = f"{data_str}.xlsx"
            caminho_saida = os.path.join(DIRETORIO_XLSX, nome_arquivo)
            
            print(f"\nGerando arquivos para: {data_str}")
            
            with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
                sheets_salvas = 0
                dados_dia_json = {}
                
                for sheet_name, df in dfs.items():
                    # Ignorar abas de Periodo (são agregadas do período total, não faz sentido em arquivo diário ou seria redundante/incorreto)
                    if "Periodo" in sheet_name:
                        # print(f"    (Ignorando aba {sheet_name} - agregação de período)")
                        continue
                        
                    # Tentar filtrar por data
                    col_data_sheet = None
                    for col in df.columns:
                        if str(col).lower().strip() == 'data':
                            col_data_sheet = col
                            break
                    
                    # Se não tem 'Data', tenta 'Data Hora Local'
                    temp_col_created = False
                    if not col_data_sheet:
                        for col in df.columns:
                            if str(col).lower().strip() == 'data hora local':
                                # Criar coluna temporária para filtro
                                df['_temp_data_filter_'] = pd.to_datetime(df[col], dayfirst=True, errors='coerce').dt.date
                                col_data_sheet = '_temp_data_filter_'
                                temp_col_created = True
                                break
                    
                    df_to_save = None
                    
                    if col_data_sheet:
                        # Se tem coluna Data, filtra
                        if not temp_col_created:
                            # Garantir tipo datetime se for a coluna original
                            df[col_data_sheet] = pd.to_datetime(df[col_data_sheet], errors='coerce')
                        
                        # Filtrar (se for temp, já é date object; se for original, é datetime)
                        # ts é datetime. Precisa comparar com date se a coluna for date, ou datetime se datetime.
                        # Melhor normalizar tudo para .date() na comparação
                        
                        if temp_col_created:
                             df_filtered = df[df[col_data_sheet] == ts.date()].copy()
                             # Remove a coluna temporária
                             df_filtered = df_filtered.drop(columns=['_temp_data_filter_'])
                        else:
                             # Comparação direta de datetime (pode falhar se tiver hora). 
                             # Vamos converter para date para garantir
                             df_temp_compare = df[col_data_sheet].dt.date
                             df_filtered = df[df_temp_compare == ts.date()].copy()
                        
                        if not df_filtered.empty:
                            df_to_save = df_filtered
                        else:
                            # Se não tem dados para esse dia nesta aba, pular
                            continue
                    else:
                        # Se não tem coluna Data
                        # Pode ser uma aba de configuração ou resumo sem data.
                        # Como o usuário pediu "todas as planilhas existentes", e removemos Periodo,
                        # se sobrar algo sem data, copiamos integralmente?
                        # Risco: duplicar dados de outros dias.
                        # Decisão: Por segurança, se não tem Data e não é Periodo, avisamos e não salvamos, 
                        # a menos que seja algo conhecido como "Operadores" (tem data), "Equipamentos" (tem data).
                        # Top5 tem data. Intervalos tem data.
                        # Se não tiver data, provavelmente não deve ir para o arquivo diário específico.
                        # print(f"    (Ignorando aba {sheet_name} - sem coluna Data)")
                        continue
                    
                    if df_to_save is not None:
                        # Salvar aba
                        df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
                        sheets_salvas += 1
                        
                        # Adicionar ao JSON (convertendo para dict serializável)
                        # Usar to_json e depois loads para garantir conversão correta de datas e NaNs (NaN vira null)
                        try:
                            json_str = df_to_save.to_json(orient='records', date_format='iso', default_handler=str)
                            dados_dia_json[sheet_name] = json.loads(json_str)
                        except Exception as e_json:
                            print(f"    AVISO: Falha ao serializar aba {sheet_name} para JSON: {e_json}")

                        # Formatação
                        worksheet = writer.sheets[sheet_name]
                        ajustar_largura_colunas(worksheet)
                        formatar_coluna_data(worksheet, df_to_save)
                
                if sheets_salvas > 0:
                    print(f"  -> Arquivo Excel salvo com {sheets_salvas} abas.")
                    arquivos_gerados.append(nome_arquivo)
                    
                    # Salvar arquivo JSON Completo
                    caminho_json = os.path.join(DIRETORIO_JSON, f"{data_str}.json")
                    try:
                        with open(caminho_json, 'w', encoding='utf-8') as f:
                            json.dump(dados_dia_json, f, ensure_ascii=False, indent=2)
                        print(f"  -> Arquivo JSON salvo em: json/{os.path.basename(caminho_json)}")
                    except Exception as e_save_json:
                        print(f"  -> ERRO ao salvar JSON: {e_save_json}")

                    # --- GERAR JSONs ESPECÍFICOS POR TIPO DE FROTA E ORGANIZAR EM PASTAS ---
                    
                    # Função auxiliar para garantir o nome do diretório normalizado
                    def obter_nome_diretorio(nome):
                        return normalizar_texto_simples(nome)

                    # Função para normalizar texto para nomes de pasta/arquivo
                    def normalizar_texto_simples(txt):
                        if not txt: return "outros"
                        # Remove acentos e caracteres especiais
                        txt = str(txt).lower().strip()
                        # Mapeamento simples para garantir consistência com o Tratamento
                        if txt == "colhedora de cana": return "colhedora"
                        if txt == "trator transbordo": return "transbordo"
                        return txt.replace(" ", "_").replace("/", "-")

                    # Iterar sobre os dados para gerar arquivos organizados
                    for key_aba, dados_lista in dados_dia_json.items():
                        
                        # 1. Abas de Resumo Diário (_Dia) -> Ex: COLHEDORA_Dia
                        if key_aba.endswith("_Dia"):
                            tipo_frota = key_aba.replace("_Dia", "").lower() # Ex: colhedora, transbordo
                            nome_pasta = obter_nome_diretorio(tipo_frota)
                            
                            # Criar pasta específica: json/colhedora/
                            dir_frota = os.path.join(DIRETORIO_JSON, nome_pasta)
                            if not os.path.exists(dir_frota):
                                os.makedirs(dir_frota)

                            # Definir colunas de interesse (Whitelist)
                            dados_filtrados = []
                            if isinstance(dados_lista, list):
                                for item in dados_lista:
                                    novo_item = {}
                                    # Chaves obrigatórias/principais
                                    if "Frota" in item: novo_item["Frota"] = item["Frota"]
                                    elif "Código Equipamento" in item: novo_item["Frota"] = item["Código Equipamento"]
                                    
                                    for k, v in item.items():
                                        k_lower = k.lower()
                                        # Inclui Horas_, Porcentagem_, Disponibilidade_, e Motor
                                        if (k.startswith("Horas_") or 
                                            k.startswith("Porcentagem_") or 
                                            k.startswith("Disponibilidade_") or
                                            k_lower == "motor ligado" or
                                            k_lower == "motor ocioso"):
                                            novo_item[k] = v
                                    
                                    if novo_item:
                                        dados_filtrados.append(novo_item)
                            
                            if dados_filtrados:
                                nome_arquivo = f"{nome_pasta}_{data_str}.json"
                                caminho_arquivo = os.path.join(dir_frota, nome_arquivo)
                                try:
                                    with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                                        json.dump(dados_filtrados, f, ensure_ascii=False, indent=2)
                                    print(f"  -> JSON salvo: json/{nome_pasta}/{nome_arquivo}")
                                except Exception as e_esp:
                                    print(f"  -> ERRO ao salvar {nome_arquivo}: {e_esp}")

                        # 2. Top5Ofensores e Intervalos (Separar por equipamento)
                        elif key_aba.startswith("Top5Ofensores") or key_aba.startswith("Intervalos"):
                            # Agrupar itens por 'Descrição do Equipamento'
                            itens_por_tipo = {}
                            if isinstance(dados_lista, list):
                                for item in dados_lista:
                                    # Obter tipo do equipamento
                                    tipo_raw = item.get("Descrição do Equipamento", "Outros")
                                    tipo_norm = obter_nome_diretorio(tipo_raw)
                                    
                                    if tipo_norm not in itens_por_tipo:
                                        itens_por_tipo[tipo_norm] = []
                                    itens_por_tipo[tipo_norm].append(item)
                            
                            # Salvar arquivos separados
                            # Define o prefixo baseado no nome da aba (Top5Ofensores ou Intervalos)
                            if "Top5Ofensores" in key_aba:
                                prefixo = "top5"
                            else:
                                prefixo = "intervalos"
                            
                            for tipo, itens in itens_por_tipo.items():
                                # Criar pasta se não existir (ex: json/colhedora/)
                                dir_frota = os.path.join(DIRETORIO_JSON, tipo)
                                if not os.path.exists(dir_frota):
                                    os.makedirs(dir_frota)
                                
                                nome_arquivo = f"{prefixo}_{tipo}_{data_str}.json"
                                caminho_arquivo = os.path.join(dir_frota, nome_arquivo)
                                
                                try:
                                    with open(caminho_arquivo, 'w', encoding='utf-8') as f:
                                        json.dump(itens, f, ensure_ascii=False, indent=2)
                                    print(f"  -> JSON salvo: json/{tipo}/{nome_arquivo}")
                                except Exception as e_esp:
                                    print(f"  -> ERRO ao salvar {nome_arquivo}: {e_esp}")
                    # --------------------------------------------------------------------------

                else:
                    print(f"  -> AVISO: Nenhuma aba gerada para {data_str}. Arquivo não salvo.")
            
        print(f"\nSucesso! Arquivos Excel e JSON gerados nas pastas 'separados/xlsx' e 'separados/json'.")
            
    except Exception as e:
        print(f"\nERRO FATAL: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
